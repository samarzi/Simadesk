#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Перед запуском:
# 1. Авторизуйтесь в Яндекс Маркете вручную в браузере
# 2. При первом запуске выберите режим "Сканер" для построения карты путей
# 3. При ошибке "Target closed" - убедитесь, что нет зависших Chrome процессов

import os
import re
import sys
import json
import time
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict, Set, Tuple

# ──────────────────────────────────────────────
# Зависимости
# ──────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("❌  pip install openpyxl")
    sys.exit(1)

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    print("❌  pip install playwright && playwright install chromium")
    sys.exit(1)


# ══════════════════════════════════════════════
#  КОНСТАНТЫ
# ══════════════════════════════════════════════
MAP_FILE       = "Карты/karta.json"
PROFILE_DIR    = os.environ.get("YD_PROFILE_DIR", "profiles/browser_profile")
LOG_FILE       = f"delivery_{os.environ.get('YD_INSTANCE_NAME', 'main')}.log"
PAUSE_FILE     = "pause.txt"
PROGRESS_FILE  = f"progress_{os.environ.get('YD_INSTANCE_NAME', 'main')}.json"
INPUT_VALUE    = "20"
EXCEL_DIR      = "Склады"       # папка с Excel-файлами городов
REPORTS_DIR    = "Отчёты"       # папка для Excel-отчётов анализа
MAPS_DIR       = "Карты"        # папка для карт karta.json

# Тайминги (мс)
AFTER_CLICK_MS    = 400     # пауза после одного клика раскрытия узла
AFTER_SPINNER_MS  = 300     # пауза после исчезновения спиннера
AFTER_SAVE_MS     = 5_000   # пауза после кнопки «Сохранить»
SCROLL_STEP_PX    = 500     # шаг прокрутки
SCROLL_PAUSE_MS   = 120     # пауза между шагами прокрутки

# ──────────────────────────────────────────────
#  СКЛАДЫ — кабинеты и URL страниц доставки
# ──────────────────────────────────────────────
WAREHOUSES = {
    "МХ": {
        "Нижний Новгород": "https://partner.market.yandex.ru/shop/149090347/delivery?warehouseId=2241450",
        "Москва":          "https://partner.market.yandex.ru/shop/149088763/delivery?warehouseId=2239087",
        "Смоленск":        "https://partner.market.yandex.ru/shop/149087710/delivery?warehouseId=2237337",
        "Пенза":           "https://partner.market.yandex.ru/shop/149085812/delivery?warehouseId=2234713",
    },
    "ХМ": {
        "Нижний Новгород": "https://partner.market.yandex.ru/shop/149090350/delivery?warehouseId=2241460",
        "Москва":          "https://partner.market.yandex.ru/shop/149088932/delivery?warehouseId=2239317",
        "Смоленск":        "https://partner.market.yandex.ru/shop/149088926/delivery?warehouseId=2239315",
        "Пенза":           "https://partner.market.yandex.ru/shop/149088920/delivery?warehouseId=2239309",
    },
    "ТС": {
        "Нижний Новгород": "https://partner.market.yandex.ru/shop/149091241/delivery?warehouseId=2243354",
        "Москва":          "https://partner.market.yandex.ru/shop/149091238/delivery?warehouseId=2243351",
        "Смоленск":        "https://partner.market.yandex.ru/shop/149091234/delivery?warehouseId=2243347",
        "Пенза":           "https://partner.market.yandex.ru/shop/149091231/delivery?warehouseId=2243346",
    },
}

# Федеральные округа — точные названия с сайта
FEDERAL_DISTRICTS: List[str] = [
    "Центральный федеральный округ",
    "Северо-Западный федеральный округ",
    "Южный федеральный округ",
    "Северо-Кавказский федеральный округ",
    "Приволжский федеральный округ",
    "Уральский федеральный округ",
    "Сибирский федеральный округ",
    "Дальневосточный федеральный округ",
]

# Настройка логгера
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)-7s] %(message)s",
    handlers=[logging.FileHandler(LOG_FILE, encoding="utf-8")],
)
logger = logging.getLogger("YD")


# ══════════════════════════════════════════════
#  UI — ТЕРМИНАЛЬНЫЙ ИНТЕРФЕЙС
# ══════════════════════════════════════════════
W = 66  # ширина рамки

def _c(code: str, text: str) -> str:
    """Применяет ANSI-цвет к тексту."""
    codes = {
        "red":    "\033[91m", "green":  "\033[92m", "yellow": "\033[93m",
        "blue":   "\033[94m", "cyan":   "\033[96m", "white":  "\033[97m",
        "grey":   "\033[90m", "bold":   "\033[1m",  "reset":  "\033[0m",
    }
    return f"{codes.get(code, '')}{text}{codes['reset']}"

def _pad(text: str, width: int) -> str:
    """Дополняет строку пробелами до нужной ширины (с учётом unicode)."""
    vis = re.sub(r'\033\[[^m]*m', '', text)  # убираем ANSI для подсчёта длины
    return text + " " * max(0, width - len(vis))

def ui_top() -> None:
    print(_c("cyan", "╔" + "═" * W + "╗"))

def ui_bot() -> None:
    print(_c("cyan", "╚" + "═" * W + "╝"))

def ui_sep(label: str = "", color: str = "cyan") -> None:
    if label:
        inner = f"  {label}  "
        pad = (W - len(inner)) // 2
        line = "═" * pad + inner + "═" * (W - pad - len(inner))
    else:
        line = "═" * W
    print(_c(color, "╠" + line + "╣"))

def ui_row(text: str = "", color: str = "white") -> None:
    inner = _c(color, text)
    pad_count = W - len(re.sub(r'\033\[[^m]*m', '', text)) - 1
    print(_c("cyan", "║ ") + inner + " " * max(0, pad_count) + _c("cyan", "║"))

def ui_empty() -> None:
    print(_c("cyan", "║") + " " * W + _c("cyan", "║"))

def ui_header() -> None:
    ui_top()
    ui_empty()
    title = "ЯНДЕКС МАРКЕТ — АВТОЗАПОЛНЕНИЕ ДОСТАВКИ"
    p = (W - len(title)) // 2
    print(_c("cyan", "║") + " " * p + _c("bold", _c("yellow", title)) + " " * (W - p - len(title)) + _c("cyan", "║"))
    sub = "v8.0  •  пошаговое раскрытие  •  все 8 округов"
    p2 = (W - len(sub)) // 2
    print(_c("cyan", "║") + " " * p2 + _c("grey", sub) + " " * (W - p2 - len(sub)) + _c("cyan", "║"))
    ui_empty()

def ui_status(icon: str, text: str, color: str = "white") -> None:
    msg = f"{icon}  {text}"
    vis_len = len(re.sub(r'\033\[[^m]*m', '', msg))
    spaces = W - vis_len - 1
    print(_c("cyan", "║ ") + _c(color, msg) + " " * max(0, spaces) + _c("cyan", "║"))

def ui_progress(done: int, total: int, label: str = "", color: str = "green") -> None:
    """Рисует прогресс-бар внутри рамки."""
    pct = done / total if total else 0
    bar_width = W - 30
    filled = int(pct * bar_width)
    bar_s = "█" * filled + "░" * (bar_width - filled)
    msg = f"{label:<10} [{bar_s}] {pct * 100:5.1f}%  {done}/{total}"
    vis_len = len(re.sub(r'\033\[[^m]*m', '', msg))
    print(_c("cyan", "║ ") + _c(color, msg) + " " * max(0, W - vis_len - 1) + _c("cyan", "║"))

def ui_ask(prompt: str, options: List[str]) -> str:
    """Интерактивный выбор из списка."""
    ui_empty()
    ui_row(prompt, "yellow")
    for i, opt in enumerate(options, 1):
        ui_row(f"  {i}.  {opt}", "white")
    ui_empty()
    while True:
        try:
            raw = input(_c("cyan", "║ ") + _c("yellow", "▶ Ваш выбор: ")).strip()
            if raw.isdigit() and 1 <= int(raw) <= len(options):
                return options[int(raw) - 1]
            ui_row("⚠  Введите номер из списка", "yellow")
        except (KeyboardInterrupt, EOFError):
            sys.exit(0)

def ui_ask_yn(prompt: str) -> bool:
    return ui_ask(prompt, ["Да", "Нет"]) == "Да"

def ui_wait_enter(msg: str = "Нажмите Enter для продолжения...") -> None:
    ui_empty()
    ui_row(msg, "yellow")
    ui_empty()
    try:
        input(_c("cyan", "║ ") + _c("grey", "[ Enter ] "))
    except (KeyboardInterrupt, EOFError):
        sys.exit(0)

def ui_dashboard(done: int, total: int, district: str, filled_cities: List[str]) -> None:
    """Дашборд — выводится каждые 50 найденных городов."""
    ui_sep("ДАШБОРД", "yellow")
    ui_row(f"📍 Округ:           {district}", "cyan")
    ui_row(f"✅ Найдено всего:   {done} из {total}", "green")
    ui_row(f"❌ Осталось:        {total - done}", "red" if total - done > 0 else "grey")
    if filled_cities:
        ui_row("Последние заполненные:", "grey")
        for c in filled_cities[-5:]:
            ui_row(f"   ✔  {c}", "green")
    ui_sep()


# ══════════════════════════════════════════════
#  ВСПОМОГАТЕЛЬНЫЕ УТИЛИТЫ
# ══════════════════════════════════════════════
def normalize(text: str) -> str:
    """Нормализует название города/региона для сравнения."""
    if not text:
        return ""
    t = text.lower().strip()
    t = t.replace("ё", "е")
    t = re.sub(r"[•·\t\r\n]+", " ", t)
    t = re.sub(r"\bг\.?\s*", "", t)
    t = re.sub(r"\bгород\b", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t

def check_pause() -> None:
    """Проверяет наличие файла паузы и ждёт его удаления."""
    if Path(PAUSE_FILE).exists():
        ui_status("⏸", f"ПАУЗА. Удалите файл {PAUSE_FILE} для продолжения.", "yellow")
        while Path(PAUSE_FILE).exists():
            time.sleep(1)
        ui_status("▶", "Продолжаем...", "green")


# ══════════════════════════════════════════════
#  EXCEL — ЗАГРУЗКА ГОРОДОВ
# ══════════════════════════════════════════════
def select_excel_file() -> str:
    files = sorted(Path(".").glob("*.xlsx"))
    # Исключаем файлы отчётов, которые создаём сами
    files = [f for f in files if not f.name.startswith("analysis_")]
    if not files:
        ui_status("❌", "Нет .xlsx файлов в текущей папке!", "red")
        ui_bot()
        sys.exit(1)
    if len(files) == 1:
        ui_status("📄", f"Файл: {files[0].name}", "green")
        return str(files[0])
    names = [f.name for f in files]
    choice = ui_ask("Выберите Excel-файл с городами:", names)
    return str(files[names.index(choice)])

def load_cities(filepath: str) -> List[str]:
    """Загружает уникальные названия городов из первого столбца Excel."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    cities: List[str] = []
    seen: Set[str] = set()
    for i, (row_,) in enumerate(ws.iter_rows(min_col=1, max_col=1, values_only=True)):
        if i == 0:
            continue  # Пропускаем заголовок
        if not row_:
            continue
        raw = re.sub(r"^г\.?\s*", "", str(row_).strip())
        raw = re.sub(r"\s+", " ", raw).strip()
        if len(raw) < 2:
            continue
        key = normalize(raw)
        if key not in seen:
            seen.add(key)
            cities.append(raw)
    wb.close()
    logger.info(f"Загружено городов из Excel: {len(cities)}")
    return cities


# ══════════════════════════════════════════════
#  КАРТА ПУТЕЙ
# ══════════════════════════════════════════════
def load_map() -> Dict:
    if not Path(MAP_FILE).exists():
        return {}
    try:
        with open(MAP_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Ошибка загрузки карты: {e}")
        return {}

def save_map(data: Dict) -> None:
    with open(MAP_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    logger.info(f"Карта сохранена: {len(data)} записей")


# ══════════════════════════════════════════════
#  ПРОГРЕСС
# ══════════════════════════════════════════════
def load_progress() -> Dict:
    if not Path(PROGRESS_FILE).exists():
        return {}
    try:
        with open(PROGRESS_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_progress(progress: Dict) -> None:
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


# ══════════════════════════════════════════════
#  БРАУЗЕР — ЗАПУСК И НАВИГАЦИЯ
# ══════════════════════════════════════════════
def launch_browser(profile_dir: str = None):
    """Запускает браузер с постоянным профилем."""
    max_retries = 3
    
    for attempt in range(max_retries):
        try:
            from playwright.sync_api import sync_playwright
            pw = sync_playwright().start()
            
            # Убедимся что директория профиля существует
            if profile_dir:
                Path(profile_dir).mkdir(parents=True, exist_ok=True)
            
            # Базовые аргументы для стабильного запуска
            args = [
                "--disable-dev-shm-usage", 
                "--no-sandbox",
                "--disable-web-security",
                "--disable-features=VizDisplayCompositor",
                "--disable-background-timer-throttling",
                "--disable-backgrounding-occluded-windows",
                "--disable-renderer-backgrounding",
                "--disable-background-networking",
                "--disable-ipc-flooding-protection",
                "--password-store=basic",
                "--use-mock-keychain",
                "--no-first-run",
                "--disable-default-apps"
            ]
            
            try:
                # Всегда используем persistent_context с профилем
                ctx = pw.chromium.launch_persistent_context(
                    user_data_dir=profile_dir or PROFILE_DIR,
                    headless=False,
                    args=args,
                    ignore_default_args=["--enable-automation"],
                )
            except Exception as launch_error:
                logger.error(f"Browser launch failed: {launch_error}")
                # Fallback на самый простой запуск
                ctx = pw.chromium.launch(
                    headless=False,
                    args=["--no-sandbox", "--disable-dev-shm-usage"],
                )
            
            page = ctx.new_page()
            page.on("pageerror", lambda e: logger.debug(f"JS error: {e}"))
            return pw, ctx, page
            
        except Exception as e:
            if attempt < max_retries - 1:
                wait_time = (attempt + 1) * 2
                logger.warning(f"Browser launch attempt {attempt + 1} failed: {e}. Retrying in {wait_time}s...")
                time.sleep(wait_time)
                continue
            else:
                logger.error(f"Final browser launch attempt failed: {e}")
                logger.error(f"Profile directory: {profile_dir}")
                logger.error(f"Working directory: {os.getcwd()}")
                raise Exception(f"Failed to launch browser after {max_retries} attempts: {e}")

def wait_for_manual_auth(page) -> None:
    """
    Ожидает ручной авторизации пользователя.
    Открывает браузер и ждет нажатия Enter после авторизации.
    """
    ui_sep("ОЖИДАНИЕ АВТОРИЗАЦИИ", "yellow")
    ui_row("🌐 Браузер открыт для авторизации", "cyan")
    ui_row("📝 Вручную авторизуйтесь в Яндекс Маркете", "yellow")
    ui_row("⏳ После успешной авторизации нажмите Enter", "green")
    ui_empty()
    ui_wait_enter("Нажмите Enter после завершения авторизации...")
    ui_sep()
    
    # Проверяем, что авторизация прошла успешно
    current_url = page.url
    if "passport" in current_url.lower() or "auth" in current_url.lower() or "login" in current_url.lower():
        ui_status("⚠️", "Вы всё ещё на странице авторизации!", "red")
        if not ui_ask_yn("Продолжить без авторизации?"):
            return wait_for_manual_auth(page)
    else:
        ui_status("✅", "Авторизация успешно пройдена", "green")

def is_already_authenticated(page) -> bool:
    """Проверяет, авторизован ли пользователь по URL и DOM."""
    try:
        url = page.url
        # Если редиректнуло на паспорт — не авторизован
        if any(x in url for x in ("passport.yandex", "auth", "login")):
            return False
        # Ищем признак залогиненного кабинета партнёра
        # Например, кнопка "Выйти" или аватар пользователя
        selectors = [
            "[data-e2e='user-menu']",
            "[data-e2e='logout']",
            "[class*='UserMenu']",
            "[class*='user-name']",
        ]
        for sel in selectors:
            if page.locator(sel).count() > 0:
                return True
        # Дополнительно — если мы уже на нужном домене без редиректа
        if "partner.market.yandex.ru" in url:
            # Ждём немного и проверяем что DOM загрузился
            page.wait_for_timeout(2_000)
            if page.locator("tr[data-tid]").count() > 0 or \
               page.locator("button[data-e2e*='delivery']").count() > 0:
                return True
    except Exception as e:
        logger.debug(f"is_already_authenticated error: {e}")
    return False

def goto_settings(page, url: str, force_manual_auth: bool = True) -> None:
    ui_status("🌐", f"Открываем: {url}", "cyan")
    page.goto(url, wait_until="domcontentloaded", timeout=60_000)
    page.wait_for_timeout(3_000)
    
    # Проверяем авторизацию автоматически
    if is_already_authenticated(page):
        ui_status("✅", "Профиль авторизован, пропускаем ручной вход", "green")
        return
    
    # Если не авторизован - ждем ручную авторизацию
    ui_status("⏸", "Требуется авторизация...", "yellow")
    wait_for_manual_auth(page)
    
    # После авторизации переходим на нужную страницу снова
    ui_status("🔄", "Переходим на страницу настроек...", "cyan")
    page.goto(url, wait_until="domcontentloaded", timeout=60_000)
    page.wait_for_timeout(3_000)

def wait_for_spinner(page, timeout_ms: int = 30_000) -> None:
    """
    Ждёт появления спиннера, затем его исчезновения.
    Если спиннер не появился — просто делает небольшую паузу.
    """
    spinner_sel = "[data-tid='7cb56119'], div[class*='spinner'], div[class*='Spinner'], svg[class*='spin']"
    try:
        page.locator(spinner_sel).first.wait_for(state="visible", timeout=1_500)
        page.locator(spinner_sel).first.wait_for(state="hidden", timeout=timeout_ms)
        page.wait_for_timeout(AFTER_SPINNER_MS)
        logger.debug("Спиннер исчез")
    except PWTimeout:
        pass  # Спиннер не появился — это нормально

def lock_drawer_open(page) -> None:
    """
    Блокирует автоматическое закрытие drawer.
    Vaul drawer закрывается при:
    1. Клике на overlay (затемнённый фон)
    2. Нажатии Escape
    3. Свайпе вниз (touch)
    Блокируем все эти события.
    """
    try:
        page.evaluate("""
            () => {
                if (window.__drawer_locked) return;
                window.__drawer_locked = true;
                
                // 1. Блокируем клики на overlay
                document.addEventListener('click', (e) => {
                    const t = e.target;
                    if (
                        t.hasAttribute('data-vaul-overlay') ||
                        t.closest('[data-vaul-overlay]') ||
                        (t.getAttribute('data-state') === 'open' && 
                         !t.closest('tr[data-tid]') &&
                         !t.closest('button') &&
                         !t.closest('input'))
                    ) {
                        e.stopImmediatePropagation();
                        e.preventDefault();
                        console.log('DRAWER LOCK: blocked overlay click');
                    }
                }, true);
                
                // 2. Блокируем Escape (только если drawer открыт)
                document.addEventListener('keydown', (e) => {
                    if (e.key === 'Escape') {
                        const rows = document.querySelectorAll('tr[data-tid]');
                        if (rows.length > 0) {
                            e.stopImmediatePropagation();
                            e.preventDefault();
                            console.log('DRAWER LOCK: blocked Escape');
                        }
                    }
                }, true);
                
                // 3. Блокируем pointer events на overlay
                const style = document.createElement('style');
                style.id = '__drawer_lock_style';
                style.textContent = '[data-vaul-overlay] { pointer-events: none !important; }';
                document.head.appendChild(style);
                
                console.log('DRAWER LOCK: installed');
            }
        """)
        logger.info("Drawer lock установлен")
    except Exception as e:
        logger.debug(f"lock_drawer_open error: {e}")


def is_drawer_open(page) -> bool:
    """
    Проверяет, открыт ли drawer с деревом регионов.
    Ищет специфичные признаки drawer: много строк tr[data-tid]
    И наличие кнопок раскрытия button[aria-expanded] внутри них.
    """
    try:
        # Drawer содержит кнопки раскрытия дерева
        count = page.locator("tr[data-tid] button[aria-expanded]").count()
        if count > 0:
            return True
        # Или много строк (>5) — признак что дерево загружено
        rows = page.locator("tr[data-tid]").count()
        return rows > 5
    except Exception:
        return False

def dismiss_overlays(page) -> None:
    """
    Закрывает попапы/оверлеи которые перехватывают клики.
    НЕ трогает stickyBar — там находится кнопка «Сохранить».
    НЕ нажимает Escape — это может закрыть drawer.
    """
    try:
        page.evaluate("""
            () => {
                const selectors = [
                    '[class*="overlay"]',
                    '[data-tid="a018d5b9"]',
                    '[class*="___content___4Arvd"]',
                ];
                for (const sel of selectors) {
                    document.querySelectorAll(sel).forEach(el => {
                        if (el.style) el.style.pointerEvents = 'none';
                    });
                }
            }
        """)
        page.wait_for_timeout(200)
    except Exception:
        pass


def js_click(page, selector: str) -> bool:
    """Кликает по элементу через JS — обходит перехват pointer events."""
    try:
        result = page.evaluate(f"""
            () => {{
                const el = document.querySelector("{selector}");
                if (!el) return false;
                el.dispatchEvent(new MouseEvent('click', {{bubbles: true, cancelable: true}}));
                return true;
            }}
        """)
        return bool(result)
    except Exception:
        return False


def open_drawer(page) -> bool:
    # Если drawer уже открыт (есть строки дерева), сразу возвращаем успех
    if is_drawer_open(page):
        ui_status("✅", "Drawer уже открыт", "green")
        return True

    ui_status("🔍", "Ищем кнопку «Настроить»...", "cyan")
    dismiss_overlays(page)

    # Все возможные способы найти кнопку «Настроить»
    btn_selectors = [
        "button[data-e2e='couriers-delivery-regions-drawer']",
        "button[data-e2e*='delivery'][data-e2e*='drawer']",
        "button[data-e2e*='courier']",
    ]

    for attempt in range(1, 6):
        ui_status("🔄", f"Попытка {attempt}/5", "grey")

        try:
            # Ждём загрузки страницы и стабилизации DOM
            page.wait_for_load_state("networkidle", timeout=15_000)
            page.wait_for_timeout(1_500)

            clicked = False

            # 1) Пробуем по data-e2e селекторам
            for sel in btn_selectors:
                try:
                    btn = page.locator(sel).first
                    if btn.count() > 0 and btn.is_visible():
                        ui_status("✅", f"Кнопка найдена: {sel}", "grey")
                        btn.scroll_into_view_if_needed(timeout=3_000)
                        page.wait_for_timeout(500)
                        dismiss_overlays(page)
                        
                        # Пробуем обычный клик
                        try:
                            btn.click(timeout=5_000, force=True)
                            clicked = True
                            break
                        except Exception:
                            # Fallback на JS клик
                            if js_click(page, sel):
                                clicked = True
                                break
                except Exception as e:
                    logger.debug(f"Ошибка с селектором {sel}: {e}")
                    continue

            # 2) Fallback — ищем по тексту «Настроить»
            if not clicked:
                ui_status("⚠️", "Кнопка не найдена по data-e2e, ищем по тексту...", "yellow")
                try:
                    text_btn = page.locator("button:has-text('Настроить')").last
                    if text_btn.count() > 0:
                        text_btn.scroll_into_view_if_needed(timeout=3_000)
                        page.wait_for_timeout(500)
                        dismiss_overlays(page)
                        text_btn.click(force=True, timeout=5_000)
                        clicked = True
                        ui_status("✅", "Кнопка найдена по тексту", "grey")
                except Exception as e:
                    logger.debug(f"Ошибка поиска по тексту: {e}")

            if not clicked:
                ui_status("⚠️", f"Кнопка не найдена (попытка {attempt})", "yellow")
                page.wait_for_timeout(3_000)
                dismiss_overlays(page)
                continue

            # Ждём появления строк дерева
            ui_status("⏳", "Ждём открытия drawer...", "cyan")
            try:
                page.wait_for_selector("tr[data-tid]", timeout=20_000)
                page.wait_for_timeout(1_000)
                ui_status("✅", "Drawer открыт!", "green")
                wait_for_spinner(page)
                lock_drawer_open(page)
                return True
            except PWTimeout:
                ui_status("⚠️", "Drawer не открылся после клика", "yellow")
                if is_drawer_open(page):
                    ui_status("✅", "Drawer открыт (проверка is_drawer_open)", "green")
                    lock_drawer_open(page)
                    return True

        except PWTimeout:
            ui_status("⚠️", f"Таймаут на попытке {attempt}", "yellow")
        except Exception as e:
            ui_status("⚠️", f"Ошибка: {e}", "yellow")

        dismiss_overlays(page)
        page.wait_for_timeout(3_000)

        if is_drawer_open(page):
            ui_status("✅", "Drawer открыт!", "green")
            lock_drawer_open(page)
            return True

    # Ручное открытие
    ui_status("❓", "Нажмите кнопку «Настроить» вручную в браузере, затем Enter.", "yellow")
    ui_wait_enter("После ручного открытия нажмите Enter")
    try:
        page.wait_for_selector("tr[data-tid]", timeout=15_000)
    except PWTimeout:
        pass
    if is_drawer_open(page):
        ui_status("✅", "Drawer открыт вручную.", "green")
        lock_drawer_open(page)
        return True
    raise RuntimeError("Не удалось открыть drawer ни автоматически, ни вручную.")

def close_drawer(page) -> bool:
    """
    Закрывает drawer.
    Пробует кнопку «Закрыть», затем Escape.
    После закрытия ждёт пока строки tr[data-tid] исчезнут.
    """
    if not is_drawer_open(page):
        return True

    selectors = [
        "button[aria-label='Закрыть']",
        "button[data-e2e='close-button']",
        "[data-tid*='close']",
    ]
    for sel in selectors:
        try:
            btn = page.locator(sel).first
            if btn.count() > 0 and btn.is_visible():
                btn.click()
                page.wait_for_timeout(2_000)
                if not is_drawer_open(page):
                    logger.info("Drawer закрыт")
                    return True
        except Exception:
            continue

    # Запасной вариант — Escape
    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(2_000)
        logger.info("Drawer закрыт через Escape")
        return True
    except Exception:
        pass
    return False

def blur_active_input(page) -> None:
    """Снимает фокус с активного поля ввода перед сохранением."""
    try:
        page.evaluate("""
            () => {
                const active = document.activeElement;
                if (active && (active.tagName === 'INPUT' || active.tagName === 'TEXTAREA')) {
                    active.blur();
                }
            }
        """)
        page.wait_for_timeout(300)
    except Exception:
        pass

def save_page_changes(page) -> bool:
    """
    Сохранение через кнопку DeliverySettingsByCouriersDrawer_submit.
    Ждёт полного закрытия drawer перед возвратом.
    """
    # Снимаем фокус с последнего заполненного поля перед сохранением
    blur_active_input(page)

    for attempt in range(4):
        try:
            page.wait_for_timeout(1_000)
            ui_status("🔍", f"Ищем кнопку «Сохранить» (попытка {attempt + 1})...", "cyan")

            # Проверяем наличие и состояние кнопки
            info = page.evaluate("""
                () => {
                    const btn = document.querySelector(
                        "button[data-e2e='DeliverySettingsByCouriersDrawer_submit']"
                    );
                    if (!btn) return {status: 'not_found'};
                    return {
                        status: 'found',
                        disabled: btn.disabled || btn.getAttribute('aria-disabled') === 'true',
                        visible: btn.offsetParent !== null
                    };
                }
            """)
            logger.info(f"Кнопка Сохранить: {info}")

            if info['status'] == 'not_found':
                ui_status("⚠️", "Кнопка не найдена в DOM", "yellow")
                page.wait_for_timeout(2_000)
                continue

            if info.get('disabled'):
                ui_status("⚠️", "Кнопка disabled, ждём активации...", "yellow")
                page.wait_for_timeout(2_500)
                continue

            # Скроллим к кнопке и кликаем
            clicked = page.evaluate("""
                () => {
                    const btn = document.querySelector(
                        "button[data-e2e='DeliverySettingsByCouriersDrawer_submit']"
                    );
                    if (!btn) return false;
                    btn.scrollIntoView({block: 'center', behavior: 'instant'});
                    // Используем все типы событий для надёжности
                    ['mousedown','mouseup','click'].forEach(type => {
                        btn.dispatchEvent(new MouseEvent(type, {bubbles:true, cancelable:true}));
                    });
                    return true;
                }
            """)

            if not clicked:
                ui_status("⚠️", "Не удалось кликнуть", "yellow")
                continue

            ui_status("⏳", "Клик выполнен, ждём закрытия drawer...", "cyan")

            # Ждём закрытия drawer — до 30 секунд
            drawer_closed = False
            for tick in range(60):
                page.wait_for_timeout(500)
                
                # Проверяем все признаки закрытия
                state = page.evaluate("""
                    () => ({
                        rows: document.querySelectorAll('tr[data-tid]').length,
                        saveBtn: !!document.querySelector("button[data-e2e='DeliverySettingsByCouriersDrawer_submit']"),
                        nastroitBtn: !!document.querySelector("button[data-e2e='couriers-delivery-regions-drawer']"),
                        nastroitVisible: document.querySelector("button[data-e2e='couriers-delivery-regions-drawer']")?.offsetParent !== null
                    })
                """)

                # Drawer закрыт если:
                # 1. Кнопка "Настроить" видна И
                # 2. Строк дерева нет ИЛИ кнопки "Сохранить" нет
                if state['nastroitVisible'] and (state['rows'] == 0 or not state['saveBtn']):
                    drawer_closed = True
                    ui_status("✅", "Drawer закрыт, изменения сохранены!", "green")
                    logger.info("Сохранение выполнено успешно")
                    # Дополнительная пауза для стабилизации DOM
                    page.wait_for_timeout(1_500)
                    return True

                # Логируем каждые 5 секунд
                if tick % 10 == 9:
                    logger.info(f"Ожидание закрытия drawer (сек {tick//2}): {state}")

            if not drawer_closed:
                ui_status("⚠️", f"Drawer не закрылся за 30 сек (попытка {attempt + 1})", "yellow")

        except Exception as e:
            logger.warning(f"Ошибка сохранения (попытка {attempt + 1}): {e}")
            page.wait_for_timeout(2_000)

    # Если автосохранение не сработало — просим пользователя
    ui_status("⚠️", "Автосохранение не сработало", "yellow")
    ui_wait_enter("💾  Нажмите жёлтую кнопку «Сохранить» вручную, затем Enter")
    
    # Ждём пока drawer точно закроется
    for _ in range(20):
        page.wait_for_timeout(500)
        if not is_drawer_open(page):
            return True
    
    return False


# ══════════════════════════════════════════════
#  РАБОТА С ДЕРЕВОМ — ПОШАГОВОЕ РАСКРЫТИЕ
# ══════════════════════════════════════════════

def get_row_text(row) -> str:
    """Получает чистый текст первой ячейки строки (название)."""
    try:
        td = row.locator("td").first
        if td.count() > 0:
            t = (td.text_content() or "").strip()
        else:
            t = (row.text_content() or "").strip()
        return re.sub(r"\s+", " ", t).strip()
    except Exception:
        return ""

def row_is_expandable(row) -> bool:
    """Возвращает True, если у строки есть кнопка раскрытия."""
    try:
        return row.locator("button[aria-expanded]").count() > 0
    except Exception:
        return False

def row_is_expanded(row) -> bool:
    """Возвращает True, если строка уже раскрыта."""
    try:
        btn = row.locator("button[aria-expanded]").first
        return btn.get_attribute("aria-expanded") == "true"
    except Exception:
        return False

def row_is_city(row) -> bool:
    """Город — строка БЕЗ кнопки раскрытия (лист дерева)."""
    return not row_is_expandable(row)

def scroll_drawer_gently(page) -> None:
    """
    Мягкий скролл drawer для подгрузки виртуализированных строк.
    Помогает асинхронно подгружаемым деревьям.
    """
    try:
        page.evaluate("""
            () => {
                // Ищем скроллируемый контейнер drawer
                const drawer = document.querySelector(
                    '[data-e2e="levitan-drawer-wrapper"]'
                );
                if (!drawer) return;
                
                const container = drawer.querySelector(
                    '[style*="overflow"]'
                ) || drawer;
                
                // Небольшой скролл вниз-вверх для триггера подгрузки
                const oldScroll = container.scrollTop;
                container.scrollTop += 50;
                setTimeout(() => {
                    container.scrollTop = oldScroll;
                }, 100);
            }
        """)
        page.wait_for_timeout(200)
    except Exception as e:
        logger.debug(f"scroll_drawer_gently error: {e}")

def scroll_to_row_in_drawer(page, row) -> None:
    """
    Скроллит контейнер drawer так, чтобы строка была видна.
    НЕ использует scroll_into_view_if_needed — это опасно в Vaul.
    """
    try:
        page.evaluate("""
            (row) => {
                // Ищем скроллируемый контейнер drawer
                const drawer = document.querySelector(
                    '[data-e2e="levitan-drawer-wrapper"]'
                );
                if (!drawer) return;
                
                const container = drawer.querySelector(
                    '[style*="overflow"]'
                ) || drawer;
                
                const rowRect = row.getBoundingClientRect();
                const contRect = container.getBoundingClientRect();
                
                // Если строка уже видна — ничего не делаем
                if (rowRect.top >= contRect.top && 
                    rowRect.bottom <= contRect.bottom) return;
                
                // Скроллим контейнер, не страницу
                container.scrollTop += (rowRect.top - contRect.top) - 100;
            }
        """, row.element_handle())
        page.wait_for_timeout(150)
    except Exception as e:
        logger.debug(f"scroll_to_row_in_drawer error: {e}")

def click_expand_row(page, row) -> bool:
    """
    Разворачивает ОДИН узел:
      1. Скроллит контейнер drawer (не scroll_into_view_if_needed)
      2. Кликает по кнопке раскрытия
      3. Ждёт AFTER_CLICK_MS
      4. Ждёт исчезновения спиннера
    Возвращает True если клик был сделан.
    """
    try:
        btn = row.locator("button[aria-expanded='false']").first
        if btn.count() == 0:
            return False
        
        # БЫЛО: btn.scroll_into_view_if_needed()  ← убираем
        # СТАЛО: мягкий скролл через контейнер
        scroll_to_row_in_drawer(page, row)
        
        btn.click()
        page.wait_for_timeout(AFTER_CLICK_MS)
        wait_for_spinner(page)
        return True
    except Exception as e:
        logger.debug(f"click_expand_row error: {e}")
        return False


def find_district_row(page, district_name: str):
    """
    Находит строку федерального округа: сначала точное совпадение,
    затем частичное (district_name содержится в тексте строки).
    """
    district_norm = normalize(district_name)
    rows = page.locator("tr[data-tid]")
    n = rows.count()

    ui_status("🔍", f"Ищем '{district_name}' среди {n} строк...", "grey")

    # Диагностика: показываем первые expandable строки
    sample = []
    for i in range(min(30, n)):
        r = rows.nth(i)
        if row_is_expandable(r):
            txt = get_row_text(r)
            sample.append(txt)
    if sample:
        ui_status("�", f"Expandable строки: {', '.join(sample[:8])}", "grey")

    # Точное совпадение
    for i in range(n):
        r = rows.nth(i)
        if not row_is_expandable(r):
            continue
        txt = normalize(get_row_text(r))
        if txt == district_norm:
            ui_status("✅", f"Найден точный округ: {get_row_text(r)}", "green")
            return r

    # Частичное совпадение
    for i in range(n):
        r = rows.nth(i)
        if not row_is_expandable(r):
            continue
        txt = normalize(get_row_text(r))
        if district_norm in txt or txt in district_norm:
            ui_status("✅", f"Найден частичный округ: {get_row_text(r)}", "green")
            return r

    # Если не нашли — выводим ВСЕ expandable строки для диагностики
    ui_status("❌", f"Округ '{district_name}' не найден. Все expandable строки:", "red")
    for i in range(min(50, n)):
        r = rows.nth(i)
        if row_is_expandable(r):
            ui_status("  →", get_row_text(r), "grey")
    return None


def _get_collapsed_batch(page, other_districts: List[str], batch: int = 3) -> List[Dict]:
    """Возвращает до `batch` свёрнутых узлов, не принадлежащих другим округам."""
    return page.evaluate("""
        ([otherDistricts, batch]) => {
            const rows = Array.from(document.querySelectorAll('tr[data-tid]'));
            const result = [];
            for (let i = 0; i < rows.length && result.length < batch; i++) {
                const btn = rows[i].querySelector("button[aria-expanded='false']");
                if (!btn || btn.offsetParent === null) continue;
                const rowText = (rows[i].querySelector('td')?.textContent || '').trim();
                let isOther = false;
                for (const d of otherDistricts) {
                    if (rowText.includes(d) || d.includes(rowText)) { isOther = true; break; }
                }
                if (!isOther) result.push({ idx: i, text: rowText });
            }
            return result;
        }
    """, [other_districts, batch])


def expand_and_fill_district(
    page,
    district_name: str,
    cities_dict: Dict[str, str],
    done_cities: Set[str],
    total_cities: int,
    filled_count: List[int],
    filled_names: List[str],
) -> Tuple[int, int]:
    """
    Раскрывает округ по 3 узла за раз и СРАЗУ заполняет города.
    """
    district_row = find_district_row(page, district_name)
    if district_row is None:
        logger.warning(f"Округ не найден: {district_name}")
        return 0, 0

    if not row_is_expanded(district_row):
        ui_status("🔓", f"Раскрываем: {district_name}", "cyan")
        click_expand_row(page, district_row)
    else:
        ui_status("ℹ", f"Уже раскрыт: {district_name}", "grey")

    total_expanded = 1
    district_filled = 0
    rounds_no_change = 0
    other_districts = [d for d in FEDERAL_DISTRICTS if d != district_name]

    def scan_and_fill():
        nonlocal district_filled
        rows = page.locator("tr[data-tid]")
        n = rows.count()
        for i in range(n):
            r = rows.nth(i)
            if not row_is_city(r):
                continue
            city_norm = normalize(get_row_text(r))
            
            # Пропускаем если: не в нашем списке ИЛИ уже заполнен
            if city_norm not in cities_dict:
                continue
            if city_norm in done_cities:  # <-- вместо filled_indices
                continue
            
            original_name = cities_dict[city_norm]
            # НЕ скроллим к строке - это может закрыть drawer
            # Строка уже должна быть видна после раскрытия узла
            if set_value_in_row(r):
                district_filled += 1
                filled_count[0] += 1
                done_cities.add(city_norm)      # <-- пишем в общий стейт
                filled_names.append(original_name)
                # НЕ делаем del cities_dict[city_norm]
                logger.info(f"✔ Заполнен: {original_name}")
                if filled_count[0] % 50 == 0:
                    ui_dashboard(filled_count[0], total_cities, district_name, filled_names)
            else:
                logger.warning(f"✘ Не удалось: {get_row_text(r)}")

    for iteration in range(100):  # Уменьшаем количество итераций для безопасности
        check_pause()

        # Если drawer закрылся во время работы — пытаемся переоткрыть
        if not is_drawer_open(page):
            ui_status("⚠️", "Drawer закрылся! Переоткрываем...", "yellow")
            logger.warning(f"Drawer закрылся на итерации {iteration}, пытаемся переоткрыть")
            
            # Пытаемся переоткрыть drawer
            try:
                if not open_drawer(page):
                    ui_status("❌", "Не удалось переоткрыть drawer, выходим", "red")
                    break
                
                # Раскрываем корень и находим наш округ заново
                ensure_root_expanded(page)
                district_row = find_district_row(page, district_name)
                if district_row is None:
                    ui_status("❌", "Не удалось найти округ после переоткрытия", "red")
                    break
                
                if not row_is_expanded(district_row):
                    click_expand_row(page, district_row)
                
                ui_status("✅", "Drawer переоткрыт, продолжаем работу", "green")
            except Exception as e:
                logger.error(f"Ошибка при переоткрытии drawer: {e}")
                break

        # Сканируем и заполняем видимые города
        scan_and_fill()

        # Берём до 3 свёрнутых узлов за раз
        batch = _get_collapsed_batch(page, other_districts, 3)

        if not batch:
            rounds_no_change += 1
            if rounds_no_change >= 2:
                # Проверяем только свёрнутые узлы внутри нашего округа (не других)
                still = page.evaluate("""
                    ([otherDistricts]) => {
                        const rows = Array.from(document.querySelectorAll('tr[data-tid]'));
                        for (const row of rows) {
                            const btn = row.querySelector("button[aria-expanded='false']");
                            if (!btn || btn.offsetParent === null) continue;
                            const rowText = (row.querySelector('td')?.textContent || '').trim();
                            let isOther = false;
                            for (const d of otherDistricts) {
                                if (rowText.includes(d) || d.includes(rowText)) {
                                    isOther = true; break;
                                }
                            }
                            if (!isOther) return true;
                        }
                        return false;
                    }
                """, [other_districts])
                if not still:
                    ui_status("✅", f"Все узлы округа раскрыты. Заполнено: {district_filled}", "green")
                    break  # все узлы округа раскрыты
                rounds_no_change = 0
            continue

        rounds_no_change = 0
        for item in batch:
            try:
                target_row = page.locator("tr[data-tid]").nth(item["idx"])
                if click_expand_row(page, target_row):
                    total_expanded += 1
            except Exception as e:
                logger.debug(f"Ошибка раскрытия: {e}")

        if total_expanded % 15 == 0:
            ui_status("🔓", f"Раскрыто: {total_expanded} узлов, заполнено: {district_filled} городов", "cyan")

    scan_and_fill()
    return total_expanded, district_filled


def expand_district_fully(page, district_name: str) -> int:
    """Для режимов Анализ и Сканер — раскрывает по 3 узла за раз."""
    # Ищем округ с повторными попытками
    district_row = None
    for attempt in range(3):
        district_row = find_district_row(page, district_name)
        if district_row:
            break
        ui_status("⏳", f"Повторная попытка {attempt+1} найти округ '{district_name}'...", "yellow")
        page.wait_for_timeout(1500)
        scroll_drawer_gently(page)
    
    if district_row is None:
        ui_status("❌", f"Округ не найден после 3 попыток: {district_name}", "red")
        return 0
    if not row_is_expanded(district_row):
        click_expand_row(page, district_row)

    total_expanded = 1
    rounds_no_change = 0
    other_districts = [d for d in FEDERAL_DISTRICTS if d != district_name]

    for iteration in range(100):  # Уменьшаем количество итераций для безопасности
        check_pause()
        if iteration % 5 == 0:
            scroll_drawer_gently(page)

        batch = _get_collapsed_batch(page, other_districts, 3)

        if not batch:
            rounds_no_change += 1
            if rounds_no_change >= 2:
                scroll_drawer_gently(page)
                page.wait_for_timeout(400)
                still = page.evaluate(
                    "() => document.querySelectorAll(\"button[aria-expanded='false']\").length"
                )
                if still == 0:
                    break
                rounds_no_change = 0
            continue

        rounds_no_change = 0
        for item in batch:
            try:
                target_row = page.locator("tr[data-tid]").nth(item["idx"])
                if click_expand_row(page, target_row):
                    total_expanded += 1
            except Exception as e:
                logger.debug(f"Ошибка раскрытия: {e}")

    ui_status("✅", f"Раскрыто узлов в '{district_name}': {total_expanded}", "green")
    return total_expanded
# ══════════════════════════════════════════════
#  ЗАПОЛНЕНИЕ ЗНАЧЕНИЯ В ПОЛЕ ВВОДА
# ══════════════════════════════════════════════
def set_value_in_row(row, value: str = INPUT_VALUE) -> bool:
    """
    Записывает значение в поле ввода строки. 
    Повторяет попытки до успеха (до 5 раз) с увеличивающейся паузой.
    Возвращает True только когда значение реально стало = value.
    """
    # Нормализуем значение
    value = str(value)

    for attempt in range(1, 6):
        try:
            inputs = row.locator(
                "input[type='text'], input[type='number'], input[inputmode='text'], input[inputmode='numeric']"
            )
            count = inputs.count()
            if count == 0:
                # Нет полей ввода – странно, но подождём
                time.sleep(0.5)
                continue

            target = None
            for i in range(count):
                inp = inputs.nth(i)
                try:
                    if not inp.is_visible():
                        continue
                    if inp.get_attribute("readonly") is not None:
                        continue
                    if inp.get_attribute("disabled") is not None:
                        continue
                    ph = (inp.get_attribute("placeholder") or "").lower()
                    val_attr = (inp.get_attribute("value") or "").lower()
                    if "единые" in ph or "единые" in val_attr:
                        continue
                    # Проверяем текущее значение
                    try:
                        current = inp.input_value()
                    except:
                        current = (inp.get_attribute("value") or "").strip()
                    if current == value:
                        return True  # уже заполнено
                    if "разное" in ph:
                        target = inp
                        break
                    if target is None:
                        target = inp
                except:
                    continue

            if target is None:
                # Если поле не найдено, возможно его ещё нет – подождём
                time.sleep(0.5)
                continue

            # Пытаемся заполнить
            target.click(timeout=3_000)
            target.fill(value)
            target.evaluate("el => el.blur()")
            time.sleep(0.2)

            # Проверка
            try:
                actual = target.input_value()
            except:
                actual = (target.get_attribute("value") or "").strip()
            if actual == value:
                return True

            # Вторая попытка: triple_click + type
            target.triple_click()
            target.type(value, delay=30)
            target.evaluate("el => el.blur()")
            time.sleep(0.2)
            try:
                actual2 = target.input_value()
            except:
                actual2 = (target.get_attribute("value") or "").strip()
            if actual2 == value:
                return True

            # Третья попытка: напрямую через JS
            target.evaluate(f"el => el.value = '{value}'")
            target.evaluate("el => el.dispatchEvent(new Event('input', {bubbles: true}))")
            target.evaluate("el => el.blur()")
            time.sleep(0.2)
            try:
                actual3 = target.input_value()
            except:
                actual3 = (target.get_attribute("value") or "").strip()
            if actual3 == value:
                return True

            # Не удалось – увеличиваем паузу и пробуем снова
            ui_status("⚠️", f"Попытка {attempt}/5: значение не записалось, ждём...", "yellow")
            time.sleep(1.5 * attempt)  # растущая пауза

        except Exception as e:
            logger.debug(f"set_value_in_row попытка {attempt} ошибка: {e}")
            time.sleep(1.0)

    # После всех попыток – ошибка
    ui_status("❌", f"Не удалось записать {value} после 5 попыток", "red")
    return False




# ══════════════════════════════════════════════
#  СКАНИРОВАНИЕ ГОРОДОВ В ОКРУГЕ
# ══════════════════════════════════════════════

# JS для извлечения городов и их путей.
# Поскольку DOM Яндекс.Маркета — плоская таблица (все tr — соседи),
# путь строим по data-атрибуту уровня вложенности (data-depth / padding-left),
# сравнивая текущую строку с предыдущими строками, у которых уровень меньше.
_JS_GET_CITIES = """
() => {
    const rows = Array.from(document.querySelectorAll('tr[data-tid]'));
    const results = [];

    // Уровень вложенности через класс __use--indent_N___
    function getIndent(row) {
        const cls = row.className || '';
        const m = cls.match(/__use--indent_(\\d+)___/);
        if (m) return parseInt(m[1]);
        // Запасной — padding-left первой ячейки
        const td = row.querySelector('td');
        if (td) {
            const inner = td.querySelector('button, span[data-e2e]');
            if (inner) {
                const pl = parseFloat(window.getComputedStyle(inner).paddingLeft) || 0;
                return Math.round(pl / 16);
            }
        }
        return 0;
    }

    function getLabel(row) {
        const span = row.querySelector("span[data-e2e^='region-name-']");
        if (span) return span.textContent.trim();
        const td = row.querySelector('td');
        return (td ? td.textContent : row.textContent).trim().replace(/\\s+/g, ' ');
    }

    const stack = [];  // [{indent, text}]

    for (let i = 0; i < rows.length; i++) {
        const row = rows[i];
        const text = getLabel(row);
        if (!text || text.length < 2) continue;

        // Фильтр мусора (дни недели и т.п.)
        const low = text.toLowerCase();
        if (['пн','вт','ср','чт','пт','сб','вс'].includes(low)) continue;

        const indent = getIndent(row);
        const isCity = !row.querySelector("button[aria-expanded]");

        while (stack.length > 0 && stack[stack.length - 1].indent >= indent) {
            stack.pop();
        }
        stack.push({ indent, text });

        if (!isCity) continue;

        const path = stack.map(s => s.text);
        results.push({ city: text, path: path });
    }
    return results;
}
"""

def ensure_root_expanded(page) -> bool:
    """Раскрывает корневой узел дерева (первый expandable элемент)."""
    page.wait_for_timeout(800)
    rows = page.locator("tr[data-tid]")
    n = rows.count()
    if n == 0:
        ui_status("❌", "Нет строк tr[data-tid] в DOM!", "red")
        return False
    for i in range(min(10, n)):
        r = rows.nth(i)
        if row_is_expandable(r):
            if row_is_expanded(r):
                ui_status("✅", f"Корень раскрыт: '{get_row_text(r)}'", "green")
                # Дополнительное ожидание для подгрузки виртуализированных строк
                page.wait_for_timeout(1000)
                return True
            txt = get_row_text(r)
            ui_status("🌳", f"Раскрываем корень: '{txt}'", "cyan")
            click_expand_row(page, r)
            wait_for_spinner(page)
            # Увеличиваем ожидание после раскрытия корня
            page.wait_for_timeout(1000)
            return True
    ui_status("⚠️", "Корневой узел не найден", "yellow")
    return False


def collect_cities_from_page(page) -> List[Dict]:
    """Собирает все города (листья) с их путями через JS."""
    try:
        results = page.evaluate(_JS_GET_CITIES)
        return results if isinstance(results, list) else []
    except Exception as e:
        logger.warning(f"collect_cities_from_page JS error: {e}")
        return []

def find_and_fill_cities_in_district(
    page,
    cities_dict: Dict[str, str],   # norm -> original
    district_name: str,
    done_cities: Set[str],
    total_cities: int,
    filled_count: List[int],        # изменяемый счётчик [filled]
    filled_names: List[str],        # список названий найденных городов
) -> int:
    """
    Проходит по всем строкам страницы, находит города из cities_dict
    и записывает значение '20'.
    Делает несколько проходов с прокруткой для виртуализированного DOM.
    """
    district_filled = 0
    processed: Set[str] = set()

    # Несколько проходов: прокручиваем и заполняем
    for pass_num in range(3):
        scroll_drawer_gently(page)
        page.wait_for_timeout(300)

        rows = page.locator("tr[data-tid]")
        n = rows.count()

        for i in range(n):
            check_pause()
            r = rows.nth(i)
            if not row_is_city(r):
                continue
            city_name = get_row_text(r)
            city_norm = normalize(city_name)

            if city_norm in processed:
                continue
            if city_norm not in cities_dict:
                continue

            original_name = cities_dict[city_norm]
            # НЕ скроллим к строке - это может закрыть drawer

            if set_value_in_row(r):
                district_filled += 1
                filled_count[0] += 1
                done_cities.add(city_norm)
                filled_names.append(original_name)
                processed.add(city_norm)
                del cities_dict[city_norm]
                logger.info(f"✔ Заполнен: {original_name}")

                if filled_count[0] % 50 == 0:
                    ui_dashboard(filled_count[0], total_cities, district_name, filled_names)
            else:
                logger.warning(f"✘ Не удалось заполнить: {city_name}")

        # Если все города найдены — выходим раньше
        if not cities_dict:
            break

    return district_filled


# ══════════════════════════════════════════════
#  РЕЖИМ 2: БЕЗ КАРТЫ
# ══════════════════════════════════════════════
def mode_without_map(page, cities: List[str]) -> None:
    ui_sep("РЕЖИМ 2 — БЕЗ КАРТЫ (пошагово по округам)", "blue")
    ui_status("ℹ", "Один округ за раз: раскрываем → заполняем → сохраняем → следующий", "cyan")
    ui_empty()

    selected_districts = select_districts()

    # Прогресс
    progress = load_progress()
    done_districts: Set[str] = set(progress.get("done_districts", []))
    done_cities: Set[str]    = set(progress.get("done_cities", []))

    cities_dict = {normalize(c): c for c in cities if normalize(c) not in done_cities}
    total_cities = len(cities)
    filled_count  = [total_cities - len(cities_dict)]
    filled_names: List[str] = []

    if filled_count[0] > 0:
        ui_status("ℹ", f"Продолжаем: уже заполнено {filled_count[0]} из {total_cities}", "grey")

    if not is_drawer_open(page):
        open_drawer(page)
    ensure_root_expanded(page)

    for idx, district in enumerate(selected_districts, 1):
        district_norm = normalize(district)
        if district_norm in done_districts:
            ui_status("⏩", f"Пропуск (уже обработан): {district}", "grey")
            continue

        if not cities_dict:
            ui_status("�", "Все города уже найдены! Выходим.", "green")
            break

        check_pause()
        ui_empty()
        ui_sep(f"{idx}/8  {district}", "cyan")
        ui_progress(filled_count[0], total_cities, "Прогресс", "green")

        # Drawer уже открыт (открыт в конце предыдущей итерации или в начале)
        # Раскрываем округ и СРАЗУ заполняем города по мере раскрытия
        expanded, district_filled = expand_and_fill_district(
            page, district,
            cities_dict, done_cities, total_cities, filled_count, filled_names
        )
        if expanded == 0:
            ui_status("⚠️", f"Округ не удалось раскрыть, пропускаем: {district}", "yellow")
            done_districts.add(district_norm)
            progress["done_districts"] = list(done_districts)
            save_progress(progress)
            # Перезагружаем страницу и открываем drawer для следующего округа
            page.reload(wait_until="domcontentloaded", timeout=60_000)
            page.wait_for_timeout(3_000)
            open_drawer(page)
            ensure_root_expanded(page)
            continue

        ui_status("✅", f"В '{district}' заполнено: {district_filled}, осталось: {len(cities_dict)}", "green")

        # Сохраняем прогресс
        done_districts.add(district_norm)
        progress["done_districts"] = list(done_districts)
        progress["done_cities"]    = list(done_cities)
        save_progress(progress)

        # Сохраняем изменения на странице
        save_success = save_page_changes(page)
        
        # Если это не последний округ — перезагружаем и открываем drawer для следующего
        if idx < len(selected_districts):
            if save_success:
                ui_status("🔄", "Перезагружаем страницу для следующего округа...", "cyan")
                page.reload(wait_until="domcontentloaded", timeout=60_000)
                page.wait_for_timeout(3_000)
                dismiss_overlays(page)
                
                # Открываем drawer для следующего округа
                if not open_drawer(page):
                    ui_status("⚠️", "Не удалось открыть drawer, пропускаем оставшиеся округа", "yellow")
                    break
                ensure_root_expanded(page)
            else:
                # Если сохранение не удалось — пробуем восстановить состояние
                ui_status("🔄", "Восстанавливаем состояние...", "cyan")
                page.reload(wait_until="domcontentloaded", timeout=60_000)
                page.wait_for_timeout(3_000)
                dismiss_overlays(page)
                if not open_drawer(page):
                    ui_status("⚠️", "Не удалось открыть drawer, пропускаем оставшиеся округа", "yellow")
                    break
                ensure_root_expanded(page)

    # ── ИТОГОВЫЙ ОТЧЁТ ──
    ui_empty()
    ui_sep("ИТОГОВЫЙ ОТЧЁТ", "blue")
    ui_progress(filled_count[0], total_cities, "Итог", "green")
    ui_row(f"✅  Успешно заполнено:  {filled_count[0]} из {total_cities}", "green")

    not_found = list(cities_dict.values())
    ui_row(f"❌  Не найдено:         {len(not_found)}", "red" if not_found else "grey")

    if not_found:
        ui_empty()
        ui_row("Города, которые не найдены:", "yellow")
        for c in not_found[:25]:
            ui_row(f"   • {c}", "grey")
        if len(not_found) > 25:
            ui_row(f"   ... и ещё {len(not_found) - 25}", "grey")
        choice = ui_ask("Что делать с ненайденными?", ["Найти вручную", "Сохранить как есть"])
        if "вручную" in choice:
            ui_wait_enter("🖱  Найдите города вручную, затем нажмите Enter")
            save_page_changes(page)
    else:
        ui_status("🎉", "Все города найдены и заполнены!", "green")
        ui_sep("РАБОТА ЗАВЕРШЕНА", "green")
        ui_row(f"✅  Заполнено городов: {filled_count[0]} из {total_cities}", "green")
        ui_row(f"📋  Все изменения сохранены автоматически", "cyan")
        # Удаляем файл прогресса — задача выполнена
        if Path(PROGRESS_FILE).exists():
            Path(PROGRESS_FILE).unlink()


# ══════════════════════════════════════════════
#  РЕЖИМ 1: С КАРТОЙ
# ══════════════════════════════════════════════
def select_districts() -> List[str]:
    """
    Предлагает выбрать: все округа или конкретные.
    Возвращает список выбранных округов.
    """
    ui_sep("ВЫБОР ОКРУГОВ", "blue")
    choice = ui_ask("Какие округа обрабатывать?", ["Все округа", "Выбрать конкретные"])
    if "Все" in choice:
        ui_status("✅", "Выбраны все 8 округов", "green")
        return list(FEDERAL_DISTRICTS)

    # Показываем нумерованный список с возможностью выбора нескольких
    ui_empty()
    ui_row("Введите номера округов через пробел (например: 1 3 5):", "yellow")
    for i, d in enumerate(FEDERAL_DISTRICTS, 1):
        ui_row(f"  {i}.  {d}", "white")
    ui_empty()

    while True:
        try:
            raw = input(_c("cyan", "║ ") + _c("yellow", "▶ Номера: ")).strip()
            parts = raw.split()
            selected = []
            valid = True
            for p in parts:
                if p.isdigit() and 1 <= int(p) <= len(FEDERAL_DISTRICTS):
                    d = FEDERAL_DISTRICTS[int(p) - 1]
                    if d not in selected:
                        selected.append(d)
                else:
                    ui_row(f"⚠  Неверный номер: {p}", "yellow")
                    valid = False
                    break
            if valid and selected:
                ui_empty()
                ui_row("Выбраны округа:", "green")
                for d in selected:
                    ui_row(f"  ✔  {d}", "green")
                return selected
            ui_row("⚠  Введите хотя бы один корректный номер", "yellow")
        except (KeyboardInterrupt, EOFError):
            sys.exit(0)



def mode_with_map(page, cities: List[str]) -> None:
    ui_sep("РЕЖИМ 1 — С КАРТОЙ (быстрый поиск)", "blue")

    city_map = load_map()
    if not city_map:
        ui_status("❌", f"Файл {MAP_FILE} не найден! Сначала запустите режим 'Сканер'.", "red")
        return

    ui_status("🗺", f"Карта загружена: {len(city_map)} записей", "green")

    selected_districts = select_districts()

    # Группируем города по округам используя путь из карты
    cities_by_district: Dict[str, List[Dict]] = {}
    not_in_map: List[str] = []
    for city in cities:
        cn = normalize(city)
        entry = city_map.get(cn)
        if not entry:
            not_in_map.append(city)
            continue
        path = entry.get("path_original", [])
        district = ""
        for p in path:
            pn = normalize(p)
            for fd in FEDERAL_DISTRICTS:
                if normalize(fd) in pn or pn in normalize(fd):
                    district = fd
                    break
            if district:
                break
        if district:
            cities_by_district.setdefault(district, []).append({"city": city, "path": path})
        else:
            not_in_map.append(city)

    if not_in_map:
        ui_status("⚠️", f"Не найдено в карте: {len(not_in_map)} городов", "yellow")

    progress = load_progress()
    done_cities: Set[str] = set(progress.get("done_cities", []))
    total_cities = len(cities)
    filled_count = [sum(1 for c in cities if normalize(c) in done_cities)]
    filled_names: List[str] = []

    open_drawer(page)
    ensure_root_expanded(page)

    for district in selected_districts:
        if district not in cities_by_district:
            ui_status("⏩", f"Нет городов для: {district}", "grey")
            continue

        district_entries = [
            e for e in cities_by_district[district]
            if normalize(e["city"]) not in done_cities
        ]
        if not district_entries:
            ui_status("⏩", f"Все города уже заполнены: {district}", "grey")
            continue

        check_pause()
        ui_sep(f"{district}", "cyan")
        ui_progress(filled_count[0], total_cities, "Прогресс", "green")
        ui_status("🗺", f"Городов для заполнения: {len(district_entries)}", "cyan")

        # Раскрываем округ с повторными попытками
        district_row = None
        for attempt in range(3):
            district_row = find_district_row(page, district)
            if district_row:
                break
            ui_status("⏳", f"Повторная попытка {attempt+1} найти округ '{district}'...", "yellow")
            page.wait_for_timeout(1500)
            scroll_drawer_gently(page)  # пробуем подгрузить строки
        
        if district_row is None:
            ui_status("❌", f"Округ не найден после 3 попыток: {district}", "red")
            page.reload(wait_until="domcontentloaded", timeout=60_000)
            page.wait_for_timeout(3_000)
            open_drawer(page)
            ensure_root_expanded(page)
            continue

        if not row_is_expanded(district_row):
            click_expand_row(page, district_row)

        district_filled = 0

        for entry in district_entries:
            city = entry["city"]
            path = entry["path"]
            city_norm = normalize(city)
            check_pause()

            # Проверяем что drawer ещё открыт
            if not is_drawer_open(page):
                ui_status("⚠️", "Drawer закрылся! Переоткрываем...", "yellow")
                open_drawer(page)
                ensure_root_expanded(page)
                # Раскрываем округ заново
                district_row = find_district_row(page, district)
                if district_row and not row_is_expanded(district_row):
                    click_expand_row(page, district_row)

            ui_status("🔍", f"Ищем: {city}", "cyan")
            ui_status("📍", f"Путь: {' → '.join(path)}", "grey")

            # Раскрываем путь и заполняем в процессе
            def strip_count_suffix(s: str) -> str:
                """Убирает числовой суффикс в конце."""
                return re.sub(r'\s+\d+\s*$', '', s).strip()

            path_opened = True
            current_row = district_row  # Начинаем с округа
            
            # Раскрываем каждый шаг пути и ищем город
            for step_idx, step_name in enumerate(path):
                step_clean = strip_count_suffix(normalize(step_name))
                
                # Ждём появления строк
                page.wait_for_timeout(300)
                rows = page.locator("tr[data-tid]")
                found_step = False
                
                for i in range(rows.count()):
                    r = rows.nth(i)
                    row_text = get_row_text(r)
                    row_clean = strip_count_suffix(normalize(row_text))
                    
                    if row_clean == step_clean:
                        if step_idx == len(path) - 1:
                            # Это последний шаг - город, заполняем его
                            if row_is_city(r):
                                if set_value_in_row(r):
                                    district_filled += 1
                                    filled_count[0] += 1
                                    done_cities.add(city_norm)
                                    filled_names.append(city)
                                    logger.info(f"✔ Заполнен: {city}")
                                    ui_status("✅", f"Заполнен: {city}", "green")
                                    if filled_count[0] % 50 == 0:
                                        ui_dashboard(filled_count[0], total_cities, district, filled_names)
                                else:
                                    ui_status("⚠️", f"Не удалось заполнить: {city}", "yellow")
                            else:
                                ui_status("⚠️", f"Строка не является городом: {city}", "yellow")
                        else:
                            # Это промежуточный шаг, раскрываем его
                            if row_is_expandable(r) and not row_is_expanded(r):
                                ui_status("🔽", f"Раскрываем: {row_text}", "grey")
                                click_expand_row(page, r)
                                page.wait_for_timeout(500)
                                wait_for_spinner(page)
                        found_step = True
                        break
                
                if not found_step:
                    ui_status("⚠️", f"Шаг пути не найден: '{step_name}'", "yellow")
                    path_opened = False
                    break

            if not path_opened:
                ui_status("⚠️", f"Не удалось раскрыть путь до: {city}", "yellow")
                continue

        ui_status("✅", f"Заполнено в '{district}': {district_filled}", "green")
        progress["done_cities"] = list(done_cities)
        save_progress(progress)
        
        # Сохраняем изменения и переходим к следующему округу
        save_page_changes(page)
        
        # Если все города в округе заполнены, сохраняем и переоткрываем настройки
        if district_filled == len(district_entries):
            ui_status("💾", f"Округ '{district}' полностью заполнен, сохраняем...", "cyan")
            page.reload(wait_until="domcontentloaded", timeout=60_000)
            page.wait_for_timeout(3_000)
            dismiss_overlays(page)
            open_drawer(page)
            ensure_root_expanded(page)

    ui_sep("ИТОГИ", "blue")
    ui_row(f"✅  Заполнено:     {filled_count[0]} из {total_cities}", "green")
    ui_row(f"❌  Не в карте:   {len(not_in_map)}", "red" if not_in_map else "grey")
    ui_row(f"📋  Все изменения сохранены автоматически", "cyan")
    if Path(PROGRESS_FILE).exists():
        Path(PROGRESS_FILE).unlink()


# ══════════════════════════════════════════════
#  РЕЖИМ 3: АНАЛИЗ
# ══════════════════════════════════════════════
def mode_analysis(page, cities: List[str], cabinet: str = "", warehouse: str = "") -> None:
    ui_sep("РЕЖИМ 3 — АНАЛИЗ + ИСПРАВЛЕНИЕ", "blue")
    ui_status("ℹ", "Анализирует все позиции, исправляет и создаёт Excel-отчёт", "cyan")
    ui_empty()

    import random
    city_norms = {normalize(c): c for c in cities}
    all_data = []

    # JS — собирает только листья-города, игнорирует:
    # - expandable строки (округа/области)
    # - чекбоксы дней недели (data-e2e="schedule-day-*")
    # - чекбоксы без name="checkbox-*" (не городские)
    _JS_COLLECT_ROWS = """
    () => {
        const rows = Array.from(document.querySelectorAll('tr[data-tid]'));
        const result = [];
        for (const row of rows) {
            // Пропускаем expandable (округа/области/районы)
            if (row.querySelector("button[aria-expanded]")) continue;

            // Название города — первая ячейка
            const td = row.querySelector('td');
            if (!td) continue;
            const name = td.textContent.trim().replace(/\\s+/g, ' ');
            if (!name || name.length < 2) continue;

            // Ищем ТОЛЬКО чекбокс города: name="checkbox-*" и БЕЗ data-e2e
            // Чекбоксы дней недели имеют data-e2e="schedule-day-MONDAY" и т.п.
            let cb = null;
            const allCbs = row.querySelectorAll("input[type='checkbox']");
            for (const c of allCbs) {
                const e2e = c.getAttribute('data-e2e') || '';
                const nm  = c.getAttribute('name') || '';
                // Пропускаем чекбоксы расписания
                if (e2e.startsWith('schedule-day')) continue;
                // Берём только checkbox-* (городские)
                if (nm.startsWith('checkbox-')) { cb = c; break; }
            }
            // Если нет городского чекбокса — пропускаем строку
            if (!cb) continue;

            const checked = cb.checked;
            const cbId    = cb.id;

            // Числовое поле доставки
            let value = '';
            const inputs = row.querySelectorAll("input[type='text'], input[type='number']");
            for (const inp of inputs) {
                if (!inp.offsetParent) continue;
                if (inp.readOnly || inp.disabled) continue;
                const ph = (inp.placeholder || '').toLowerCase();
                if (ph.includes('единые')) continue;
                value = inp.value || '';
                break;
            }

            result.push({ name, checked, cbId, value });
        }
        return result;
    }
    """

    # JS — снятие/установка чекбокса по id через прямое изменение + React event
    _JS_SET_CHECKBOX = """
    ([cbId, wantChecked]) => {
        const cb = document.getElementById(cbId);
        if (!cb) return false;
        if (cb.checked === wantChecked) return true;
        // Прямой клик через JS — обходит overlay
        cb.dispatchEvent(new MouseEvent('click', {bubbles: true, cancelable: true}));
        return true;
    }
    """

    # Открываем drawer один раз
    if not open_drawer(page):
        ui_status("❌", "Не удалось открыть drawer", "red")
        return
    ensure_root_expanded(page)


    for idx, district in enumerate(FEDERAL_DISTRICTS, 1):
        check_pause()
        ui_sep(f"{idx}/8  {district}", "cyan")

        expand_district_fully(page, district)
        page.wait_for_timeout(500)

        ui_status("��", "Читаем данные...", "grey")
        try:
            rows_data = page.evaluate(_JS_COLLECT_ROWS)
        except Exception as e:
            ui_status("⚠️", f"Ошибка чтения: {e}", "yellow")
            rows_data = []

        ui_status("📊", f"Найдено строк: {len(rows_data)}", "cyan")

        changed = 0
        for row_info in rows_data:
            city_name   = row_info["name"]
            is_checked  = row_info["checked"]
            cb_id       = row_info["cbId"]
            current_val = row_info["value"]
            city_norm   = normalize(city_name)
            in_excel    = city_norm in city_norms
            original    = city_norms.get(city_norm, city_name)
            action_taken = "—"

            # Если город в Excel и уже отмечен с правильным значением - пропускаем
            if in_excel and is_checked and current_val == INPUT_VALUE:
                action_taken = "✅ Уже ок"
            
            # Если город в Excel и не отмечен - отмечаем и заполняем
            elif in_excel and not is_checked:
                try:
                    if cb_id:
                        page.evaluate(_JS_SET_CHECKBOX, [cb_id, True])
                        page.wait_for_timeout(500)  # даём время появиться полю
                        
                        # Ждём, пока поле ввода станет доступно (до 5 секунд)
                        start = time.time()
                        while time.time() - start < 5:
                            r_loc = page.locator(f"tr[data-tid]:has(input[id='{cb_id}'])")
                            if r_loc.count() > 0:
                                # Пробуем заполнить – функция сама повторит попытки
                                if set_value_in_row(r_loc.first):
                                    action_taken = "✅ Отмечен + заполнен"
                                    is_checked = True
                                    current_val = INPUT_VALUE
                                    changed += 1
                                    break
                            page.wait_for_timeout(500)
                        else:
                            # Не дождались
                            action_taken = "⚠️ Поле не появилось, значение не записано"
                    else:
                        action_taken = "⚠️ Нет cb_id"
                except Exception as e:
                    action_taken = f"⚠️ Ошибка: {e}"

            # Если город НЕ в Excel но отмечен - убираем галочку (значение удалится само)
            elif not in_excel and is_checked:
                try:
                    if cb_id:
                        page.evaluate(_JS_SET_CHECKBOX, [cb_id, False])
                        page.wait_for_timeout(150)
                    action_taken = "🗑 Галочка снята"
                    is_checked = False
                    current_val = ""
                    changed += 1
                except Exception as e:
                    action_taken = f"⚠️ Ошибка снятия: {e}"

            # Если город в Excel, отмечен но значение неправильное - исправляем значение
            elif in_excel and is_checked and current_val != INPUT_VALUE:
                try:
                    if cb_id:
                        r_loc = page.locator(f"tr[data-tid]:has(input[id='{cb_id}'])")
                        if r_loc.count() > 0:
                            set_value_in_row(r_loc.first)
                    action_taken = "✏️ Значение исправлено"
                    current_val = INPUT_VALUE
                    changed += 1
                except Exception as e:
                    action_taken = f"⚠️ Ошибка: {e}"

            if in_excel and is_checked and current_val == INPUT_VALUE:
                status = "✅ Всё ок"
            elif in_excel and is_checked and current_val != INPUT_VALUE:
                status = "⚠️ Отмечен, значение ≠ 20"
            elif in_excel and not is_checked:
                status = "❌ Есть в Excel, не отмечен"
            elif not in_excel and is_checked:
                status = "🔴 Нет в Excel, но отмечен"
            else:
                status = "— Нет в Excel, не отмечен"

            all_data.append({
                "district":    district,
                "city_page":   city_name,
                "city_excel":  original if in_excel else "",
                "in_excel":    in_excel,
                "was_checked": is_checked,
                "value":       current_val,
                "status":      status,
                "action":      action_taken,
            })

        ui_status("✅", f"{district}: {len(rows_data)} позиций, изменено: {changed}", "green")

        blur_active_input(page)
        save_page_changes(page)

        if idx < len(FEDERAL_DISTRICTS):
            page.reload(wait_until="domcontentloaded", timeout=60_000)
            page.wait_for_timeout(3_000)
            open_drawer(page)
            ensure_root_expanded(page)

    # ── Excel-отчёт ──
    import openpyxl as _xl
    from openpyxl.styles import PatternFill as _PF, Font as _Fnt, Alignment as _Al, Border as _Br, Side as _Sd
    wb = _xl.Workbook()
    ws = wb.active
    ws.title = "Анализ и исправления"

    hdr_fill = _PF("solid", fgColor="2B2D42")
    hdr_font = _Fnt(bold=True, color="FFFFFF", size=11)
    ok_fill  = _PF("solid", fgColor="C8F7C5")
    fix_fill = _PF("solid", fgColor="FFF3CD")
    err_fill = _PF("solid", fgColor="FADADD")
    ext_fill = _PF("solid", fgColor="FFE0B2")
    no_fill  = _PF("solid", fgColor="F0F0F0")
    thin = _Br(left=_Sd(style="thin"), right=_Sd(style="thin"),
               top=_Sd(style="thin"),  bottom=_Sd(style="thin"))
    center = _Al(horizontal="center", vertical="center", wrap_text=True)
    left_a = _Al(horizontal="left", vertical="center")

    headers = ["№", "Федеральный округ", "Город (сайт)", "Город (Excel)",
               "Есть в Excel", "Отмечен", "Значение", "Статус", "Действие"]
    col_widths = [5, 32, 30, 30, 12, 10, 12, 30, 30]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = w
        cell = ws.cell(1, ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[1].height = 24

    for ri, d in enumerate(all_data, 2):
        s, a = d["status"], d["action"]
        if "Всё ок" in s and "Уже ок" in a:
            fill = ok_fill
        elif "исправлен" in a or "Отмечен +" in a:
            fill = fix_fill
        elif "🔴" in s:
            fill = ext_fill
        elif "❌" in s or "⚠️" in s:
            fill = err_fill
        else:
            fill = no_fill

        for ci, v in enumerate([
            ri-1, d["district"], d["city_page"], d["city_excel"],
            "Да" if d["in_excel"] else "Нет",
            "Да" if d["was_checked"] else "Нет",
            d["value"], d["status"], d["action"]
        ], 1):
            cell = ws.cell(ri, ci, value=v)
            cell.fill = fill
            cell.border = thin
            cell.alignment = left_a if ci in (2,3,4,8,9) else center

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(all_data)+1}"

    ws2 = wb.create_sheet("Сводка")
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 15
    total     = len(all_data)
    ok_cnt    = sum(1 for d in all_data if "Всё ок" in d["status"])
    fixed_cnt = sum(1 for d in all_data if d["action"] not in ("—","✅ Уже ок") and "Ошибка" not in d["action"])
    extra_cnt = sum(1 for d in all_data if not d["in_excel"] and d["was_checked"])
    miss_cnt  = sum(1 for d in all_data if d["in_excel"] and not d["was_checked"])
    err_cnt   = sum(1 for d in all_data if "Ошибка" in d["action"])
    ws2.cell(1,1,"Параметр").font = _Fnt(bold=True)
    ws2.cell(1,2,"Кол-во").font = _Fnt(bold=True)
    for i, (label, val) in enumerate([
        ("Всего позиций на сайте", total),
        ("✅ Всё ок (без изменений)", ok_cnt),
        ("✏️ Исправлено автоматически", fixed_cnt),
        ("🔴 Лишние (нет в Excel, сняты)", extra_cnt),
        ("❌ Отсутствовали в Excel", miss_cnt),
        ("⚠️ Ошибки при исправлении", err_cnt),
    ], 2):
        ws2.cell(i, 1, label)
        ws2.cell(i, 2, val)

    import random as _rnd
    rand2 = _rnd.randint(10, 99)
    date_str = datetime.now().strftime("%d%m%Y")
    cab_part = f"{cabinet}_{warehouse}".replace(" ", "_") if cabinet else ""
    fname = f"Анализ_{cab_part}_{rand2}_{date_str}.xlsx" if cab_part else f"Анализ_{rand2}_{date_str}.xlsx"
    wb.save(fname)

    ui_empty()
    ui_sep("ИТОГИ АНАЛИЗА", "green")
    ui_row(f"📊  Всего позиций:       {total}", "white")
    ui_row(f"✅  Уже были ок:         {ok_cnt}", "green")
    ui_row(f"✏️   Исправлено:          {fixed_cnt}", "cyan")
    ui_row(f"🔴  Лишних (снято):      {extra_cnt}", "yellow")
    ui_row(f"❌  Не было в Excel:     {miss_cnt}", "red")
    ui_row(f"⚠️   Ошибок:             {err_cnt}", "yellow")
    ui_empty()
    ui_status("💾", f"Отчёт сохранён: {fname}", "green")
    logger.info(f"Анализ сохранён: {fname}, строк: {total}")



# ══════════════════════════════════════════════
#  РЕЖИМ 4: СКАНЕР (построение карты)
# ══════════════════════════════════════════════
def mode_scanner(page) -> None:
    ui_sep("РЕЖИМ 4 — СКАНЕР (построение карты)", "blue")
    ui_status("ℹ", f"Строим карту путей и сохраняем в {MAP_FILE}", "cyan")
    ui_empty()

    full_map = load_map()
    progress = load_progress()
    scanned_districts: Set[str] = set(progress.get("scanned_districts", []))

    # Открываем drawer один раз
    open_drawer(page)
    ensure_root_expanded(page)

    for idx, district in enumerate(FEDERAL_DISTRICTS, 1):
        district_norm = normalize(district)
        if district_norm in scanned_districts:
            ui_status("⏩", f"Пропуск (уже отсканирован): {district}", "grey")
            continue

        check_pause()
        ui_sep(f"{idx}/8  {district}", "cyan")

        expanded = expand_district_fully(page, district)
        if expanded == 0:
            ui_status("⚠️", f"Не удалось раскрыть: {district}", "yellow")
            scanned_districts.add(district_norm)
            progress["scanned_districts"] = list(scanned_districts)
            save_progress(progress)
            page.reload(wait_until="domcontentloaded", timeout=60_000)
            page.wait_for_timeout(3_000)
            open_drawer(page)
            ensure_root_expanded(page)
            continue

        scroll_drawer_gently(page)
        page.wait_for_timeout(500)

        # Собираем города через JS — путь строится через DOM-родителей
        results = collect_cities_from_page(page)
        district_count = 0
        for item in results:
            city_text = (item.get("city") or "").strip()
            path = item.get("path") or []
            if not city_text or not path:
                continue
            key = normalize(city_text)
            if key not in full_map:
                full_map[key] = {
                    "city_original": city_text,
                    "path_original": path,
                    "path_norm":     [normalize(x) for x in path],
                    "path_display":  " / ".join(path),
                }
                district_count += 1

        ui_status("✅", f"{district}: +{district_count} городов (всего: {len(full_map)})", "green")
        logger.info(f"Сканер: '{district}' — {district_count} городов")

        scanned_districts.add(district_norm)
        progress["scanned_districts"] = list(scanned_districts)
        save_progress(progress)
        save_map(full_map)

        # Перезагружаем страницу для чистого DOM
        page.reload(wait_until="domcontentloaded", timeout=60_000)
        page.wait_for_timeout(3_000)
        dismiss_overlays(page)
        open_drawer(page)
        ensure_root_expanded(page)

    save_map(full_map)
    ui_empty()
    ui_sep("СКАНИРОВАНИЕ ЗАВЕРШЕНО", "blue")
    ui_row(f"🗺  Карта сохранена в {MAP_FILE}", "green")
    ui_row(f"📍 Всего городов в карте: {len(full_map)}", "green")
    if Path(PROGRESS_FILE).exists():
        Path(PROGRESS_FILE).unlink()
    ui_wait_enter("Нажмите Enter для завершения")


# ══════════════════════════════════════════════
#  ГЛАВНОЕ МЕНЮ
# ══════════════════════════════════════════════
def select_warehouse() -> Tuple[str, str, str]:
    """
    Двухшаговый выбор: сначала кабинет (МХ / ХМ / ТС),
    затем склад (Нижний / Москва / Смоленск / Пенза).
    Возвращает (url, cabinet, warehouse).
    """
    ui_sep("ШАГ 0 — ВЫБОР СКЛАДА", "blue")

    cabinets = list(WAREHOUSES.keys())
    cabinet = ui_ask("Выберите кабинет:", cabinets)

    warehouses = WAREHOUSES[cabinet]
    warehouse_names = list(warehouses.keys())
    warehouse = ui_ask(f"Кабинет «{cabinet}» — выберите склад:", warehouse_names)

    url = warehouses[warehouse]
    ui_status("✅", f"Выбран: {cabinet} / {warehouse}", "green")
    ui_status("🔗", url, "grey")
    logger.info(f"Выбран склад: {cabinet} / {warehouse} -> {url}")
    return url, cabinet, warehouse


def find_excel_for_warehouse(cabinet: str, warehouse: str) -> Optional[str]:
    """
    Ищет Excel файл для указанного склада.
    Возвращает путь к файлу или None если не найден.
    """
    # Сначала ищем в основной папке Склады
    excel_dir = Path(EXCEL_DIR)
    patterns = [
        f"{warehouse}.xlsx",
        f"{cabinet}_{warehouse}.xlsx",
        f"{cabinet}-{warehouse}.xlsx"
    ]
    
    for pattern in patterns:
        file_path = excel_dir / pattern
        if file_path.exists():
            return str(file_path)
    
    # Затем ищем в подпапках кабинетов
    cabinet_dir = excel_dir / cabinet
    if cabinet_dir.exists():
        for pattern in patterns:
            file_path = cabinet_dir / pattern
            if file_path.exists():
                return str(file_path)
        
        # Ищем любой Excel файл в папке кабинета
        for file_path in cabinet_dir.glob("*.xlsx"):
            return str(file_path)
    
    return None


def select_multi_warehouses() -> List[Tuple[str, str, str]]:
    """
    Выбор нескольких складов для параллельного запуска.
    Возвращает список (url, cabinet, warehouse).
    """
    ui_sep("ВЫБОР СКЛАДОВ (ПАРАЛЛЕЛЬНЫЙ РЕЖИМ)", "blue")

    # Строим плоский список всех складов
    all_warehouses: List[Tuple[str, str, str]] = []
    for cab, whs in WAREHOUSES.items():
        for wh, url in whs.items():
            all_warehouses.append((cab, wh, url))

    ui_empty()
    ui_row("Введите номера складов через пробел (например: 1 3 5):", "yellow")
    for i, (cab, wh, _) in enumerate(all_warehouses, 1):
        ui_row(f"  {i:2d}.  {cab} — {wh}", "white")
    ui_empty()

    while True:
        try:
            raw = input(_c("cyan", "║ ") + _c("yellow", "▶ Номера: ")).strip()
            parts = raw.split()
            selected = []
            valid = True
            for p in parts:
                if p.isdigit() and 1 <= int(p) <= len(all_warehouses):
                    entry = all_warehouses[int(p) - 1]
                    if entry not in selected:
                        selected.append(entry)
                else:
                    ui_row(f"⚠  Неверный номер: {p}", "yellow")
                    valid = False
                    break
            if valid and selected:
                ui_empty()
                ui_row("Выбраны склады:", "green")
                for cab, wh, _ in selected:
                    ui_row(f"  ✔  {cab} — {wh}", "green")
                return selected
            ui_row("⚠  Введите хотя бы один корректный номер", "yellow")
        except (KeyboardInterrupt, EOFError):
            sys.exit(0)


def run_parallel(selected: List[Tuple[str, str, str]], mode: str) -> None:
    """
    Запускает несколько экземпляров script1.py параллельно.
    selected - список (cabinet, warehouse, url)
    """
    import subprocess
    import tempfile
    import json
    
    ui_sep("ЗАПУСК ПАРАЛЛЕЛЬНЫХ ПРОЦЕССОВ", "blue")
    ui_status("ℹ", f"Будет запущено {len(selected)} экземпляров", "cyan")
    ui_empty()

    # Создаем временные файлы конфигурации для каждого экземпляра
    temp_files = []
    for i, (cabinet, warehouse, url) in enumerate(selected):
        # Загружаем города для этого склада
        cities: List[str] = []
        if "Сканер" not in mode:
            excel_path = find_excel_for_warehouse(cabinet, warehouse)
            if excel_path:
                cities = load_cities(excel_path)
                ui_status("📋", f"{cabinet}/{warehouse}: {len(cities)} городов", "grey")
            else:
                ui_status("⚠️", f"{cabinet}/{warehouse}: Excel файл не найден", "yellow")
        
        # Создаем временный файл конфигурации
        config = {
            "YD_START_URL": url,
            "YD_CABINET": cabinet,
            "YD_WAREHOUSE": warehouse,
            "YD_MODE": mode,
            "YD_CITIES": cities,
            "YD_PROFILE_DIR": f"profiles/browser_profile_{cabinet}_{warehouse}".replace(" ", "_"),
            "YD_INSTANCE_NAME": f"{cabinet}-{warehouse}",
        }
        
        # Создаем временный файл конфигурации (не удаляем автоматически)
        temp_dir = tempfile.mkdtemp()
        temp_file = os.path.join(temp_dir, f'config_{i}.json')
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        temp_files.append(temp_file)
    
    ui_empty()
    ui_status("🚀", "Запускаем экземпляры в отдельных терминалах...", "cyan")
    ui_empty()
    
    # Запускаем каждый экземпляр в отдельном процессе
    procs = []
    for i, (cabinet, warehouse, _) in enumerate(selected):
        name = f"{cabinet}-{warehouse}"
        temp_file = temp_files[i]
        
        # Запускаем в новом терминале для ручной авторизации
        if sys.platform == "darwin":  # macOS
            script_content = f'''#!/bin/bash
cd "{os.getcwd()}"
echo "=== Терминал для {name} ==="
echo "Авторизуйтесь в браузере и нажмите Enter для продолжения"
"{sys.executable}" "{__file__}" --instance "{temp_file}"
echo "Нажмите Enter для закрытия..."
read
'''
            # Создаем временный скрипт и запускаем через osascript с ожиданием
            script_path = f"/tmp/terminal_{name}.sh"
            with open(script_path, 'w') as f:
                f.write(script_content)
            os.chmod(script_path, 0o755)
            # Используем osascript для точного контроля над терминалами
            cmd = [
                "osascript", "-e",
                f'tell application "Terminal" to do script "exec \\"{script_path}\\""'
            ]
        elif sys.platform == "win32":  # Windows
            cmd = ["start", "cmd", "/k", sys.executable, __file__, "--instance", temp_file]
        else:  # Linux
            cmd = ["gnome-terminal", "--", sys.executable, __file__, "--instance", temp_file]
        
        ui_status(f"▶", f"Запускаем терминал для: {name}", "cyan")
        try:
            proc = subprocess.Popen(cmd, cwd=os.getcwd(), 
                                 stdout=subprocess.DEVNULL, 
                                 stderr=subprocess.DEVNULL)
            procs.append(proc)
            time.sleep(3)  # Задержка между запусками терминалов
        except Exception as e:
            ui_status("❌", f"Ошибка запуска {name}: {e}", "red")
            continue
    
    ui_empty()
    ui_status("✅", f"Запущено {len(procs)} экземпляров в отдельных терминалах", "green")
    ui_status("ℹ", "Ожидаем завершения всех процессов...", "cyan")
    ui_status("ℹ", "Авторизуйтесь в каждом браузере и нажмите Enter", "yellow")
    ui_empty()
    
    # Ожидаем завершения всех процессов с проверкой статуса
    try:
        running_procs = procs.copy()
        while running_procs:
            time.sleep(2)  # Проверяем каждые 2 секунды
            still_running = []
            
            for proc in running_procs:
                if proc.poll() is None:  # Процесс еще работает
                    still_running.append(proc)
                else:  # Процесс завершился
                    if proc.returncode != 0:
                        ui_status("❌", f"Процесс завершился с ошибкой (код {proc.returncode})", "red")
                    else:
                        ui_status("✅", f"Процесс завершился успешно", "green")
            
            running_procs = still_running
            
            if running_procs:
                ui_status("⏳", f"Работает {len(running_procs)} процессов...", "cyan")
        
        ui_status("🎉", "Все процессы завершены!", "green")
    except KeyboardInterrupt:
        ui_status("⚠️", "Прерывание процессов...", "yellow")
        for proc in procs:
            proc.terminate()
    
    # Удаляем временные файлы
    for temp_file in temp_files:
        try:
            Path(temp_file).unlink(missing_ok=True)
        except Exception:
            pass


def run_single_instance(url: str, cabinet: str, warehouse: str, mode: str, cities: List[str], instance_name: str = "main", debug_port: int = None) -> None:
    """
    Запускает один экземпляр автоматизации для указанного склада и режима.
    """
    # Используем профиль из переменной окружения, установленной в main()
    profile_dir = os.environ.get("YD_PROFILE_DIR", PROFILE_DIR)
    
    # Для одиночного режима используем профиль с данными авторизации
    if not profile_dir or profile_dir == PROFILE_DIR:
        profile_dir = f"profiles/{cabinet}_{warehouse}".replace(" ", "_")
        logger.info(f"Using profile with auth data: {profile_dir}")
    
    try:
        ui_status("🔧", f"Запуск браузера с профилем: {profile_dir}", "cyan")
        pw, ctx, page = launch_browser(profile_dir)
        ui_status("✅", "Браузер успешно запущен", "green")
    except Exception as e:
        ui_status("❌", f"Ошибка запуска браузера: {e}", "red")
        logger.error(f"Browser launch failed: {e}")
        return
    try:
        goto_settings(page, url, force_manual_auth=True)
        
        if "С картой" in mode:
            mode_with_map(page, cities)
        elif "Без карты" in mode:
            mode_without_map(page, cities)
        elif "Анализ" in mode:
            mode_analysis(page, cities, cabinet, warehouse)
        elif "Сканер" in mode:
            mode_scanner(page)
        else:
            ui_status("❌", f"Неизвестный режим: {mode}", "red")
            return
            
    except Exception as e:
        logger.error(f"Ошибка в run_single_instance: {e}")
        ui_status("❌", f"Ошибка: {e}", "red")
    finally:
        ctx.close()
        pw.stop()


def main() -> None:
    # ── Дочерний процесс (параллельный режим) ──
    if "--instance" in sys.argv:
        idx = sys.argv.index("--instance")
        tmp_file = sys.argv[idx + 1]
        print(f"� Чтение конфига из: {tmp_file}")
        
        try:
            with open(tmp_file, encoding="utf-8") as f:
                data = json.load(f)
            # Удаляем временный файл
            Path(tmp_file).unlink(missing_ok=True)
            print("✅ Конфиг успешно загружен")
        except Exception as e:
            print(f"❌ Ошибка чтения конфига: {e}")
            sys.exit(1)

        url       = data.get("YD_START_URL", "")
        cabinet   = data.get("YD_CABINET", "")
        warehouse = data.get("YD_WAREHOUSE", "")
        mode      = data.get("YD_MODE", "")
        cities    = data.get("YD_CITIES", [])
        profile   = data.get("YD_PROFILE_DIR", "browser_profile")
        name      = data.get("YD_INSTANCE_NAME", "main")
        debug_port = data.get("YD_DEBUG_PORT", None)

        print(f"🎯 Параметры: {cabinet}/{warehouse} - {mode}")
        print(f"🌐 URL: {url}")
        print(f"👥 Города: {len(cities)}")
        print(f"📁 Профиль: {profile}")

        # Устанавливаем переменные для логов/прогресса
        os.environ["YD_PROFILE_DIR"]   = profile
        os.environ["YD_INSTANCE_NAME"] = name

        print("🚀 Запуск run_single_instance...")
        run_single_instance(url, cabinet, warehouse, mode, cities, instance_name=name, debug_port=debug_port)
        print("✅ run_single_instance завершен")
        return

    os.system("cls" if os.name == "nt" else "clear")
    ui_header()

    # Создаём нужные папки при старте
    for d in [EXCEL_DIR, REPORTS_DIR, MAPS_DIR]:
        Path(d).mkdir(exist_ok=True)

    # ── Выбор: единичный или параллельный ──
    ui_sep("ШАГ 0 — РЕЖИМ ЗАПУСКА", "blue")
    launch_mode = ui_ask(
        "Как запустить?",
        ["Единичный — один склад", "Параллельный — несколько складов одновременно"]
    )

    # ── Выбор режима работы ──
    ui_sep("ШАГ 1 — РЕЖИМ РАБОТЫ", "blue")
    mode = ui_ask(
        "Выберите режим:",
        [
            "С картой     — быстрый поиск по karta.json",
            "Без карты    — пошагово по округам + заполнение 20",
            "Анализ       — сбор данных + Excel-отчёт",
            "Сканер       — построить карту karta.json",
        ]
    )

    if "Параллельный" in launch_mode:
        # ── Параллельный режим ──
        selected = select_multi_warehouses()
        run_parallel(selected, mode)
        ui_bot()
        return

    # ── Единичный режим ──
    start_url, cabinet, warehouse = select_warehouse()

    cities: List[str] = []
    if "Сканер" not in mode:
        ui_sep("ШАГ 2 — ФАЙЛ С ГОРОДАМИ", "blue")
        # Сначала пробуем найти автоматически
        excel_path = find_excel_for_warehouse(cabinet, warehouse)
        if excel_path:
            ui_status("📄", f"Найден файл: {Path(excel_path).name}", "green")
        else:
            ui_status("⚠️", f"Файл для {cabinet}/{warehouse} не найден автоматически", "yellow")
            excel_path = select_excel_file()
        cities = load_cities(excel_path)
        ui_status("📋", f"Загружено {len(cities)} уникальных городов", "green")
        ui_empty()

    logger.info("Starting progress check...")
    progress = load_progress()
    logger.info(f"Progress loaded: {bool(progress)}")
    if progress and ui_ask_yn("Найден сохранённый прогресс. Продолжить с него?"):
        ui_status("ℹ", "Продолжаем с сохранённого прогресса", "cyan")
    else:
        save_progress({})
        ui_status("ℹ", "Начинаем с начала", "cyan")

    logger.info("About to call run_single_instance...")
    run_single_instance(start_url, cabinet, warehouse, mode, cities)
    logger.info("run_single_instance completed")
    ui_bot()


if __name__ == "__main__":
    main()