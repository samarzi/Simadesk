#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Автоматизация тарифов на доставку по зонам
# Заполнение тарифов на основе Excel-файлов с зонами
#
# Структура страницы «Тарифы на доставку» (Яндекс Маркет):
#   Кнопка «Настроить» → data-e2e="levitan-button tariff-setting-setup-button"
#   Drawer → data-e2e="levitan-drawer-wrapper"
#   Таблица регионов → data-e2e="levitan-table"
#   Строки → tr[data-e2e="levitan-table-row"] с классами __use--indent_N___
#   Чекбоксы → span[data-e2e^="levitan-clickable checkbox-"]
#   Кнопка сохранения → data-e2e="levitan-button tariff-setting-save-button"

import os
import re
import sys
import json
import time
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict, Set, Tuple

# ──────────────────────────────────────────────
# Зависимости
# ──────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("❌ pip install openpyxl")
    sys.exit(1)

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    print("❌ pip install playwright && playwright install chromium")
    sys.exit(1)


# ══════════════════════════════════════════════
#  КОНСТАНТЫ
# ══════════════════════════════════════════════
MAP_FILE       = "../Карты/karta.json"
PROFILE_DIR    = os.environ.get("YD_PROFILE_DIR", "../profiles/tariffs_profile")
LOG_FILE       = f"tariffs_{os.environ.get('YD_INSTANCE_NAME', 'main')}.log"
PAUSE_FILE     = "pause.txt"
PROGRESS_FILE  = f"progress_tariffs_{os.environ.get('YD_INSTANCE_NAME', 'main')}.json"
INPUT_VALUE    = "20"
EXCEL_DIR      = "../Склады"
CITIES_DIR     = "../Склады города"
REPORTS_DIR    = "../Отчёты"
MAPS_DIR       = "../Карты"

SKIPPED_CITIES = []

# Тайминги (мс)
AFTER_CLICK_MS    = 150
AFTER_SPINNER_MS  = 100
AFTER_SAVE_MS     = 5_000
SCROLL_STEP_PX    = 500
SCROLL_PAUSE_MS   = 120

# ──────────────────────────────────────────────
#  СКЛАДЫ — кабинеты и URL страниц доставки
# ──────────────────────────────────────────────
WAREHOUSES = {
    "МХ": {
        "Нижний Новгород": "https://partner.market.yandex.ru/shop/149090347/delivery?warehouseId=2241450",
        "Москва":          "https://partner.market.yandex.ru/shop/149088763/delivery?warehouseId=2239087",
        "Смоленск":        "https://partner.market.yandex.ru/shop/149087710/delivery?warehouseId=2237337",
        "Пенза":           "https://partner.market.yandex.ru/shop/149085812/delivery?warehouseId=2234713",
    },
    "ХМ": {
        "Нижний Новгород": "https://partner.market.yandex.ru/shop/149090350/delivery?warehouseId=2241460",
        "Москва":          "https://partner.market.yandex.ru/shop/149088932/delivery?warehouseId=2239317",
        "Смоленск":        "https://partner.market.yandex.ru/shop/149088926/delivery?warehouseId=2239315",
        "Пенза":           "https://partner.market.yandex.ru/shop/149088920/delivery?warehouseId=2239309",
    },
    "ТС": {
        "Нижний Новгород": "https://partner.market.yandex.ru/shop/149091241/delivery?warehouseId=2243354",
        "Москва":          "https://partner.market.yandex.ru/shop/149091238/delivery?warehouseId=2243351",
        "Смоленск":        "https://partner.market.yandex.ru/shop/149091234/delivery?warehouseId=2243347",
        "Пенза":           "https://partner.market.yandex.ru/shop/149091231/delivery?warehouseId=2243346",
    },
}

# Настройка логгера
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)-7s] %(message)s",
    handlers=[logging.FileHandler(LOG_FILE, encoding="utf-8")],
)
logger = logging.getLogger("TARIFFS")

FEDERAL_DISTRICTS: List[str] = [
    "Центральный федеральный округ",
    "Северо-Западный федеральный округ",
    "Южный федеральный округ",
    "Северо-Кавказский федеральный округ",
    "Приволжский федеральный округ",
    "Уральский федеральный округ",
    "Сибирский федеральный округ",
    "Дальневосточный федеральный округ",
]


# ══════════════════════════════════════════════
#  UI — ТЕРМИНАЛЬНЫЙ ИНТЕРФЕЙС
# ══════════════════════════════════════════════
W = 66

def _c(code: str, text: str) -> str:
    codes = {
        "red": "\033[91m", "green": "\033[92m", "yellow": "\033[93m",
        "blue": "\033[94m", "cyan": "\033[96m", "white": "\033[97m",
        "grey": "\033[90m", "bold": "\033[1m", "reset": "\033[0m",
    }
    return f"{codes.get(code, '')}{text}{codes['reset']}"

def ui_top():
    print(_c("cyan", "╔" + "═" * W + "╗"))

def ui_bot():
    print(_c("cyan", "╚" + "═" * W + "╝"))

def ui_sep(label: str = "", color: str = "cyan"):
    if label:
        inner = f"  {label}  "
        pad = (W - len(inner)) // 2
        line = "═" * pad + inner + "═" * (W - pad - len(inner))
    else:
        line = "═" * W
    print(_c(color, "╠" + line + "╣"))

def ui_row(text: str = "", color: str = "white"):
    vis = re.sub(r'\033\[[^m]*m', '', text)
    pad = W - len(vis) - 1
    print(_c("cyan", "║ ") + _c(color, text) + " " * max(0, pad) + _c("cyan", "║"))

def ui_empty():
    print(_c("cyan", "║") + " " * W + _c("cyan", "║"))

def ui_header():
    ui_top()
    ui_empty()
    title = "ТАРИФЫ НА ДОСТАВКУ — АВТОЗАПОЛНЕНИЕ ПО ЗОНАМ"
    p = (W - len(title)) // 2
    print(_c("cyan", "║") + " " * p + _c("bold", _c("yellow", title)) + " " * (W - p - len(title)) + _c("cyan", "║"))
    ui_empty()

def ui_status(icon: str, text: str, color: str = "white"):
    vis = len(icon + "  " + text)
    spaces = W - vis - 1
    msg = f"{icon}  {text}"
    print(_c("cyan", "║ ") + _c(color, msg) + " " * max(0, spaces) + _c("cyan", "║"))
    logger.info(f"[UI] {icon} {text}")

def ui_wait_enter(msg: str = "Нажмите Enter для продолжения...") -> None:
    ui_empty()
    ui_row(msg, "yellow")
    ui_empty()
    try:
        input(_c("cyan", "║ ") + _c("grey", "[ Enter ] "))
    except (KeyboardInterrupt, EOFError):
        sys.exit(0)

def ui_ask(question: str, options: List[str]) -> str:
    ui_sep(question, "blue")
    for i, opt in enumerate(options, 1):
        ui_row(f"  {i}.  {opt}", "white")
    ui_empty()
    while True:
        try:
            raw = input(_c("cyan", "║ ") + _c("yellow", "▶ Выберите: ")).strip()
            if raw.isdigit() and 1 <= int(raw) <= len(options):
                return options[int(raw) - 1]
        except (KeyboardInterrupt, EOFError):
            sys.exit(0)
        ui_row("⚠  Неверный выбор, попробуйте ещё раз", "yellow")

def ui_ask_yn(question: str) -> bool:
    ui_row(question, "yellow")
    try:
        ans = input(_c("cyan", "║ ") + _c("grey", "[ y/n ] ")).strip().lower()
        return ans in ("y", "yes", "да", "д")
    except (KeyboardInterrupt, EOFError):
        sys.exit(0)


# ══════════════════════════════════════════════
#  ВСПОМОГАТЕЛЬНЫЕ УТИЛИТЫ
# ══════════════════════════════════════════════
def normalize(text: str) -> str:
    if not text:
        return ""
    t = text.lower().strip()
    t = t.replace("ё", "е")
    t = re.sub(r"[•·\t\r\n]+", " ", t)
    t = re.sub(r"\bг\.?\s*", "", t)
    t = re.sub(r"\bгород\b", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t

def strip_count_suffix(s: str) -> str:
    s = re.sub(r'[•·]\s*\d+\s*$', '', s)
    s = re.sub(r'\s+[•·]\s*\d+\s*$', '', s)
    s = re.sub(r'\s+\d+\s*$', '', s)
    return s.strip()

def check_pause():
    if Path(PAUSE_FILE).exists():
        ui_status("⏸", f"ПАУЗА. Удалите файл {PAUSE_FILE} для продолжения.", "yellow")
        while Path(PAUSE_FILE).exists():
            time.sleep(1)
        ui_status("▶", "Продолжаем...", "green")

def load_map() -> Dict:
    if not Path(MAP_FILE).exists():
        return {}
    try:
        with open(MAP_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Ошибка загрузки карты: {e}")
        return {}

def save_map(data: Dict) -> None:
    Path(MAPS_DIR).mkdir(exist_ok=True)
    with open(MAP_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    logger.info(f"Карта сохранена: {len(data)} записей")


# ══════════════════════════════════════════════
#  РАБОТА С DOM — ОБЩИЕ ФУНКЦИИ ДЛЯ ДЕРЕВА РЕГИОНОВ
#  (Используют levitan-table селекторы)
# ══════════════════════════════════════════════
def get_row_text(row) -> str:
    try:
        td = row.locator("td").first
        if td.count() > 0:
            t = (td.text_content() or "").strip()
        else:
            t = (row.text_content() or "").strip()
        return re.sub(r"\s+", " ", t).strip()
    except Exception:
        return ""

def get_row_clean_name(row) -> str:
    return strip_count_suffix(normalize(get_row_text(row)))

def row_is_expandable(row) -> bool:
    try:
        return row.locator("button[aria-expanded]").count() > 0
    except Exception:
        return False

def row_is_expanded(row) -> bool:
    try:
        btn = row.locator("button[aria-expanded]").first
        return btn.get_attribute("aria-expanded") == "true"
    except Exception:
        return False

def row_is_city(row) -> bool:
    return not row_is_expandable(row)

def click_expand_row(page, row, wait_spinner: bool = False) -> bool:
    try:
        btn = row.locator("button[aria-expanded='false']").first
        if btn.count() == 0:
            return False
        scroll_row_in_drawer(page, row)
        btn.click(timeout=3_000)
        page.wait_for_timeout(AFTER_CLICK_MS)
        if wait_spinner:
            wait_for_spinner(page)
        return True
    except Exception as e:
        logger.debug(f"click_expand_row error: {e}")
        return False


def scroll_row_in_drawer(page, row) -> None:
    """Скроллит строку в видимую область ВНУТРИ drawer-контейнера."""
    try:
        page.evaluate("""
            (el) => {
                const drawer = el.closest('[data-e2e="levitan-drawer-content"]')
                    || el.closest('[data-e2e="levitan-drawer-wrapper"]')
                    || el.closest('.___content___4Arvd');
                if (drawer) {
                    const rect = el.getBoundingClientRect();
                    const dRect = drawer.getBoundingClientRect();
                    if (rect.bottom > dRect.bottom || rect.top < dRect.top) {
                        el.scrollIntoView({block: 'center', behavior: 'instant'});
                    }
                }
            }
        """, row.element_handle())
    except Exception:
        pass


def wait_for_spinner(page, timeout_ms: int = 10_000) -> None:
    """Ждёт исчезновения спиннера. Если спиннер не появился — выходит мгновенно."""
    try:
        # Проверяем через JS — спиннер видим прямо сейчас?
        visible = page.evaluate("""
            () => {
                const sp = document.querySelector("[data-e2e='levitan-preloader'] [aria-hidden='false']")
                    || document.querySelector("div[class*='spinner']:not([style*='display: none'])")
                    || document.querySelector("div[class*='Spinner']:not([style*='display: none'])");
                if (!sp) return false;
                const rect = sp.getBoundingClientRect();
                return rect.width > 0 && rect.height > 0;
            }
        """)
        if not visible:
            return  # Спиннера нет — выходим мгновенно
        # Спиннер виден — ждём пока скроется
        spinner_sel = "[data-e2e='levitan-preloader'], div[class*='spinner'], div[class*='Spinner']"
        page.locator(spinner_sel).first.wait_for(state="hidden", timeout=timeout_ms)
        page.wait_for_timeout(AFTER_SPINNER_MS)
    except (PWTimeout, Exception):
        pass

def find_row_by_text(page, target_text: str, table=None, expandable_only: bool = False):
    target_norm = strip_count_suffix(normalize(target_text))
    if table:
        rows = table.locator("tr[data-e2e='levitan-table-row']")
    else:
        rows = page.locator("tr[data-e2e='levitan-table-row']")
    n = rows.count()

    for i in range(n):
        r = rows.nth(i)
        if expandable_only and not row_is_expandable(r):
            continue
        row_norm = get_row_clean_name(r)
        if row_norm == target_norm:
            return r

    for i in range(n):
        r = rows.nth(i)
        if expandable_only and not row_is_expandable(r):
            continue
        row_norm = get_row_clean_name(r)
        if target_norm in row_norm or row_norm in target_norm:
            return r

    return None

def find_district_row(page, district_name: str, table=None):
    district_norm = normalize(district_name)
    if table:
        rows = table.locator("tr[data-e2e='levitan-table-row']")
    else:
        rows = page.locator("tr[data-e2e='levitan-table-row']")
    n = rows.count()

    logger.info(f"Ищем округ '{district_name}' среди {n} строк")
    ui_status("🔍", f"Ищем '{district_name}' среди {n} строк...", "grey")

    for i in range(n):
        r = rows.nth(i)
        if not row_is_expandable(r):
            continue
        txt = normalize(get_row_text(r))
        if txt == district_norm:
            ui_status("✅", f"Найден точный округ: {get_row_text(r)}", "green")
            return r

    for i in range(n):
        r = rows.nth(i)
        if not row_is_expandable(r):
            continue
        txt = normalize(get_row_text(r))
        if district_norm in txt or txt in district_norm:
            ui_status("✅", f"Найден частичный округ: {get_row_text(r)}", "green")
            return r

    ui_status("❌", f"Округ '{district_name}' НЕ найден (строк: {n})", "red")
    # Диагностика — показываем что реально видно в дереве
    sample_expandable = []
    sample_all = []
    for i in range(min(15, n)):
        r = rows.nth(i)
        txt = get_row_text(r)
        if row_is_expandable(r):
            sample_expandable.append(txt)
        if i < 5:
            sample_all.append(f"[{'E' if row_is_expandable(r) else 'L'}] {txt}")
    if sample_expandable:
        logger.warning(f"Доступные expandable строки ({len(sample_expandable)}): {sample_expandable}")
        ui_status("📋", f"Видимые раскрываемые: {', '.join(sample_expandable[:5])}", "grey")
    else:
        logger.warning(f"НЕТ expandable строк! Первые строки: {sample_all}")
        ui_status("📋", f"Нет раскрываемых строк! Все строки: {'; '.join(sample_all)}", "grey")
    return None


# ══════════════════════════════════════════════
#  АВТОРИЗАЦИЯ
# ══════════════════════════════════════════════
def wait_for_manual_auth(page) -> None:
    ui_sep("ОЖИДАНИЕ АВТОРИЗАЦИИ", "yellow")
    ui_row("🌐 Браузер открыт для авторизации", "cyan")
    ui_row("📝 Вручную авторизуйтесь в Яндекс Маркете", "yellow")
    ui_row("⏳ После успешной авторизации нажмите Enter", "green")
    ui_empty()
    ui_wait_enter("Нажмите Enter после завершения авторизации...")
    ui_sep()

    current_url = page.url
    if "passport" in current_url.lower() or "auth" in current_url.lower() or "login" in current_url.lower():
        ui_status("⚠️", "Вы всё ещё на странице авторизации!", "red")
        if not ui_ask_yn("Продолжить без авторизации?"):
            return wait_for_manual_auth(page)
    else:
        ui_status("✅", "Авторизация успешно пройдена", "green")

def is_already_authenticated(page) -> bool:
    try:
        url = page.url
        if any(x in url for x in ("passport.yandex", "auth", "login")):
            return False
        selectors = [
            "[data-e2e='user-menu']",
            "[data-e2e='logout']",
            "[class*='UserMenu']",
            "[class*='user-name']",
        ]
        for sel in selectors:
            if page.locator(sel).count() > 0:
                return True
        if "partner.market.yandex.ru" in url:
            page.wait_for_timeout(2_000)
            if page.locator("[data-e2e='dbs-delivery-page']").count() > 0:
                return True
            if page.locator("[data-e2e='tariff-setting-card']").count() > 0:
                return True
    except Exception as e:
        logger.debug(f"is_already_authenticated error: {e}")
    return False

def goto_settings(page, url: str) -> None:
    ui_status("🌐", f"Открываем: {url}", "cyan")
    page.goto(url, wait_until="domcontentloaded", timeout=60_000)
    page.wait_for_timeout(3_000)

    if is_already_authenticated(page):
        ui_status("✅", "Профиль авторизован", "green")
        return

    ui_status("⏸", "Требуется авторизация...", "yellow")
    wait_for_manual_auth(page)

    ui_status("🔄", "Переходим на страницу настроек...", "cyan")
    page.goto(url, wait_until="domcontentloaded", timeout=60_000)
    page.wait_for_timeout(3_000)


# ══════════════════════════════════════════════
#  EXCEL — ЗАГРУЗКА ДАННЫХ
# ══════════════════════════════════════════════
def validate_excel_format(filepath: str) -> bool:
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        wb.close()
        zone_headers = [h for h in headers if "зона" in h.lower()]
        if not zone_headers:
            return False
        logger.info(f"Файл валиден: {len(zone_headers)} зон в {filepath}")
        return True
    except Exception as e:
        logger.error(f"Ошибка валидации файла: {e}")
        return False

def find_excel_for_warehouse(cabinet: str, warehouse: str, mode: str) -> Optional[str]:
    dir_path = CITIES_DIR if mode == "cities" else EXCEL_DIR
    excel_dir = Path(dir_path)
    patterns = [
        f"{warehouse}.xlsx",
        f"{cabinet}_{warehouse}.xlsx",
        f"{cabinet}-{warehouse}.xlsx"
    ]

    logger.info(f"Поиск Excel файла для {cabinet}/{warehouse} в {dir_path}")

    for pattern in patterns:
        file_path = excel_dir / pattern
        if file_path.exists():
            logger.info(f"Найден файл: {file_path}")
            return str(file_path)

    cabinet_dir = excel_dir / cabinet
    if cabinet_dir.exists():
        for pattern in patterns:
            file_path = cabinet_dir / pattern
            if file_path.exists():
                return str(file_path)

    logger.warning(f"Excel файл не найден для {cabinet}/{warehouse} в {dir_path}")
    return None

def select_excel_file(mode: str, cabinet: str = None, warehouse: str = None) -> str:
    dir_path = CITIES_DIR if mode == "cities" else EXCEL_DIR
    excel_dir = Path(dir_path)

    if cabinet and warehouse:
        auto_file = find_excel_for_warehouse(cabinet, warehouse, mode)
        if auto_file:
            ui_status("📄", f"Автоматически выбран файл: {Path(auto_file).name}", "green")
            return auto_file

    files = sorted(excel_dir.glob("*.xlsx"))
    if not files:
        ui_status("❌", f"Нет .xlsx файлов в папке {dir_path}!", "red")
        ui_bot()
        sys.exit(1)

    valid_files = [f for f in files if validate_excel_format(str(f))]
    if not valid_files:
        ui_status("❌", f"Нет файлов с колонками зон в {dir_path}!", "red")
        ui_bot()
        sys.exit(1)

    if len(valid_files) == 1:
        ui_status("📄", f"Файл: {valid_files[0].name}", "green")
        return str(valid_files[0])

    names = [f.name for f in valid_files]
    print(_c("cyan", "║ ") + _c("yellow", "Выберите Excel-файл:"))
    for i, name in enumerate(names, 1):
        ui_row(f"  {i}.  {name}", "white")
    ui_empty()
    while True:
        try:
            raw = input(_c("cyan", "║ ") + _c("yellow", "▶ Ваш выбор: ")).strip()
            if raw.isdigit() and 1 <= int(raw) <= len(names):
                return str(valid_files[int(raw) - 1])
        except (KeyboardInterrupt, EOFError):
            sys.exit(0)

def load_cities_data(excel_path: str) -> Dict[str, List[str]]:
    logger.info(f"Загрузка городов из файла: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    cities_data = {}
    zone_columns = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header and isinstance(header, str) and "зона" in header.lower():
            zone_columns[header.strip()] = col

    if not zone_columns:
        logger.warning(f"В файле {excel_path} не найдены колонки зон")
        return cities_data

    logger.info(f"Найдены зоны: {list(zone_columns.keys())}")

    for zone_name, col in zone_columns.items():
        cities = []
        for row in range(2, ws.max_row + 1):
            city = ws.cell(row, col).value
            if city and isinstance(city, str) and city.strip():
                cities.append(city.strip())
        if cities:
            cities_data[zone_name] = cities
            logger.info(f"  {zone_name}: {len(cities)} городов (первые 3: {cities[:3]})")

    wb.close()
    logger.info(f"Загружены города для {len(cities_data)} зон")
    return cities_data

def select_mode() -> str:
    ui_sep("ВЫБОР РЕЖИМА", "yellow")
    ui_row("  1.  Заполнение городов по зонам (чекбоксы)", "white")
    ui_empty()
    while True:
        try:
            raw = input(_c("cyan", "║ ") + _c("yellow", "▶ Выберите режим: ")).strip()
            if raw == "1":
                return "cities"
        except (KeyboardInterrupt, EOFError):
            sys.exit(0)

def select_zones(all_zones: Dict, mode: str) -> Dict:
    ui_sep("ВЫБОР ЗОН", "yellow")
    ui_row("  0.  Все зоны", "white")

    zone_list = list(all_zones.keys())
    for i, zone_name in enumerate(zone_list, 1):
        count = len(all_zones[zone_name])
        ui_row(f"  {i}.  {zone_name} ({count} городов)", "white")
    ui_empty()

    while True:
        try:
            raw = input(_c("cyan", "║ ") + _c("yellow", "▶ Выберите зоны (через запятую, например: 1,3,5): ")).strip()
            if raw == "0":
                ui_status("✅", "Выбраны все зоны", "green")
                return all_zones

            selected_indices = [int(x.strip()) for x in raw.split(",") if x.strip().isdigit()]
            if not selected_indices:
                continue

            selected_zones = {}
            for idx in selected_indices:
                if 1 <= idx <= len(zone_list):
                    zone_name = zone_list[idx - 1]
                    selected_zones[zone_name] = all_zones[zone_name]

            if selected_zones:
                ui_status("✅", f"Выбрано зон: {len(selected_zones)}", "green")
                return selected_zones
        except (KeyboardInterrupt, EOFError):
            sys.exit(0)
        except ValueError:
            ui_status("⚠️", "Некорректный ввод", "yellow")

def find_duplicate_cities(cities_data: Dict) -> Set[str]:
    city_map = load_map()
    if not city_map:
        return set()

    city_names = {}
    for zone_name, cities in cities_data.items():
        for city in cities:
            cn = normalize(city)
            entry = city_map.get(cn)
            if entry:
                path = entry.get("path_original", [])
                if len(path) >= 2:
                    region = normalize(path[-2])
                else:
                    region = "unknown"
                if cn not in city_names:
                    city_names[cn] = set()
                city_names[cn].add(region)

    duplicates = set()
    for cn, regions in city_names.items():
        if len(regions) > 1:
            duplicates.add(cn)

    if duplicates:
        ui_status("⚠️", f"Найдено городов-дубликатов: {len(duplicates)}", "yellow")
        logger.info(f"Дубликаты: {sorted(duplicates)}")

    return duplicates

def show_skipped_report():
    if not SKIPPED_CITIES:
        return
    ui_sep("ОТЧЁТ О ПРОПУЩЕННЫХ ГОРОДАХ", "yellow")
    ui_status("⚠️", f"Всего пропущено: {len(SKIPPED_CITIES)} городов", "yellow")
    by_zone = {}
    for item in SKIPPED_CITIES:
        zone = item["zone"]
        if zone not in by_zone:
            by_zone[zone] = []
        by_zone[zone].append(item)
    for zone, items in sorted(by_zone.items()):
        ui_row(f"  {zone}:", "cyan")
        for item in items:
            ui_row(f"    •  {item['city']} - {item['reason']}", "white")
    ui_sep()


# ══════════════════════════════════════════════
#  БРАУЗЕР — ЗАПУСК
# ══════════════════════════════════════════════
def launch_browser(profile_dir: str = None):
    max_retries = 3
    for attempt in range(max_retries):
        try:
            pw = sync_playwright().start()
            if profile_dir:
                Path(profile_dir).mkdir(parents=True, exist_ok=True)

            args = [
                "--disable-dev-shm-usage",
                "--no-sandbox",
                "--disable-web-security",
                "--disable-features=VizDisplayCompositor",
                "--disable-background-timer-throttling",
                "--disable-backgrounding-occluded-windows",
                "--disable-renderer-backgrounding",
                "--disable-background-networking",
                "--disable-ipc-flooding-protection",
                "--password-store=basic",
                "--use-mock-keychain",
                "--no-first-run",
                "--disable-default-apps"
            ]

            ctx = pw.chromium.launch_persistent_context(
                user_data_dir=profile_dir or PROFILE_DIR,
                headless=False,
                args=args,
                ignore_default_args=["--enable-automation"],
            )

            page = ctx.new_page()
            page.on("pageerror", lambda e: logger.debug(f"JS error: {e}"))
            return pw, ctx, page

        except Exception as e:
            if attempt < max_retries - 1:
                wait_time = (attempt + 1) * 2
                logger.warning(f"Browser launch attempt {attempt + 1} failed: {e}. Retrying in {wait_time}s...")
                time.sleep(wait_time)
                continue
            else:
                raise Exception(f"Failed to launch browser after {max_retries} attempts: {e}")


# ══════════════════════════════════════════════
#  DRAWER — УПРАВЛЕНИЕ ПАНЕЛЬЮ ТАРИФОВ
# ══════════════════════════════════════════════
def dismiss_overlays(page) -> None:
    try:
        page.evaluate("""
            () => {
                document.querySelectorAll('[class*="overlay"]').forEach(el => {
                    if (el.style) el.style.pointerEvents = 'none';
                });
            }
        """)
        page.wait_for_timeout(200)
    except Exception:
        pass

def lock_drawer_open(page) -> None:
    try:
        page.evaluate("""
            () => {
                if (window.__drawer_locked) return;
                window.__drawer_locked = true;
                document.addEventListener('click', (e) => {
                    const t = e.target;
                    if (
                        t.hasAttribute('data-vaul-overlay') ||
                        t.closest('[data-vaul-overlay]')
                    ) {
                        e.stopImmediatePropagation();
                        e.preventDefault();
                    }
                }, true);
                document.addEventListener('keydown', (e) => {
                    if (e.key === 'Escape') {
                        const rows = document.querySelectorAll("tr[data-e2e='levitan-table-row']");
                        if (rows.length > 0) {
                            e.stopImmediatePropagation();
                            e.preventDefault();
                        }
                    }
                }, true);
                const style = document.createElement('style');
                style.textContent = '[data-vaul-overlay] { pointer-events: none !important; }';
                document.head.appendChild(style);
            }
        """)
        logger.info("Drawer lock установлен")
    except Exception as e:
        logger.debug(f"lock_drawer_open error: {e}")

def is_tariff_drawer_open(page) -> bool:
    """Проверяет что drawer открыт И дерево интерактивно (есть кнопки раскрытия)."""
    try:
        # Надёжный признак — кнопки раскрытия в дереве (как в kur.py)
        expand_btns = page.locator("tr[data-e2e='levitan-table-row'] button[aria-expanded]").count()
        if expand_btns > 0:
            return True
        # Или хотя бы кнопка сохранения видна
        save_btn = page.locator("[data-e2e*='tariff-setting-save-button']").count()
        return save_btn > 0
    except Exception:
        return False


def scroll_to_tree_section(page) -> bool:
    """
    Скроллит внутри drawer к секции 'Где применяется' с деревом регионов.
    Секция может быть ниже видимой области drawer.
    """
    try:
        # Попытка 1: scroll к заголовку секции "Где применяется"
        title_sel = "[data-e2e-i18n-key*='tariff-table.title']"
        title_el = page.locator(title_sel).first
        if title_el.count() > 0:
            title_el.scroll_into_view_if_needed(timeout=2_000)
            logger.info("Проскроллили к секции 'Где применяется' (по i18n key)")
            return True
    except Exception:
        pass

    try:
        # Попытка 2: scroll к таблице дерева
        table_sel = "[data-levitan-table]"
        table_el = page.locator(table_sel).first
        if table_el.count() > 0:
            table_el.scroll_into_view_if_needed(timeout=2_000)
            logger.info("Проскроллили к таблице дерева (по data-levitan-table)")
            return True
    except Exception:
        pass

    try:
        # Попытка 3: scroll к первой строке дерева
        first_row = page.locator("tr[data-e2e='levitan-table-row']").first
        if first_row.count() > 0:
            first_row.scroll_into_view_if_needed(timeout=2_000)
            logger.info("Проскроллили к первой строке дерева")
            return True
    except Exception:
        pass

    # Попытка 4: JS скролл внутри drawer content
    try:
        page.evaluate("""
            () => {
                const wrapper = document.querySelector('[data-e2e="levitan-drawer-content"]')
                    || document.querySelector('[class*="content-wrapper"]');
                if (wrapper) {
                    wrapper.scrollTop = wrapper.scrollHeight;
                    return true;
                }
                // Скроллим весь drawer
                const drawer = document.querySelector('[data-e2e="levitan-drawer-wrapper"]');
                if (drawer) {
                    const scrollable = drawer.querySelector('[class*="content"]');
                    if (scrollable) scrollable.scrollTop = scrollable.scrollHeight;
                }
                return false;
            }
        """)
        page.wait_for_timeout(1_000)
        logger.info("JS-скролл внутри drawer контейнера")
        return True
    except Exception as e:
        logger.debug(f"scroll_to_tree_section JS error: {e}")

    return False


def wait_for_tree_interactive(page, timeout_s: int = 30) -> bool:
    """
    Ждёт пока в дереве появятся ИНТЕРАКТИВНЫЕ элементы (кнопки раскрытия).
    React может рендерить <tr> раньше чем <button aria-expanded> внутри них.
    """
    interactive_sel = "tr[data-e2e='levitan-table-row'] button[aria-expanded]"
    deadline = time.time() + timeout_s
    attempt = 0

    while time.time() < deadline:
        attempt += 1
        try:
            count = page.locator(interactive_sel).count()
            if count > 0:
                ui_status("✅", f"Дерево интерактивно: {count} кнопок раскрытия", "green")
                logger.info(f"Дерево интерактивно: {count} кнопок (попытка {attempt})")
                return True
        except Exception:
            pass

        # Если не нашли — скроллим к дереву и ждём
        if attempt <= 3:
            scroll_to_tree_section(page)

        remaining = int(deadline - time.time())
        if attempt % 8 == 0:
            ui_status("⏳", f"Ждём загрузки дерева... ({remaining}с осталось)", "grey")

        page.wait_for_timeout(400)

    # Последний шанс — проверяем сколько строк вообще есть
    rows = page.locator("tr[data-e2e='levitan-table-row']").count()
    ui_status("⚠️", f"Дерево НЕ стало интерактивным за {timeout_s}с (строк: {rows})", "yellow")
    logger.warning(f"Дерево не интерактивно за {timeout_s}с: {rows} строк, 0 кнопок")
    return False

def _drawer_wrapper_visible(page) -> bool:
    """Проверяет что drawer wrapper элемент существует в DOM."""
    try:
        return page.locator("[data-e2e='levitan-drawer-wrapper']").count() > 0
    except Exception:
        return False


def _diagnose_drawer_state(page) -> Dict:
    """Диагностика состояния drawer — что видно в DOM."""
    try:
        return page.evaluate("""
            () => {
                const result = {};
                result.drawerWrapper = !!document.querySelector('[data-e2e="levitan-drawer-wrapper"]');
                result.drawerContent = !!document.querySelector('[data-e2e="levitan-drawer-content"]');
                result.saveBtn = !!document.querySelector('[data-e2e*="tariff-setting-save-button"]');
                result.closeBtn = !!document.querySelector('[data-e2e*="tariff-setting-drawer-close-button"]');
                result.tableRows = document.querySelectorAll('tr[data-e2e="levitan-table-row"]').length;
                result.expandBtns = document.querySelectorAll('tr[data-e2e="levitan-table-row"] button[aria-expanded]').length;
                result.leviTable = !!document.querySelector('[data-levitan-table]');
                // Текст первой строки
                const firstRow = document.querySelector('tr[data-e2e="levitan-table-row"] td');
                result.firstRowText = firstRow ? firstRow.textContent.trim().substring(0, 60) : '';
                // Есть ли заголовок 'Где применяется'
                const treeTitle = document.querySelector('[data-e2e-i18n-key*="tariff-table.title"]');
                result.treeTitle = !!treeTitle;
                // Есть ли название тарифа
                const nameInput = document.querySelector('[data-e2e="levitan-text-field name"] input');
                result.tariffName = nameInput ? nameInput.value : '';
                return result;
            }
        """)
    except Exception as e:
        return {"error": str(e)}


def _is_in_editing_mode(page) -> bool:
    """Проверяет что drawer в режиме РЕДАКТИРОВАНИЯ (не в списке тарифов)."""
    state = _diagnose_drawer_state(page)
    # В режиме редактирования: есть таблица дерева ИЛИ название тарифа заполнено ИЛИ есть save button
    if state.get("leviTable"):
        return True
    if state.get("tariffName"):
        return True
    if state.get("treeTitle"):
        return True
    return False


def _select_tariff_in_drawer(page, zone_name: str) -> bool:
    """
    Drawer в режиме СПИСКА тарифов.
    Кнопки-меню: data-e2e='levitan-button tariff-menu-toggler' — по одной на каждый тариф.
    Порядок кнопок = порядок тарифов (0=Зона 1, 1=Зона 2, ...).
    Также ищем по тексту DIV рядом с кнопкой.
    """
    ui_status("📋", f"Drawer в режиме списка. Ищем тариф '{zone_name}'...", "cyan")
    logger.info(f"Ищем тариф '{zone_name}' в списке drawer")

    zone_match = re.search(r'(\d+)', zone_name)
    zone_num = zone_match.group(1) if zone_match else ""

    # ── Стратегия: кнопки tariff-menu-toggler идут по порядку зон ──
    # Из логов: data-e2e='levitan-button tariff-menu-toggler' — по одной на зону
    # btnIdx=0 → Зона 1, btnIdx=1 → Зона 2, и т.д.

    toggler_sel = "[data-e2e='levitan-drawer-wrapper'] [data-e2e='levitan-button tariff-menu-toggler']"
    togglers = page.locator(toggler_sel)
    togglers_count = togglers.count()
    logger.info(f"Найдено tariff-menu-toggler кнопок: {togglers_count}")

    if togglers_count == 0:
        # Fallback: ищем любые кнопки без текста в drawer
        toggler_sel = "[data-e2e='levitan-drawer-wrapper'] button[data-e2e*='menu']"
        togglers = page.locator(toggler_sel)
        togglers_count = togglers.count()
        logger.info(f"Fallback menu buttons: {togglers_count}")

    if togglers_count == 0:
        ui_status("❌", f"Не найдены кнопки-меню тарифов в drawer", "red")
        return False

    # Определяем индекс нужной зоны
    target_idx = -1

    # Сначала пробуем найти по тексту TR-строки рядом с кнопкой
    if zone_num:
        # Ищем через JS: для каждого toggler-а проверяем текст его TR-родителя
        result = page.evaluate("""
            (args) => {
                const [sel, zoneNum] = args;
                const btns = document.querySelectorAll(sel);
                const pattern = new RegExp('зона\\\\s*' + zoneNum + '(?!\\\\d)', 'i');
                for (let i = 0; i < btns.length; i++) {
                    // Ищем ближайший TR или контейнер с текстом зоны
                    let el = btns[i];
                    for (let d = 0; d < 6; d++) {
                        el = el.parentElement;
                        if (!el) break;
                        const tag = el.tagName;
                        if (tag === 'TR' || (tag === 'DIV' && el.querySelector('button') === btns[i])) {
                            const text = (el.textContent || '');
                            if (pattern.test(text)) {
                                return {found: true, idx: i, text: text.substring(0, 80)};
                            }
                        }
                    }
                }
                // Fallback: по индексу (зона N → кнопка N-1)
                const idx = parseInt(zoneNum) - 1;
                if (idx >= 0 && idx < btns.length) {
                    return {found: true, idx: idx, text: 'by index', method: 'index'};
                }
                return {found: false, count: btns.length};
            }
        """, [toggler_sel, zone_num])

        if result and result.get("found"):
            target_idx = result["idx"]
            method = result.get("method", "text-match")
            ui_status("🖱", f"Кнопка-меню зоны {zone_num} найдена ({method}), idx={target_idx}", "grey")
            logger.info(f"Зона {zone_num}: toggler idx={target_idx}, text='{result.get('text', '')}'")

    if target_idx < 0:
        # Крайний fallback: по номеру зоны
        if zone_num:
            target_idx = int(zone_num) - 1
            if target_idx >= togglers_count:
                target_idx = 0
        else:
            target_idx = 0

    # Кликаем нужную кнопку-меню
    try:
        btn = togglers.nth(target_idx)
        btn.scroll_into_view_if_needed(timeout=3_000)
        page.wait_for_timeout(300)
        btn.click(timeout=5_000)
        ui_status("🖱", f"Кликнули menu-toggler #{target_idx} для '{zone_name}'", "grey")
        logger.info(f"Кликнули tariff-menu-toggler[{target_idx}]")
        page.wait_for_timeout(1_000)
    except Exception as e:
        logger.error(f"Не удалось кликнуть toggler[{target_idx}]: {e}")
        ui_status("❌", f"Не удалось кликнуть menu-toggler #{target_idx}", "red")
        return False

    return _click_edit_in_option_list(page, zone_name)


def _click_tariff_menu_button(page, container) -> bool:
    """Находит и кликает кнопку-меню (kebab/три точки) внутри контейнера."""
    # Ищем кнопку-меню по разным признакам
    menu_selectors = [
        "button[aria-haspopup='true']",
        "button[aria-haspopup='listbox']",
        "button[aria-haspopup='menu']",
        "button[data-e2e*='menu']",
        "button[data-e2e*='kebab']",
        "button[data-e2e*='more']",
        "button[data-e2e*='action']",
        "button[aria-label*='Действия']",
        "button[aria-label*='действия']",
        "button[aria-label*='Меню']",
        "button[aria-label*='меню']",
    ]
    for sel in menu_selectors:
        try:
            btn = container.locator(sel).first
            if btn.count() > 0 and btn.is_visible(timeout=1_000):
                ui_status("🖱", f"Кнопка-меню найдена: {sel}", "grey")
                btn.scroll_into_view_if_needed(timeout=3_000)
                page.wait_for_timeout(300)
                btn.click(timeout=5_000)
                page.wait_for_timeout(1_000)
                return True
        except Exception:
            continue

    # Fallback: кнопка без текста (иконка) — часто это kebab
    try:
        all_btns = container.locator("button")
        for bi in range(all_btns.count()):
            btn = all_btns.nth(bi)
            try:
                btn_text = (btn.text_content(timeout=1_000) or "").strip()
                if not btn_text:  # Кнопка без текста = иконка (kebab)
                    ui_status("🖱", f"Кликаем кнопку-иконку #{bi}", "grey")
                    btn.scroll_into_view_if_needed(timeout=3_000)
                    page.wait_for_timeout(300)
                    btn.click(timeout=5_000)
                    page.wait_for_timeout(1_000)
                    # Проверяем появился ли option-list
                    option_list = page.locator("[data-e2e='option-list']")
                    if option_list.count() > 0 and option_list.is_visible(timeout=2_000):
                        ui_status("✅", "Option-list появился!", "green")
                        return True
                    # Может popup без data-e2e
                    edit_visible = page.locator("button:has-text('Редактировать')").first
                    if edit_visible.count() > 0 and edit_visible.is_visible(timeout=1_000):
                        return True
            except Exception:
                continue
    except Exception:
        pass

    return False


def _get_current_tariff_name(page) -> str:
    """Возвращает текущее имя тарифа из поля ввода в режиме редактирования."""
    try:
        return page.evaluate("""
            () => {
                const input = document.querySelector('[data-e2e="levitan-text-field name"] input')
                    || document.querySelector('input[value*="Зона"]')
                    || document.querySelector('input[value*="зона"]');
                return input ? input.value : '';
            }
        """) or ""
    except Exception:
        return ""


def _set_tariff_name_in_field(page, zone_name: str) -> None:
    """Устанавливает имя тарифа в поле ввода. Используется чтобы переключить зону."""
    try:
        # Очищаем поле и вводим новое имя
        name_input = page.locator("[data-e2e='levitan-text-field name'] input").first
        if name_input.count() > 0:
            name_input.fill("")
            page.wait_for_timeout(100)
            name_input.fill(zone_name)
            page.wait_for_timeout(200)
            logger.info(f"Имя тарифа изменено на: '{zone_name}'")
            ui_status("✏️", f"Имя тарифа → '{zone_name}'", "cyan")
    except Exception as e:
        logger.warning(f"Не удалось изменить имя тарифа: {e}")


def _click_edit_in_option_list(page, expected_zone: str = "") -> bool:
    """Кликает 'Редактировать' в выпадающем option-list."""
    edit_selectors = [
        "[data-e2e='levitan-link tariff-edit-action']",
        "[data-e2e*='tariff-edit-action']",
        "[data-value='tariff-edit-action'] button",
        "[data-value='tariff-edit-action']",
        "[data-e2e='option-list'] button:has-text('Редактировать')",
        "button:has-text('Редактировать')",
    ]
    for sel in edit_selectors:
        try:
            btn = page.locator(sel).first
            if btn.count() > 0 and btn.is_visible(timeout=2_000):
                ui_status("🖱", f"Кликаем 'Редактировать': {sel}", "cyan")
                logger.info(f"Кликаем 'Редактировать': {sel}")
                btn.click(timeout=5_000)
                page.wait_for_timeout(2_000)
                wait_for_spinner(page)
                if _is_in_editing_mode(page):
                    # Проверяем что открылась ПРАВИЛЬНАЯ зона
                    if expected_zone:
                        actual_name = _get_current_tariff_name(page)
                        zone_match_exp = re.search(r'(\d+)', expected_zone)
                        zone_num_exp = zone_match_exp.group(1) if zone_match_exp else ""
                        zone_match_act = re.search(r'(\d+)', actual_name) if actual_name else None
                        zone_num_act = zone_match_act.group(1) if zone_match_act else ""
                        if actual_name and zone_num_exp and zone_num_act and zone_num_exp != zone_num_act:
                            logger.warning(f"Открылась зона '{actual_name}' (#{zone_num_act}) вместо '{expected_zone}' (#{zone_num_exp})!")
                            ui_status("⚠️", f"Открылась '{actual_name}' вместо '{expected_zone}' — нужный kebab не сработал", "yellow")
                    logger.info("Перешли в режим редактирования!")
                    ui_status("✅", "Перешли в режим редактирования!", "green")
                    return True
                # Может нужно подождать ещё
                for tick in range(8):
                    page.wait_for_timeout(1_000)
                    if _is_in_editing_mode(page):
                        if expected_zone:
                            actual_name = _get_current_tariff_name(page)
                            zone_match2 = re.search(r'(\d+)', expected_zone)
                        zn2_exp = zone_match2.group(1) if zone_match2 else ""
                        zone_match2a = re.search(r'(\d+)', actual_name) if actual_name else None
                        zn2_act = zone_match2a.group(1) if zone_match2a else ""
                        if actual_name and zn2_exp and zn2_act and zn2_exp != zn2_act:
                            logger.warning(f"Открылась зона '{actual_name}' вместо '{expected_zone}'")
                        logger.info(f"Режим редактирования загрузился (после {tick+1}с)")
                        ui_status("✅", "Режим редактирования загрузился!", "green")
                        return True
        except Exception:
            continue
    return False


def open_tariff_drawer(page, zone_name: str = "") -> bool:
    """
    Открывает drawer тарифов И переходит в режим РЕДАКТИРОВАНИЯ нужного тарифа.
    Этапы:
    1. Кликает «Настроить» → drawer открывается (может быть в режиме СПИСКА)
    2. Если drawer в режиме списка → кликает по нужному тарифу
    3. Скроллит к дереву регионов
    4. Ждёт пока дерево станет интерактивным
    """
    # Проверяем — может drawer уже открыт и дерево готово
    if is_tariff_drawer_open(page):
        ui_status("✅", "Drawer тарифов уже открыт, дерево интерактивно", "green")
        lock_drawer_open(page)
        return True

    ui_status("🔍", "Ищем кнопку «Настроить» тарифы...", "cyan")

    btn_selectors = [
        "[data-e2e='levitan-button tariff-setting-setup-button']",
        "[data-e2e*='tariff-setting-setup-button']",
    ]

    for attempt in range(1, 6):
        ui_status("🔄", f"Попытка {attempt}/5 открыть drawer тарифов", "grey")

        try:
            page.wait_for_load_state("domcontentloaded", timeout=15_000)
            page.wait_for_timeout(500)
            dismiss_overlays(page)

            # ── Запоминаем состояние ДО клика ──
            rows_before = page.locator("tr[data-e2e='levitan-table-row']").count()

            # ── Ищем и кликаем кнопку ──
            clicked = False
            for sel in btn_selectors:
                try:
                    btn = page.locator(sel).first
                    if btn.count() > 0 and btn.is_visible(timeout=2_000):
                        ui_status("✅", f"Кнопка найдена: {sel}", "grey")
                        btn.scroll_into_view_if_needed(timeout=2_000)
                        try:
                            btn.click(timeout=3_000)
                        except Exception:
                            dismiss_overlays(page)
                            btn.click(timeout=5_000, force=True)
                        clicked = True
                        ui_status("🖱", "Клик по кнопке выполнен", "cyan")
                        break
                except Exception as e:
                    logger.debug(f"Клик по {sel} не удался: {e}")
                    continue

            if not clicked:
                try:
                    text_btn = page.locator("button:has-text('Настроить')").first
                    if text_btn.count() > 0 and text_btn.is_visible(timeout=2_000):
                        text_btn.scroll_into_view_if_needed(timeout=3_000)
                        page.wait_for_timeout(300)
                        text_btn.click(timeout=5_000)
                        clicked = True
                        ui_status("✅", "Клик по 'Настроить' (по тексту)", "grey")
                except Exception:
                    pass

            if not clicked:
                ui_status("⚠️", f"Кнопка не найдена (попытка {attempt})", "yellow")
                page.wait_for_timeout(2_000)
                continue

            # ── Этап 1: ждём появления DRAWER WRAPPER ──
            ui_status("⏳", "Ждём открытия drawer...", "cyan")
            drawer_opened = False

            for wait_tick in range(20):
                page.wait_for_timeout(500)
                if _drawer_wrapper_visible(page):
                    drawer_opened = True
                    ui_status("✅", "Drawer появился!", "green")
                    break
                rows_now = page.locator("tr[data-e2e='levitan-table-row']").count()
                save_visible = page.locator("[data-e2e*='tariff-setting-save-button']").count() > 0
                if save_visible or rows_now > rows_before:
                    drawer_opened = True
                    ui_status("✅", f"Drawer открыт", "green")
                    break
                if wait_tick == 8:
                    dismiss_overlays(page)
                    for sel in btn_selectors:
                        try:
                            btn2 = page.locator(sel).first
                            if btn2.count() > 0 and btn2.is_visible(timeout=1_000):
                                btn2.click(timeout=3_000)
                                ui_status("🖱", "Повторный клик", "cyan")
                                break
                        except Exception:
                            continue

            if not drawer_opened:
                state = _diagnose_drawer_state(page)
                ui_status("⚠️", f"Drawer НЕ открылся", "yellow")
                logger.warning(f"Drawer не открылся (попытка {attempt}): {state}")
                page.wait_for_timeout(2_000)
                continue

            # ── Этап 2: drawer открыт — определяем режим ──
            wait_for_spinner(page)
            lock_drawer_open(page)

            state = _diagnose_drawer_state(page)
            logger.info(f"Drawer state после открытия: {state}")
            ui_status("📋", f"Drawer: таблица={state.get('leviTable')}, "
                      f"строк={state.get('tableRows')}, кнопок={state.get('expandBtns')}, "
                      f"тариф='{state.get('tariffName','')}'", "grey")

            # ── Этап 2.5: Если drawer в РЕЖИМЕ СПИСКА — выбираем тариф ──
            if not _is_in_editing_mode(page):
                ui_status("📋", "Drawer в режиме СПИСКА тарифов", "cyan")
                if zone_name:
                    if _select_tariff_in_drawer(page, zone_name):
                        ui_status("✅", f"Тариф '{zone_name}' выбран, перешли в редактирование", "green")
                    else:
                        ui_status("⚠️", f"Не удалось выбрать тариф '{zone_name}'", "yellow")
                        ui_wait_enter(f"Выберите тариф '{zone_name}' вручную и нажмите Enter")
                else:
                    # Нет имени зоны — пробуем кликнуть первый доступный тариф
                    ui_status("🔍", "Пробуем открыть первый тариф...", "cyan")
                    try:
                        first_card = page.locator("[data-e2e*='tariff-setting-card']").first
                        if first_card.count() > 0:
                            first_card.click(timeout=5_000)
                            page.wait_for_timeout(2_000)
                            wait_for_spinner(page)
                    except Exception:
                        pass

                    if not _is_in_editing_mode(page):
                        ui_wait_enter("Выберите тариф вручную и нажмите Enter")

                # Перепроверяем состояние
                state = _diagnose_drawer_state(page)
                logger.info(f"Drawer state после выбора тарифа: {state}")

            # ── Этап 3: скроллим к дереву и ждём ИНТЕРАКТИВНОСТИ ──
            lock_drawer_open(page)
            ui_status("⏳", "Скроллим к дереву регионов...", "cyan")
            scroll_to_tree_section(page)

            if wait_for_tree_interactive(page, timeout_s=15):
                ui_status("✅", "Drawer открыт, дерево готово!", "green")
                return True

            # Доп. ожидание
            scroll_to_tree_section(page)
            page.wait_for_timeout(2_000)
            if wait_for_tree_interactive(page, timeout_s=10):
                ui_status("✅", "Дерево загрузилось!", "green")
                return True

            # Даже если дерево не готово — возвращаем True, ensure_root попробует
            state2 = _diagnose_drawer_state(page)
            ui_status("⚠️", f"Дерево не готово. строк={state2.get('tableRows')}, "
                      f"тариф='{state2.get('tariffName','')}'", "yellow")
            logger.warning(f"Дерево не интерактивно: {state2}")
            return True

        except PWTimeout:
            ui_status("⚠️", f"Таймаут на попытке {attempt}", "yellow")
        except Exception as e:
            ui_status("⚠️", f"Ошибка: {e}", "yellow")
            logger.error(f"open_tariff_drawer attempt {attempt} error: {e}")

        page.wait_for_timeout(2_000)

    ui_status("❓", "Нажмите «Настроить» вручную, затем выберите тариф и нажмите Enter.", "yellow")
    ui_wait_enter("После открытия тарифа нажмите Enter")
    return _drawer_wrapper_visible(page)

def save_tariff_changes(page) -> bool:
    logger.info("Сохраняем изменения тарифа...")
    ui_status("💾", "Сохраняем изменения тарифа...", "cyan")

    try:
        page.evaluate("""
            () => {
                const active = document.activeElement;
                if (active && (active.tagName === 'INPUT' || active.tagName === 'TEXTAREA')) {
                    active.blur();
                }
            }
        """)
        page.wait_for_timeout(300)
    except Exception:
        pass

    save_selectors = [
        "[data-e2e='levitan-button tariff-setting-save-button']",
        "[data-e2e*='tariff-setting-save-button']",
        "button:has-text('Сохранить')",
    ]

    for attempt in range(4):
        for sel in save_selectors:
            try:
                btn = page.locator(sel).first
                if btn.count() > 0 and btn.is_visible(timeout=2_000):
                    ui_status("✅", f"Кнопка сохранения найдена", "grey")
                    btn.scroll_into_view_if_needed(timeout=3_000)
                    page.wait_for_timeout(500)
                    btn.click(timeout=5_000, force=True)

                    ui_status("⏳", "Ждём сохранения...", "cyan")
                    page.wait_for_timeout(AFTER_SAVE_MS)

                    if not is_tariff_drawer_open(page):
                        logger.info("Drawer закрылся — изменения сохранены!")
                        ui_status("✅", "Drawer закрылся — изменения сохранены!", "green")
                        return True

                    logger.info("Сохранение выполнено (drawer остался открыт)")
                    ui_status("✅", "Сохранение выполнено", "green")
                    return True
            except Exception as e:
                logger.debug(f"Ошибка сохранения с {sel}: {e}")
                continue
        page.wait_for_timeout(2_000)

    logger.warning("Автосохранение не сработало!")
    ui_status("⚠️", "Автосохранение не сработало", "yellow")
    ui_wait_enter("💾 Нажмите «Сохранить» вручную, затем Enter")
    return True

def close_tariff_drawer(page) -> bool:
    close_selectors = [
        "[data-e2e*='tariff-setting-drawer-close-button']",
        "button[aria-label='Закрыть']",
        "[data-e2e='close-button']",
    ]
    for sel in close_selectors:
        try:
            btn = page.locator(sel).first
            if btn.count() > 0 and btn.is_visible():
                btn.click(timeout=3_000)
                page.wait_for_timeout(2_000)
                if not is_tariff_drawer_open(page):
                    logger.info("Drawer закрыт")
                    return True
        except Exception:
            continue
    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(2_000)
        return True
    except Exception:
        pass
    return False

def ensure_root_expanded(page) -> bool:
    """
    Находит корневой узел дерева ('Все регионы') и раскрывает его.
    Если кнопки раскрытия ещё не появились — ждёт и скроллит к дереву.
    """
    for retry in range(5):
        # Скроллим к дереву на каждой попытке
        scroll_to_tree_section(page)
        page.wait_for_timeout(500)

        rows = page.locator("tr[data-e2e='levitan-table-row']")
        n = rows.count()

        if n == 0:
            ui_status("⏳", f"Нет строк в дереве (попытка {retry+1}/5), ждём...", "yellow")
            page.wait_for_timeout(2_000)
            continue

        # Ищем expandable корень
        for i in range(min(5, n)):
            r = rows.nth(i)
            if row_is_expandable(r):
                if row_is_expanded(r):
                    txt = get_row_text(r)
                    ui_status("✅", f"Корень раскрыт: '{txt}'", "green")
                    logger.info(f"Корень уже раскрыт: '{txt}', строк в дереве: {n}")
                    return True
                txt = get_row_text(r)
                ui_status("🌳", f"Раскрываем корень: '{txt}'", "cyan")
                logger.info(f"Раскрываем корень: '{txt}'")
                click_expand_row(page, r)
                # Проверяем что дочерние строки появились
                new_n = page.locator("tr[data-e2e='levitan-table-row']").count()
                if new_n > n:
                    ui_status("✅", f"Корень раскрыт: {n} → {new_n} строк", "green")
                    logger.info(f"Корень раскрыт: {n} → {new_n} строк")
                    return True
                else:
                    ui_status("⚠️", f"Клик по корню, но строк не прибавилось ({new_n})", "yellow")
                    # Всё равно продолжаем — может просто ещё не загрузились
                    return True

        # Строки есть, но ни одна не expandable — кнопки ещё не загрузились
        ui_status("⏳", f"Строк: {n}, но кнопок раскрытия нет (попытка {retry+1}/5)...", "yellow")
        logger.warning(f"ensure_root: {n} строк, 0 expandable (попытка {retry+1})")

        # Ждём появления кнопок раскрытия
        try:
            page.wait_for_selector(
                "tr[data-e2e='levitan-table-row'] button[aria-expanded]",
                timeout=5_000
            )
            ui_status("✅", "Кнопки раскрытия появились!", "green")
            # Повторяем цикл — теперь найдём expandable
            continue
        except PWTimeout:
            pass

        # Попытка через JS — может кнопка скрыта или ещё рендерится
        try:
            result = page.evaluate("""
                () => {
                    const rows = document.querySelectorAll("tr[data-e2e='levitan-table-row']");
                    const info = [];
                    for (let i = 0; i < Math.min(rows.length, 5); i++) {
                        const r = rows[i];
                        const btn = r.querySelector("button[aria-expanded]");
                        const text = (r.querySelector('td')?.textContent || '').trim().substring(0, 50);
                        info.push({
                            idx: i,
                            text: text,
                            hasBtn: !!btn,
                            expanded: btn ? btn.getAttribute('aria-expanded') : null,
                            classes: r.className.substring(0, 100)
                        });
                    }
                    return info;
                }
            """)
            logger.info(f"JS-диагностика строк дерева: {result}")
            for item in result:
                if item.get("hasBtn"):
                    ui_status("🔍", f"JS нашёл кнопку в строке {item['idx']}: '{item['text']}'", "cyan")
        except Exception as e:
            logger.debug(f"JS диагностика ошибка: {e}")

    ui_status("❌", "Корневой узел НЕ найден после 5 попыток!", "red")
    logger.error("Корневой узел не найден после 5 попыток")
    return False


# ══════════════════════════════════════════════
#  ПОЛНОЕ РАСКРЫТИЕ ДЕРЕВА В ОКРУГЕ
#  (Пошагово, как в kur.py — по 3 узла за раз)
# ══════════════════════════════════════════════
def _get_collapsed_batch(page, other_districts: List[str], batch: int = 3) -> List[Dict]:
    return page.evaluate("""
        ([otherDistricts, batch]) => {
            const rows = Array.from(document.querySelectorAll("tr[data-e2e='levitan-table-row']"));
            const result = [];
            for (let i = 0; i < rows.length && result.length < batch; i++) {
                const btn = rows[i].querySelector("button[aria-expanded='false']");
                if (!btn || btn.offsetParent === null) continue;
                const rowText = (rows[i].querySelector('td')?.textContent || '').trim();
                let isOther = false;
                for (const d of otherDistricts) {
                    if (rowText.includes(d) || d.includes(rowText)) { isOther = true; break; }
                }
                if (!isOther) result.push({ idx: i, text: rowText });
            }
            return result;
        }
    """, [other_districts, batch])

def expand_district_fully(page, district_name: str) -> int:
    district_row = None
    for attempt in range(3):
        district_row = find_district_row(page, district_name)
        if district_row:
            break
        ui_status("⏳", f"Повтор поиска округа ({attempt+1}/3)...", "yellow")
        page.wait_for_timeout(1500)

    if district_row is None:
        ui_status("❌", f"Округ не найден: {district_name}", "red")
        return 0
    if not row_is_expanded(district_row):
        click_expand_row(page, district_row)

    total_expanded = 1
    rounds_no_change = 0
    other_districts = [d for d in FEDERAL_DISTRICTS if d != district_name]

    for iteration in range(200):
        check_pause()

        batch = _get_collapsed_batch(page, other_districts, 5)

        if not batch:
            rounds_no_change += 1
            if rounds_no_change >= 3:
                still = page.evaluate("""
                    ([otherDistricts]) => {
                        const rows = Array.from(document.querySelectorAll("tr[data-e2e='levitan-table-row']"));
                        for (const row of rows) {
                            const btn = row.querySelector("button[aria-expanded='false']");
                            if (!btn || btn.offsetParent === null) continue;
                            const rowText = (row.querySelector('td')?.textContent || '').trim();
                            let isOther = false;
                            for (const d of otherDistricts) {
                                if (rowText.includes(d) || d.includes(rowText)) { isOther = true; break; }
                            }
                            if (!isOther) return true;
                        }
                        return false;
                    }
                """, [other_districts])
                if not still:
                    break
                rounds_no_change = 0
            continue

        rounds_no_change = 0
        for item in batch:
            try:
                target_row = page.locator("tr[data-e2e='levitan-table-row']").nth(item["idx"])
                if click_expand_row(page, target_row):
                    total_expanded += 1
                    if total_expanded % 10 == 0:
                        logger.info(f"Раскрыто {total_expanded} узлов в '{district_name}'")
            except Exception as e:
                logger.debug(f"Ошибка раскрытия: {e}")

    total_rows = page.locator("tr[data-e2e='levitan-table-row']").count()
    ui_status("✅", f"Раскрыт '{district_name}': {total_expanded} узлов, {total_rows} строк всего", "green")
    logger.info(f"Раскрыт '{district_name}': {total_expanded} узлов, {total_rows} строк всего")
    return total_expanded


# ══════════════════════════════════════════════
#  ОТМЕТКА ЧЕКБОКСОВ ГОРОДОВ
# ══════════════════════════════════════════════
def check_city_checkbox(page, row, city_name: str) -> str:
    """
    Отмечает чекбокс города. Возвращает статус:
    'checked' — только что отмечен
    'already' — уже был отмечен
    'failed'  — не удалось отметить
    """
    try:
        checkbox_span = row.locator("span[data-e2e^='levitan-clickable checkbox-']").first
        if checkbox_span.count() == 0:
            checkbox_span = row.locator("[data-e2e='levitan-checkbox']").first
        if checkbox_span.count() == 0:
            logger.warning(f"Чекбокс не найден для города: {city_name}")
            return "failed"

        checkbox_input = row.locator("input[type='checkbox']").first
        if checkbox_input.count() > 0:
            try:
                if checkbox_input.is_checked():
                    return "already"
            except Exception:
                pass

        scroll_row_in_drawer(page, row)

        try:
            checkbox_span.click(timeout=5_000)
        except Exception:
            try:
                checkbox_span.evaluate("el => el.click()")
            except Exception:
                try:
                    checkbox_span.click(force=True, timeout=5_000)
                except Exception as e:
                    logger.error(f"Не удалось кликнуть чекбокс для {city_name}: {e}")
                    return "failed"

        page.wait_for_timeout(200)

        if checkbox_input.count() > 0:
            try:
                if checkbox_input.is_checked():
                    return "checked"
                else:
                    logger.warning(f"Клик выполнен, но чекбокс НЕ отмечен: {city_name}")
                    return "failed"
            except Exception:
                pass

        return "checked"

    except Exception as e:
        logger.error(f"Ошибка отметки чекбокса для {city_name}: {e}")
        return "failed"


# ══════════════════════════════════════════════
#  РЕЖИМ: ЗАПОЛНЕНИЕ ГОРОДОВ С КАРТОЙ
# ══════════════════════════════════════════════
def fill_cities_with_map_for_zone(page, cities: List[str], city_map: Dict,
                                   zone_name: str, all_duplicates: Set[str]) -> Tuple[int, int, List[str]]:
    """
    Режим С КАРТОЙ: раскрывает ТОЛЬКО путь к каждому городу из karta.json.
    Не раскрывает весь округ — идёт точно по пути: Все регионы → Округ → Область → Город.
    """
    # Счётчики
    checked_cities: List[str] = []     # Только что отмечены
    already_cities: List[str] = []     # Уже были отмечены
    failed_cities: List[str] = []      # Не удалось отметить
    not_found_dom: List[str] = []      # Не найдены в DOM (путь не раскрылся)
    not_in_map: List[str] = []         # Нет в карте
    skipped_dup: List[str] = []        # Пропущены (дубликаты)
    total_cities = len(cities)

    logger.info(f"[{zone_name}] Начинаем заполнение С КАРТОЙ: {total_cities} городов")

    # ── Группировка по округам ──
    cities_by_district: Dict[str, List[Dict]] = {}
    for city in cities:
        cn = normalize(city)
        if cn in all_duplicates:
            SKIPPED_CITIES.append({"city": city, "zone": zone_name, "reason": "Дубликат в разных областях"})
            skipped_dup.append(city)
            logger.debug(f"[{zone_name}] Пропуск дубликата: {city}")
            continue

        entry = city_map.get(cn)
        if not entry:
            not_in_map.append(city)
            logger.debug(f"[{zone_name}] Город не в карте: {city}")
            continue

        path = entry.get("path_original", [])
        if len(path) < 2:
            not_in_map.append(city)
            logger.warning(f"[{zone_name}] Слишком короткий путь для {city}: {path}")
            continue

        district = ""
        for p in path:
            pn = normalize(p)
            for fd in FEDERAL_DISTRICTS:
                if normalize(fd) in pn or pn in normalize(fd):
                    district = fd
                    break
            if district:
                break
        if district:
            cities_by_district.setdefault(district, []).append({"city": city, "path": path})
        else:
            not_in_map.append(city)
            logger.warning(f"[{zone_name}] Не определён округ для: {city}, путь: {path}")

    # Сводка
    for d, ents in cities_by_district.items():
        logger.info(f"[{zone_name}] Округ {d}: {len(ents)} городов")
    if not_in_map:
        ui_status("⚠️", f"Не в карте: {len(not_in_map)} городов", "yellow")
        logger.warning(f"[{zone_name}] Не в карте ({len(not_in_map)}): {not_in_map}")
    if skipped_dup:
        ui_status("⏭", f"Дубликаты: {len(skipped_dup)}", "grey")
        logger.info(f"[{zone_name}] Дубликаты ({len(skipped_dup)}): {skipped_dup}")

    processed = 0

    # ── Проверяем что корень раскрыт ──
    ensure_root_expanded(page)

    for district, entries in cities_by_district.items():
        ui_status("📍", f"Округ: {district} ({len(entries)} городов)", "cyan")
        logger.info(f"[{zone_name}] === Округ: {district} ({len(entries)} городов) ===")

        # Раскрываем округ (первый уровень)
        district_row = find_district_row(page, district)
        if district_row is None:
            scroll_to_tree_section(page)
            ensure_root_expanded(page)
            district_row = find_district_row(page, district)

        if district_row is None:
            logger.error(f"[{zone_name}] Округ {district} НЕ найден!")
            not_found_dom.extend([e["city"] for e in entries])
            processed += len(entries)
            for e in entries:
                ui_status("❌", f"[{processed}/{total_cities}] Округ не найден: {e['city']}", "red")
            continue

        if not row_is_expanded(district_row):
            logger.info(f"[{zone_name}] Раскрываем округ: {district}")
            rows_before = page.locator("tr[data-e2e='levitan-table-row']").count()
            click_expand_row(page, district_row)
            # Ждём появления дочерних строк округа
            for _w in range(20):
                page.wait_for_timeout(300)
                if page.locator("tr[data-e2e='levitan-table-row']").count() > rows_before:
                    break

        # ── Идём по каждому городу — раскрываем ТОЛЬКО его путь ──
        for entry in entries:
            city = entry["city"]
            path = entry["path"]
            processed += 1

            logger.info(f"[{zone_name}] [{processed}/{total_cities}] Город: {city}, путь: {' → '.join(path)}")
            ui_status("🔍", f"[{processed}/{total_cities}] {city}", "cyan")

            # Идём по шагам пути (пропускаем "Все регионы" и округ — они уже раскрыты)
            path_ok = True
            for step_idx, step_name in enumerate(path):
                step_clean = strip_count_suffix(normalize(step_name))

                # Пропускаем "Все регионы" и федеральный округ
                is_root = "все регионы" in step_clean or "все_регионы" in step_clean
                is_district = False
                for fd in FEDERAL_DISTRICTS:
                    if normalize(fd) in step_clean or step_clean in normalize(fd):
                        is_district = True
                        break
                if is_root or is_district:
                    continue

                # Ищем строку в текущем дереве (с повторами — дочерние строки могут загружаться)
                found_step = False
                for search_attempt in range(10):
                    rows = page.locator("tr[data-e2e='levitan-table-row']")
                    row_count = rows.count()

                    for ri in range(row_count):
                        r = rows.nth(ri)
                        row_text_raw = get_row_text(r)
                        row_clean = strip_count_suffix(normalize(row_text_raw))

                        if row_clean != step_clean:
                            continue

                        # Нашли нужную строку
                        if step_idx == len(path) - 1:
                            # Последний шаг = город → отмечаем чекбокс
                            result = check_city_checkbox(page, r, city)
                            if result == "checked":
                                checked_cities.append(city)
                                logger.info(f"[{zone_name}] ✅ [{processed}/{total_cities}] ОТМЕЧЕН: {city}")
                                ui_status("✅", f"[{processed}/{total_cities}] Отмечен: {city}", "green")
                            elif result == "already":
                                already_cities.append(city)
                                logger.info(f"[{zone_name}] ⏭ [{processed}/{total_cities}] Уже отмечен: {city}")
                                ui_status("⏭", f"[{processed}/{total_cities}] Уже отмечен: {city}", "grey")
                            else:
                                failed_cities.append(city)
                                logger.warning(f"[{zone_name}] ⚠️ [{processed}/{total_cities}] НЕ УДАЛОСЬ: {city}")
                                ui_status("⚠️", f"[{processed}/{total_cities}] Не удалось: {city}", "yellow")
                        else:
                            # Промежуточный шаг → раскрываем если свёрнут
                            if row_is_expandable(r) and not row_is_expanded(r):
                                logger.debug(f"[{zone_name}] Раскрываем шаг: {row_text_raw}")
                                rows_before = row_count
                                click_expand_row(page, r)
                                # Ждём появления дочерних строк
                                for wait_i in range(15):
                                    page.wait_for_timeout(300)
                                    new_count = page.locator("tr[data-e2e='levitan-table-row']").count()
                                    if new_count > rows_before:
                                        break
                                else:
                                    logger.debug(f"[{zone_name}] Дочерние строки не появились после раскрытия {row_text_raw}")

                        found_step = True
                        break

                    if found_step:
                        break
                    # Не нашли — ждём и пробуем снова
                    page.wait_for_timeout(500)
                    if search_attempt == 4:
                        logger.info(f"[{zone_name}] Шаг '{step_name}' не найден (попытка 5/10), ждём загрузки...")

                if not found_step:
                    logger.warning(f"[{zone_name}] Шаг пути не найден после 10 попыток: '{step_name}' (для {city})")
                    path_ok = False
                    break

            if not path_ok:
                not_found_dom.append(city)
                logger.warning(f"[{zone_name}] ❌ [{processed}/{total_cities}] Путь не раскрылся: {city}")
                ui_status("❌", f"[{processed}/{total_cities}] Не найден: {city}", "red")

            # Промежуточный прогресс
            if processed % 20 == 0:
                done = len(checked_cities) + len(already_cities)
                logger.info(f"[{zone_name}] Прогресс: {done}/{total_cities} "
                            f"(отмечено={len(checked_cities)}, уже_были={len(already_cities)}, "
                            f"не_найдено={len(not_found_dom)}, ошибки={len(failed_cities)})")
                ui_status("📊", f"Прогресс: {done}/{total_cities} "
                          f"(✅{len(checked_cities)} ⏭{len(already_cities)} "
                          f"❌{len(not_found_dom)} ⚠️{len(failed_cities)})", "cyan")

        logger.info(f"[{zone_name}] Округ {district} завершён. "
                    f"Отмечено: {len(checked_cities)}, уже_были: {len(already_cities)}")

    # ── ОТЧЁТ ПО ЗОНЕ ──
    all_not_found = not_found_dom + not_in_map + failed_cities
    total_found = len(checked_cities) + len(already_cities)

    ui_sep(f"ОТЧЁТ: {zone_name}", "cyan")
    ui_status("📊", f"Всего городов в Excel: {total_cities}", "cyan")
    ui_status("✅", f"Отмечено сейчас: {len(checked_cities)}", "green")
    ui_status("⏭", f"Уже были отмечены: {len(already_cities)}", "grey")
    ui_status("🔍", f"Всего найдено: {total_found} из {total_cities}", "cyan")
    if not_found_dom:
        ui_status("❌", f"Не найдено в DOM: {len(not_found_dom)}", "red")
        for c in not_found_dom[:10]:
            ui_row(f"   • {c}", "grey")
        if len(not_found_dom) > 10:
            ui_row(f"   ... и ещё {len(not_found_dom) - 10}", "grey")
    if not_in_map:
        ui_status("⚠️", f"Не в карте: {len(not_in_map)}", "yellow")
        for c in not_in_map[:10]:
            ui_row(f"   • {c}", "grey")
        if len(not_in_map) > 10:
            ui_row(f"   ... и ещё {len(not_in_map) - 10}", "grey")
    if failed_cities:
        ui_status("⚠️", f"Ошибка отметки: {len(failed_cities)}", "yellow")
        for c in failed_cities[:10]:
            ui_row(f"   • {c}", "grey")
    if skipped_dup:
        ui_status("⏭", f"Пропущено (дубликаты): {len(skipped_dup)}", "grey")

    logger.info(f"[{zone_name}] === ИТОГ: отмечено={len(checked_cities)}, "
                f"уже_были={len(already_cities)}, не_найдено_DOM={len(not_found_dom)}, "
                f"не_в_карте={len(not_in_map)}, ошибки={len(failed_cities)}, "
                f"дубликаты={len(skipped_dup)} ===")
    if checked_cities:
        logger.info(f"[{zone_name}] Отмечены: {checked_cities}")
    if not_found_dom:
        logger.info(f"[{zone_name}] Не найдены в DOM: {not_found_dom}")
    if not_in_map:
        logger.info(f"[{zone_name}] Не в карте: {not_in_map}")

    selected_count = len(checked_cities) + len(already_cities)
    return selected_count, len(all_not_found), all_not_found


def fill_cities_expand_all(page, cities: List[str], zone_name: str,
                           all_duplicates: Set[str]) -> Tuple[int, int, List[str]]:
    checked_cities: List[str] = []
    already_cities: List[str] = []
    failed_cities: List[str] = []
    found_set: Set[str] = set()  # нормализованные имена найденных городов
    total_cities = len(cities)
    processed = 0

    logger.info(f"[{zone_name}] Начинаем заполнение БЕЗ карты: {total_cities} городов")
    ensure_root_expanded(page)

    for district in FEDERAL_DISTRICTS:
        ui_status("📍", f"Раскрываем округ: {district}", "cyan")
        logger.info(f"[{zone_name}] === Округ: {district} ===")
        expanded = expand_district_fully(page, district)
        if expanded == 0:
            logger.warning(f"[{zone_name}] Не удалось раскрыть округ: {district}")
            continue

        rows = page.locator("tr[data-e2e='levitan-table-row']")
        n = rows.count()
        logger.info(f"[{zone_name}] Округ {district} раскрыт: {expanded} узлов, строк: {n}")

        district_found = 0
        for city in cities:
            cn = normalize(city)
            if cn in all_duplicates or cn in found_set:
                continue

            for i in range(n):
                r = rows.nth(i)
                if not row_is_city(r):
                    continue
                row_name = get_row_clean_name(r)
                if row_name == strip_count_suffix(cn):
                    found_set.add(cn)
                    processed += 1
                    district_found += 1
                    result = check_city_checkbox(page, r, city)
                    if result == "checked":
                        checked_cities.append(city)
                        logger.info(f"[{zone_name}] ✅ [{processed}/{total_cities}] ОТМЕЧЕН: {city}")
                        ui_status("✅", f"[{processed}/{total_cities}] Отмечен: {city}", "green")
                    elif result == "already":
                        already_cities.append(city)
                        logger.info(f"[{zone_name}] ⏭ [{processed}/{total_cities}] Уже отмечен: {city}")
                        ui_status("⏭", f"[{processed}/{total_cities}] Уже отмечен: {city}", "grey")
                    else:
                        failed_cities.append(city)
                        logger.warning(f"[{zone_name}] ⚠️ [{processed}/{total_cities}] НЕ УДАЛОСЬ: {city}")
                        ui_status("⚠️", f"[{processed}/{total_cities}] Не удалось: {city}", "yellow")
                    break

            if processed % 20 == 0 and processed > 0:
                done = len(checked_cities) + len(already_cities)
                logger.info(f"[{zone_name}] Прогресс: {done}/{total_cities} "
                            f"(отмечено={len(checked_cities)}, уже_были={len(already_cities)}, "
                            f"ошибки={len(failed_cities)})")
                ui_status("📊", f"Прогресс: {done}/{total_cities}", "cyan")

        logger.info(f"[{zone_name}] Округ {district} завершён: найдено {district_found} городов")

        page.reload(wait_until="domcontentloaded", timeout=60_000)
        page.wait_for_timeout(3_000)
        open_tariff_drawer(page, zone_name=zone_name)
        ensure_root_expanded(page)

    # Города не найденные в DOM
    not_found_dom = [c for c in cities if normalize(c) not in found_set and normalize(c) not in all_duplicates]

    # ── ОТЧЁТ ПО ЗОНЕ ──
    all_not_found = not_found_dom + failed_cities
    total_found = len(checked_cities) + len(already_cities)

    ui_sep(f"ОТЧЁТ: {zone_name}", "cyan")
    ui_status("📊", f"Всего городов в Excel: {total_cities}", "cyan")
    ui_status("✅", f"Отмечено сейчас: {len(checked_cities)}", "green")
    ui_status("⏭", f"Уже были отмечены: {len(already_cities)}", "grey")
    ui_status("🔍", f"Найдено в дереве: {total_found} из {total_cities}", "cyan")
    logger.info(f"[{zone_name}] ОТЧЁТ: всего={total_cities}, отмечено={len(checked_cities)}, "
                f"уже_были={len(already_cities)}, найдено={total_found}, "
                f"не_найдено_DOM={len(not_found_dom)}, ошибки={len(failed_cities)}")
    if not_found_dom:
        ui_status("❌", f"Не найдено в DOM: {len(not_found_dom)}", "red")
        logger.info(f"[{zone_name}] Не найдены в DOM: {not_found_dom}")
        for c in not_found_dom[:10]:
            ui_row(f"   • {c}", "grey")
        if len(not_found_dom) > 10:
            ui_row(f"   ... и ещё {len(not_found_dom) - 10}", "grey")
    if failed_cities:
        ui_status("⚠️", f"Ошибка отметки: {len(failed_cities)}", "yellow")
        for c in failed_cities[:10]:
            ui_row(f"   • {c}", "grey")

    logger.info(f"=== ОТЧЁТ {zone_name}: отмечено={len(checked_cities)}, "
                f"уже_были={len(already_cities)}, не_найдено={len(not_found_dom)}, "
                f"ошибки={len(failed_cities)} ===")
    if checked_cities:
        logger.info(f"  Отмечены: {checked_cities}")
    if not_found_dom:
        logger.info(f"  Не найдены: {not_found_dom}")

    selected_count = len(checked_cities) + len(already_cities)
    return selected_count, len(all_not_found), all_not_found


# ══════════════════════════════════════════════
#  ПРОГРЕСС
# ══════════════════════════════════════════════
def save_progress(progress_file: str, data: Dict) -> None:
    try:
        with open(progress_file, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.info(f"Прогресс сохранён: {progress_file}")
    except Exception as e:
        logger.error(f"Ошибка сохранения прогресса: {e}")

def load_progress(progress_file: str) -> Dict:
    try:
        if Path(progress_file).exists():
            with open(progress_file, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception as e:
        logger.error(f"Ошибка загрузки прогресса: {e}")
    return {}


# ══════════════════════════════════════════════
#  ОПРЕДЕЛЕНИЕ ЗОНЫ ИЗ НАЗВАНИЯ ТАРИФА
# ══════════════════════════════════════════════
def determine_zone_from_name(tariff_name: str) -> Optional[str]:
    if not tariff_name:
        return None
    clean = re.sub(r'[\n\r]+', ' ', tariff_name).strip()
    clean = re.sub(r'\s*(Включен|Выключен|Активен|Неактивен)\s*', '', clean, flags=re.IGNORECASE).strip()

    match = re.search(r'[Зз]она\s*(\d+)|[Zz]one\s*(\d+)', clean, re.IGNORECASE)
    if match:
        zone_num = match.group(1) or match.group(2)
        result = f"Зона {zone_num}"
        logger.info(f"Определена зона из имени '{clean}' → {result}")
        return result

    logger.warning(f"Не удалось определить зону из имени тарифа: '{clean}'")
    return None


# ══════════════════════════════════════════════
#  ПОЛУЧЕНИЕ СПИСКА ТАРИФОВ ИЗ DRAWER
# ══════════════════════════════════════════════

_JS_GET_TARIFF_CONDITIONS = """
() => {
    const results = [];
    // Ищем все блоки условий по data-e2e атрибутам
    let idx = 0;
    while (true) {
        const priceField = document.querySelector(
            `[data-e2e="levitan-text-field region_tariffs[${idx}].prices_rur.courier"]`
        );
        if (!priceField) break;

        const priceInput = priceField.querySelector('input');
        const freeField = document.querySelector(
            `[data-e2e="levitan-text-field region_tariffs[${idx}].free_delivery_price_rur"]`
        );
        const freeInput = freeField ? freeField.querySelector('input') : null;

        const weightFromField = document.querySelector(
            `[data-e2e="levitan-text-field region_tariffs[${idx}].conditions.chargeable_weight_g.from"]`
        );
        const weightFromInput = weightFromField ? weightFromField.querySelector('input') : null;

        const weightToField = document.querySelector(
            `[data-e2e="levitan-text-field region_tariffs[${idx}].conditions.chargeable_weight_g.to"]`
        );
        const weightToInput = weightToField ? weightToField.querySelector('input') : null;

        results.push({
            index: idx,
            price: priceInput ? priceInput.value : '',
            free_delivery: freeInput ? freeInput.value : '',
            weight_from: weightFromInput ? weightFromInput.value : '',
            weight_to: weightToInput ? weightToInput.value : '',
        });
        idx++;
    }
    return results;
}
"""


# ══════════════════════════════════════════════
#  ОСНОВНОЙ FLOW — ЕДИНИЧНЫЙ ЭКЗЕМПЛЯР
# ══════════════════════════════════════════════
def run_single_instance(url: str, cabinet: str, warehouse: str, mode: str,
                        fill_mode: str = "map", cities_data: Dict = None,
                        instance_name: str = "main") -> None:
    specific_profile = f"../profiles/tariffs_{cabinet}_{warehouse}".replace(" ", "_")
    if not Path(specific_profile).exists():
        specific_profile = f"../profiles/browser_profile_{cabinet}_{warehouse}".replace(" ", "_")
    if not Path(specific_profile).exists():
        specific_profile = PROFILE_DIR

    profile_dir = os.environ.get("YD_PROFILE_DIR", specific_profile)

    pw = None
    ctx = None
    page = None

    try:
        ui_status("🔧", f"Запуск браузера с профилем: {profile_dir}", "cyan")
        pw, ctx, page = launch_browser(profile_dir)
        ui_status("✅", "Браузер запущен", "green")
    except Exception as e:
        ui_status("❌", f"Ошибка запуска браузера: {e}", "red")
        logger.error(f"Browser launch failed: {e}")
        return

    try:
        goto_settings(page, url)

        if not cities_data:
            ui_status("❌", "Нет данных для обработки", "red")
            return

        all_duplicates = find_duplicate_cities(cities_data)
        if all_duplicates:
            ui_status("⚠️", f"Будет пропущено {len(all_duplicates)} городов-дубликатов", "yellow")
            logger.info(f"Дубликаты между зонами ({len(all_duplicates)}): {list(all_duplicates)[:20]}")

        city_map = load_map()
        use_map = fill_mode == "map" and bool(city_map)
        if use_map:
            ui_status("🗺", f"Карта загружена: {len(city_map)} записей", "green")
            logger.info(f"Режим с картой: {len(city_map)} записей")
        else:
            ui_status("ℹ", "Режим без карты — полное раскрытие дерева", "cyan")
            logger.info("Режим без карты")

        total_selected = 0
        total_not_found = 0
        all_not_found = []

        for zone_name, cities in cities_data.items():
            ui_sep(f"{zone_name} — {len(cities)} городов", "blue")
            logger.info(f"{'='*60}")
            logger.info(f"=== Обработка {zone_name}: {len(cities)} городов ===")
            logger.info(f"Города: {cities[:10]}{'...' if len(cities) > 10 else ''}")

            if not open_tariff_drawer(page, zone_name=zone_name):
                ui_status("❌", "Не удалось открыть drawer тарифов", "red")
                ui_wait_enter("Откройте drawer вручную и нажмите Enter")
                if not is_tariff_drawer_open(page):
                    logger.error("Drawer не открыт после ручной попытки, пропускаем зону")
                    continue

            root_ok = ensure_root_expanded(page)
            if not root_ok:
                ui_status("⚠️", "Корень не раскрыт! Пробуем продолжить...", "yellow")
                logger.warning("Корень не раскрыт, пробуем работать с текущим деревом")
                # Последняя попытка — ждём ещё
                page.wait_for_timeout(3_000)
                scroll_to_tree_section(page)
                wait_for_tree_interactive(page, timeout_s=10)
                ensure_root_expanded(page)

            if use_map:
                selected, not_found_cnt, not_found = fill_cities_with_map_for_zone(
                    page, cities, city_map, zone_name, all_duplicates
                )
            else:
                selected, not_found_cnt, not_found = fill_cities_expand_all(
                    page, cities, zone_name, all_duplicates
                )

            total_selected += selected
            total_not_found += not_found_cnt
            all_not_found.extend(not_found)

            ui_status("📊", f"{zone_name}: отмечено {selected}, не найдено {not_found_cnt}", "cyan")
            logger.info(f"{zone_name}: отмечено {selected}, не найдено {not_found_cnt}")
            if not_found:
                logger.info(f"  Не найдены: {not_found}")

            logger.info(f"[{zone_name}] Сохраняем изменения тарифа...")
            save_tariff_changes(page)
            logger.info(f"[{zone_name}] Изменения сохранены")

            page.wait_for_timeout(2_000)
            logger.info(f"[{zone_name}] Перезагрузка страницы...")
            page.reload(wait_until="domcontentloaded", timeout=60_000)
            page.wait_for_timeout(3_000)

        ui_sep("ИТОГОВЫЙ ОТЧЁТ", "green")
        ui_status("✅", f"Всего отмечено: {total_selected}", "green")
        ui_status("❌", f"Не найдено: {total_not_found}", "red" if total_not_found > 0 else "grey")
        if all_not_found:
            ui_row("Не найденные города:", "yellow")
            for c in all_not_found[:20]:
                ui_row(f"   • {c}", "grey")
            if len(all_not_found) > 20:
                ui_row(f"   ... и ещё {len(all_not_found) - 20}", "grey")

        show_skipped_report()

    except Exception as e:
        logger.error(f"Ошибка в run_single_instance: {e}", exc_info=True)
        ui_status("❌", f"Ошибка: {e}", "red")
    finally:
        if ctx:
            try:
                ctx.close()
            except Exception:
                pass
        if pw:
            try:
                pw.stop()
            except Exception:
                pass


# ══════════════════════════════════════════════
#  ПАРАЛЛЕЛЬНЫЙ ЗАПУСК
# ══════════════════════════════════════════════
def select_multi_warehouses() -> List[Tuple[str, str, str]]:
    ui_sep("ВЫБОР СКЛАДОВ (ПАРАЛЛЕЛЬНЫЙ РЕЖИМ)", "blue")
    all_warehouses: List[Tuple[str, str, str]] = []
    for cab, whs in WAREHOUSES.items():
        for wh, url in whs.items():
            all_warehouses.append((cab, wh, url))

    ui_empty()
    ui_row("Введите номера складов через пробел (например: 1 3 5):", "yellow")
    for i, (cab, wh, _) in enumerate(all_warehouses, 1):
        ui_row(f"  {i:2d}.  {cab} — {wh}", "white")
    ui_empty()

    while True:
        try:
            raw = input(_c("cyan", "║ ") + _c("yellow", "▶ Номера: ")).strip()
            parts = raw.split()
            selected = []
            valid = True
            for p in parts:
                if p.isdigit() and 1 <= int(p) <= len(all_warehouses):
                    entry = all_warehouses[int(p) - 1]
                    if entry not in selected:
                        selected.append(entry)
                else:
                    ui_row(f"⚠  Неверный номер: {p}", "yellow")
                    valid = False
                    break
            if valid and selected:
                ui_empty()
                ui_row("Выбраны склады:", "green")
                for cab, wh, _ in selected:
                    ui_row(f"  ✔  {cab} — {wh}", "green")
                return selected
            ui_row("⚠  Введите хотя бы один корректный номер", "yellow")
        except (KeyboardInterrupt, EOFError):
            sys.exit(0)

def run_parallel(selected: List[Tuple[str, str, str]], mode: str, fill_mode: str) -> None:
    import subprocess
    import tempfile

    ui_sep("ЗАПУСК ПАРАЛЛЕЛЬНЫХ ПРОЦЕССОВ", "blue")
    ui_status("ℹ", f"Будет запущено {len(selected)} экземпляров", "cyan")

    first_cabinet, first_warehouse, _ = selected[0]
    first_excel_path = find_excel_for_warehouse(first_cabinet, first_warehouse, mode)
    if not first_excel_path:
        ui_status("❌", f"Excel файл не найден для {first_cabinet}/{first_warehouse}", "red")
        return

    first_cities_data = load_cities_data(first_excel_path)
    selected_zones = select_zones(first_cities_data, mode)

    temp_files = []
    selected_with_config = []
    for i, (cabinet, warehouse, url) in enumerate(selected):
        specific_profile = f"../profiles/tariffs_{cabinet}_{warehouse}".replace(" ", "_")
        if not Path(specific_profile).exists():
            specific_profile = f"../profiles/browser_profile_{cabinet}_{warehouse}".replace(" ", "_")
        if not Path(specific_profile).exists():
            specific_profile = PROFILE_DIR

        excel_path = find_excel_for_warehouse(cabinet, warehouse, mode)
        if not excel_path:
            ui_status("⚠️", f"{cabinet}/{warehouse}: Excel файл не найден", "yellow")
            continue

        cities_data = load_cities_data(excel_path)
        filtered_cities = {}
        for zn in selected_zones.keys():
            if zn in cities_data:
                filtered_cities[zn] = cities_data[zn]

        data = {
            "YD_START_URL": url,
            "YD_CABINET": cabinet,
            "YD_WAREHOUSE": warehouse,
            "YD_MODE": mode,
            "YD_FILL_MODE": fill_mode,
            "YD_CITIES_DATA": filtered_cities,
            "YD_PROFILE_DIR": specific_profile,
            "YD_INSTANCE_NAME": f"tariffs-{cabinet}-{warehouse}",
        }

        temp_dir = tempfile.mkdtemp()
        temp_file = os.path.join(temp_dir, f'config_{i}.json')
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        temp_files.append(temp_file)
        selected_with_config.append((cabinet, warehouse, url))

    procs = []
    for i, (cabinet, warehouse, _) in enumerate(selected_with_config):
        name = f"tariffs-{cabinet}-{warehouse}"
        temp_file = temp_files[i]

        if sys.platform == "darwin":
            script_content = f'''#!/bin/bash
cd "{os.getcwd()}"
echo "=== Терминал для {name} ==="
"{sys.executable}" "{__file__}" --instance "{temp_file}"
EXIT_CODE=$?
echo "=== Завершено с кодом: $EXIT_CODE ==="
echo "Нажмите Enter для закрытия..."
read
'''
            script_path = f"/tmp/terminal_{name}.sh"
            with open(script_path, 'w') as f:
                f.write(script_content)
            os.chmod(script_path, 0o755)
            cmd = ["open", "-a", "Terminal.app", script_path]
        elif sys.platform == "win32":
            cmd = ["start", "cmd", "/k", sys.executable, __file__, "--instance", temp_file]
        else:
            cmd = ["gnome-terminal", "--", sys.executable, __file__, "--instance", temp_file]

        ui_status("▶", f"Запускаем: {name}", "cyan")
        try:
            proc = subprocess.Popen(cmd, cwd=os.getcwd(),
                                    stdout=subprocess.DEVNULL,
                                    stderr=subprocess.DEVNULL)
            procs.append(proc)
            time.sleep(3)
        except Exception as e:
            ui_status("❌", f"Ошибка запуска {name}: {e}", "red")

    ui_status("✅", f"Запущено {len(procs)} экземпляров", "green")
    ui_status("ℹ", "Авторизуйтесь в каждом браузере и дождитесь завершения", "yellow")


# ══════════════════════════════════════════════
#  ГЛАВНАЯ ФУНКЦИЯ
# ══════════════════════════════════════════════
def main():
    if "--instance" in sys.argv:
        idx = sys.argv.index("--instance")
        tmp_file = sys.argv[idx + 1]
        print(f"📂 Чтение конфига из: {tmp_file}")

        try:
            with open(tmp_file, encoding="utf-8") as f:
                data = json.load(f)
            Path(tmp_file).unlink(missing_ok=True)
        except Exception as e:
            print(f"❌ Ошибка чтения конфига: {e}")
            sys.exit(1)

        url       = data.get("YD_START_URL", "")
        cabinet   = data.get("YD_CABINET", "")
        warehouse = data.get("YD_WAREHOUSE", "")
        mode      = data.get("YD_MODE", "cities")
        fill_mode = data.get("YD_FILL_MODE", "map")
        cities_data = data.get("YD_CITIES_DATA", {})
        profile   = data.get("YD_PROFILE_DIR", PROFILE_DIR)
        name      = data.get("YD_INSTANCE_NAME", "main")

        os.environ["YD_PROFILE_DIR"]   = profile
        os.environ["YD_INSTANCE_NAME"] = name

        total_cities = sum(len(c) for c in cities_data.values())
        print(f"🎯 {cabinet}/{warehouse}, зон: {len(cities_data)}, городов: {total_cities}")

        run_single_instance(url, cabinet, warehouse, mode, fill_mode,
                            cities_data, instance_name=name)
        return

    os.system("cls" if os.name == "nt" else "clear")
    ui_header()

    for d in [EXCEL_DIR, CITIES_DIR, REPORTS_DIR, MAPS_DIR]:
        Path(d).mkdir(exist_ok=True)

    # ── Режим запуска ──
    ui_sep("ШАГ 0 — РЕЖИМ ЗАПУСКА", "blue")
    launch_mode = ui_ask(
        "Как запустить?",
        ["Единичный — один склад", "Параллельный — несколько складов одновременно"]
    )

    # ── Выбор режима заполнения ──
    fill_mode = "map"
    ui_sep("ВЫБОР РЕЖИМА ЗАПОЛНЕНИЯ", "yellow")
    ui_row("  1.  С картой (быстрый поиск по karta.json)", "white")
    ui_row("  2.  Без карты (полное раскрытие дерева)", "white")
    ui_empty()
    while True:
        try:
            raw = input(_c("cyan", "║ ") + _c("yellow", "▶ Выберите: ")).strip()
            if raw == "1":
                fill_mode = "map"
                break
            elif raw == "2":
                fill_mode = "expand_all"
                break
        except (KeyboardInterrupt, EOFError):
            sys.exit(0)
    ui_status("📝", f"Режим заполнения: {fill_mode}", "cyan")

    mode = "cities"

    # ── Параллельный режим ──
    if "Параллельный" in launch_mode:
        selected = select_multi_warehouses()
        run_parallel(selected, mode, fill_mode)
        show_skipped_report()
        ui_bot()
        return

    # ── Единичный режим ──
    ui_sep("ВЫБОР СКЛАДА", "yellow")
    warehouse_options = []
    for company, cities in WAREHOUSES.items():
        for city in cities.keys():
            warehouse_options.append(f"{company} - {city}")

    for i, opt in enumerate(warehouse_options, 1):
        ui_row(f"  {i}.  {opt}", "white")
    ui_empty()

    while True:
        try:
            raw = input(_c("cyan", "║ ") + _c("yellow", "▶ Выберите склад: ")).strip()
            if raw.isdigit() and 1 <= int(raw) <= len(warehouse_options):
                selected_opt = warehouse_options[int(raw) - 1]
                company, city = selected_opt.split(" - ")
                url = WAREHOUSES[company][city]
                break
        except (KeyboardInterrupt, EOFError):
            sys.exit(0)

    ui_status("🌐", f"Склад: {company}/{city}", "cyan")

    excel_file = select_excel_file(mode, company, city)
    cities_data = load_cities_data(excel_file)

    if not cities_data:
        ui_status("❌", "Нет данных о городах в файле!", "red")
        ui_bot()
        sys.exit(1)

    ui_sep("ЗАГРУЖЕННЫЕ ЗОНЫ", "yellow")
    for zone_name, zone_cities in cities_data.items():
        ui_row(f"  •  {zone_name}: {len(zone_cities)} городов", "white")
    ui_sep()

    cities_data = select_zones(cities_data, mode)

    run_single_instance(url, company, city, mode, fill_mode, cities_data)
    show_skipped_report()
    ui_bot()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        ui_status("⚠️", "Прервано пользователем", "yellow")
        sys.exit(0)
    except Exception as e:
        ui_status("❌", f"Ошибка: {e}", "red")
        logger.exception("Fatal error")
        ui_bot()
        sys.exit(1)
