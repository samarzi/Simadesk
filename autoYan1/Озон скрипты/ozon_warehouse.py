#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Автоматизация складов Ozon — настройка зон доставки
#
# Страница складов: https://seller.ozon.ru/app/warehouse
#
# Структура страницы:
#   Карточки складов: <li> внутри <ul class="nd4-af8">
#   Название склада: [data-widget="warehouse-card-title"]
#   Зоны доставки: div.nd4-ea6[title="..."]  (напр. "Пенза Зона 1")
#   Кнопка "Показать ещё N методов": button:has-text("Показать ещё")
#   Редактировать метод (карандаш): button[title="Редактировать метод"]
#   Кнопка "Далее": button:has-text("Далее")
#   Кнопка "Сохранить": button:has-text("Сохранить")
#   Поиск города: input[placeholder="Направление доставки"]
#   Чекбокс города: input[type="checkbox"] внутри строки таблицы

import os
import re
import sys
import json
import time
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict, Set, Tuple

# ──────────────────────────────────────────────
# Зависимости
# ──────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    print("pip install playwright && playwright install chromium")
    sys.exit(1)


# ══════════════════════════════════════════════
#  КОНСТАНТЫ
# ══════════════════════════════════════════════
OZON_WAREHOUSE_URL = "https://seller.ozon.ru/app/warehouse"
PROFILE_DIR = os.environ.get("OZ_PROFILE_DIR", "../profiles/ozon_profile")
LOG_FILE = f"ozon_{os.environ.get('OZ_INSTANCE_NAME', 'main')}.log"
CITIES_DIR = "../Склады города"

# Тайминги (мс)
AFTER_CLICK_MS = 300
SEARCH_DEBOUNCE_MS = 800
AFTER_SAVE_MS = 3_000

# ──────────────────────────────────────────────
#  Логгер
# ──────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)-7s] %(message)s",
    handlers=[logging.FileHandler(LOG_FILE, encoding="utf-8")],
)
logger = logging.getLogger("OZON")


# ══════════════════════════════════════════════
#  UI — красивый вывод в терминал
# ══════════════════════════════════════════════
COLORS = {
    "red": "\033[91m", "green": "\033[92m", "yellow": "\033[93m",
    "blue": "\033[94m", "magenta": "\033[95m", "cyan": "\033[96m",
    "white": "\033[97m", "grey": "\033[90m", "reset": "\033[0m",
    "bold": "\033[1m",
}

def _c(color: str, text: str) -> str:
    return f"{COLORS.get(color, '')}{text}{COLORS['reset']}"

def ui_sep(label: str = "", color: str = "cyan"):
    if label:
        pad = 60 - len(label)
        left = pad // 2
        right = pad - left
        print(_c(color, f"{'=' * left}  {label}  {'=' * right}"))
    else:
        print(_c(color, "=" * 66))

def ui_status(icon: str, text: str, color: str = "white"):
    line = f" {icon}  {text}"
    print(_c(color, f" {line:<65}"))
    logger.info(f"[UI] {icon} {text}")

def ui_wait_enter(msg: str = "Enter...") -> None:
    print()
    print(_c("yellow", f" {msg}"))
    print()
    try:
        input(_c("cyan", " [ Enter ] "))
    except (KeyboardInterrupt, EOFError):
        sys.exit(0)

def ui_ask(question: str, options: List[str]) -> str:
    print()
    for i, opt in enumerate(options, 1):
        print(_c("white", f"  {i}.  {opt}"))
    print()
    while True:
        try:
            raw = input(_c("yellow", f" {question} ")).strip()
            if raw.isdigit() and 1 <= int(raw) <= len(options):
                return options[int(raw) - 1]
        except (KeyboardInterrupt, EOFError):
            sys.exit(0)


# ══════════════════════════════════════════════
#  EXCEL — загрузка городов по зонам
# ══════════════════════════════════════════════
def load_cities_data(excel_path: str) -> Dict[str, List[str]]:
    """Загружает города из Excel. Формат: строка 1 = заголовки зон, ниже — города."""
    logger.info(f"Загрузка городов из: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    cities_data = {}
    zone_columns = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header and isinstance(header, str) and "зона" in header.lower():
            zone_columns[header.strip()] = col

    if not zone_columns:
        logger.warning(f"Не найдены колонки зон в {excel_path}")
        return cities_data

    logger.info(f"Найдены зоны: {list(zone_columns.keys())}")

    for zone_name, col in zone_columns.items():
        cities = []
        for row in range(2, ws.max_row + 1):
            city = ws.cell(row, col).value
            if city and isinstance(city, str) and city.strip():
                cities.append(city.strip())
        if cities:
            cities_data[zone_name] = cities
            logger.info(f"  {zone_name}: {len(cities)} городов")

    wb.close()
    return cities_data


def select_excel_file() -> str:
    """Выбор Excel файла с городами."""
    cities_dir = Path(CITIES_DIR)
    if not cities_dir.exists():
        cities_dir.mkdir(parents=True, exist_ok=True)

    xlsx_files = sorted(cities_dir.glob("*.xlsx"))
    if not xlsx_files:
        ui_status("", f"Нет .xlsx файлов в {CITIES_DIR}", "red")
        sys.exit(1)

    ui_sep("ВЫБОР ФАЙЛА С ГОРОДАМИ", "yellow")
    for i, f in enumerate(xlsx_files, 1):
        print(_c("white", f"  {i}.  {f.name}"))
    print()

    while True:
        try:
            raw = input(_c("yellow", " Выберите файл: ")).strip()
            if raw.isdigit() and 1 <= int(raw) <= len(xlsx_files):
                return str(xlsx_files[int(raw) - 1])
        except (KeyboardInterrupt, EOFError):
            sys.exit(0)


def select_zones(cities_data: Dict[str, List[str]]) -> Dict[str, List[str]]:
    """Выбор зон для обработки."""
    ui_sep("ВЫБОР ЗОН", "yellow")
    zone_names = list(cities_data.keys())
    print(_c("white", f"  0.  Все зоны ({len(zone_names)})"))
    for i, z in enumerate(zone_names, 1):
        print(_c("white", f"  {i}.  {z} ({len(cities_data[z])} городов)"))
    print()

    while True:
        try:
            raw = input(_c("yellow", " Выберите зоны (через запятую или 0=все): ")).strip()
            if raw == "0":
                return cities_data
            indices = [int(x.strip()) for x in raw.split(",") if x.strip().isdigit()]
            selected = {}
            for idx in indices:
                if 1 <= idx <= len(zone_names):
                    name = zone_names[idx - 1]
                    selected[name] = cities_data[name]
            if selected:
                return selected
        except (KeyboardInterrupt, EOFError):
            sys.exit(0)


# ══════════════════════════════════════════════
#  БРАУЗЕР
# ══════════════════════════════════════════════
def launch_browser(profile_dir: str = None):
    """Запускает Chromium с persistent profile."""
    profile = profile_dir or PROFILE_DIR
    Path(profile).mkdir(parents=True, exist_ok=True)

    ui_status("", f"Запуск браузера с профилем: {profile}", "cyan")
    logger.info(f"Запуск браузера: {profile}")

    pw = sync_playwright().start()
    ctx = pw.chromium.launch_persistent_context(
        profile,
        headless=False,
        viewport={"width": 1400, "height": 900},
        args=[
            "--disable-blink-features=AutomationControlled",
            "--start-maximized",
        ],
    )
    page = ctx.pages[0] if ctx.pages else ctx.new_page()
    ui_status("", "Браузер запущен", "green")
    return pw, ctx, page


def check_auth(page) -> bool:
    """Проверяет авторизацию на Ozon Seller."""
    try:
        # Если есть элементы страницы складов — авторизован
        page.wait_for_selector("[data-widget='warehouse-card-title']", timeout=10_000)
        return True
    except PWTimeout:
        # Проверяем URL — может быть редирект на логин
        if "login" in page.url or "passport" in page.url:
            return False
        # Попробуем подождать ещё
        try:
            page.wait_for_selector("[data-widget='warehouse-card-title']", timeout=15_000)
            return True
        except PWTimeout:
            return False


# ══════════════════════════════════════════════
#  ПОИСК СКЛАДОВ И ЗОН НА СТРАНИЦЕ
# ══════════════════════════════════════════════
def scan_warehouses(page) -> List[Dict]:
    """
    Сканирует все склады на странице.
    Возвращает список: [{name, element_index, zones: [{name, title}]}]
    """
    ui_status("", "Сканируем склады на странице...", "cyan")
    logger.info("Сканируем склады")

    result = page.evaluate("""
        () => {
            const cards = document.querySelectorAll('[data-widget="warehouse-card-title"]');
            const warehouses = [];
            for (let i = 0; i < cards.length; i++) {
                const card = cards[i];
                const name = card.textContent.trim();
                // Ищем зоны внутри карточки склада (в ближайшем li)
                let li = card.closest('li');
                const zones = [];
                if (li) {
                    const zoneDivs = li.querySelectorAll('.nd4-ea6');
                    for (const zd of zoneDivs) {
                        zones.push({
                            name: zd.textContent.trim(),
                            title: zd.getAttribute('title') || zd.textContent.trim()
                        });
                    }
                }
                warehouses.push({name, index: i, zones, zonesCount: zones.length});
            }
            return warehouses;
        }
    """)

    for wh in result:
        logger.info(f"  Склад: '{wh['name']}', зон видно: {wh['zonesCount']}")
        for z in wh.get("zones", []):
            logger.info(f"    Зона: '{z['name']}'")

    ui_status("", f"Найдено складов: {len(result)}", "green")
    return result


def expand_hidden_methods(page, warehouse_name: str):
    """
    Раскрывает скрытые методы доставки для склада.
    Кнопка: "Показать ещё N методов"
    """
    ui_status("", f"Раскрываем скрытые методы для '{warehouse_name}'...", "cyan")
    logger.info(f"Раскрываем методы для: {warehouse_name}")

    max_clicks = 10
    for attempt in range(max_clicks):
        # Найти кнопку "Показать ещё" в карточке этого склада
        clicked = page.evaluate("""
            (whName) => {
                const cards = document.querySelectorAll('[data-widget="warehouse-card-title"]');
                for (const card of cards) {
                    if (card.textContent.trim() === whName) {
                        const li = card.closest('li');
                        if (!li) continue;
                        const buttons = li.querySelectorAll('button');
                        for (const btn of buttons) {
                            const text = btn.textContent.trim();
                            if (/Показать ещё/i.test(text)) {
                                btn.click();
                                return {clicked: true, text: text};
                            }
                        }
                    }
                }
                return {clicked: false};
            }
        """, warehouse_name)

        if clicked and clicked.get("clicked"):
            ui_status("", f"Кликнули: '{clicked.get('text', '')}'", "grey")
            logger.info(f"Expand click: {clicked.get('text')}")
            page.wait_for_timeout(1_000)
        else:
            logger.info(f"Нет больше кнопок 'Показать ещё' для {warehouse_name}")
            break

    ui_status("", "Все методы раскрыты", "green")


def find_zone_edit_button(page, warehouse_name: str, zone_title: str) -> bool:
    """
    Находит зону по title и кликает карандаш (Редактировать метод).
    zone_title — полное название зоны, напр. "Пенза Зона 1"
    """
    ui_status("", f"Ищем зону '{zone_title}' в складе '{warehouse_name}'...", "cyan")
    logger.info(f"Ищем зону: '{zone_title}' в складе '{warehouse_name}'")

    result = page.evaluate("""
        (args) => {
            const [whName, zoneTitle] = args;
            const cards = document.querySelectorAll('[data-widget="warehouse-card-title"]');
            for (const card of cards) {
                if (card.textContent.trim() !== whName) continue;
                const li = card.closest('li');
                if (!li) continue;

                // Ищем div с title зоны
                const zoneDivs = li.querySelectorAll('.nd4-ea6');
                for (const zd of zoneDivs) {
                    const title = zd.getAttribute('title') || zd.textContent.trim();
                    if (title !== zoneTitle) continue;

                    // Нашли зону — ищем кнопку-карандаш рядом
                    // Кнопка title="Редактировать метод" в том же контейнере
                    const parent = zd.closest('.nd4-ea5') || zd.closest('.sc8131-a');
                    if (!parent) continue;

                    const editBtn = parent.querySelector('button[title="Редактировать метод"]');
                    if (editBtn) {
                        editBtn.scrollIntoView({block: 'center'});
                        editBtn.click();
                        return {found: true, method: 'title-attr'};
                    }

                    // Fallback: ищем кнопку с SVG карандаша (path содержит "4.989")
                    const buttons = parent.querySelectorAll('button');
                    for (const btn of buttons) {
                        const svg = btn.querySelector('svg');
                        if (svg && svg.innerHTML.includes('4.989')) {
                            btn.scrollIntoView({block: 'center'});
                            btn.click();
                            return {found: true, method: 'svg-pencil'};
                        }
                    }
                }
            }
            return {found: false};
        }
    """, [warehouse_name, zone_title])

    if result and result.get("found"):
        ui_status("", f"Кнопка редактирования найдена ({result.get('method')})", "green")
        logger.info(f"Edit button found: {result}")
        page.wait_for_timeout(2_000)
        return True

    ui_status("", f"Кнопка редактирования НЕ найдена для '{zone_title}'", "red")
    logger.error(f"Edit button not found for zone '{zone_title}' in '{warehouse_name}'")
    return False


# ══════════════════════════════════════════════
#  РЕДАКТИРОВАНИЕ ЗОНЫ — поиск и выбор городов
# ══════════════════════════════════════════════
def click_next_button(page, timeout: int = 5_000) -> bool:
    """Кликает кнопку 'Далее'."""
    try:
        btn = page.locator("button:has-text('Далее')").first
        if btn.count() > 0 and btn.is_visible(timeout=timeout):
            btn.scroll_into_view_if_needed(timeout=3_000)
            page.wait_for_timeout(300)
            btn.click(timeout=5_000)
            ui_status("", "Кликнули 'Далее'", "grey")
            logger.info("Click: Далее")
            page.wait_for_timeout(1_500)
            return True
    except Exception as e:
        logger.warning(f"Кнопка 'Далее' не найдена: {e}")
    return False


def click_save_button(page) -> bool:
    """Кликает кнопку 'Сохранить'."""
    try:
        btn = page.locator("button:has-text('Сохранить')").first
        if btn.count() > 0 and btn.is_visible(timeout=5_000):
            btn.scroll_into_view_if_needed(timeout=3_000)
            page.wait_for_timeout(300)
            btn.click(timeout=5_000)
            ui_status("", "Кликнули 'Сохранить'", "green")
            logger.info("Click: Сохранить")
            page.wait_for_timeout(AFTER_SAVE_MS)
            return True
    except Exception as e:
        logger.warning(f"Кнопка 'Сохранить' не найдена: {e}")
    return False


def wait_for_search_page(page, timeout_s: int = 15) -> bool:
    """Ждёт появления поля поиска города."""
    ui_status("", "Ждём страницу поиска городов...", "cyan")
    for tick in range(timeout_s * 2):
        try:
            search_input = page.locator("input[placeholder='Направление доставки']").first
            if search_input.count() > 0 and search_input.is_visible(timeout=500):
                ui_status("", "Поле поиска городов готово!", "green")
                return True
        except Exception:
            pass
        page.wait_for_timeout(500)
    ui_status("", "Поле поиска не появилось!", "red")
    return False


def search_and_select_city(page, city_name: str) -> str:
    """
    Ищет город через поле поиска и отмечает чекбокс.
    Возвращает: "checked", "already", "not_found", "failed"
    """
    logger.info(f"Поиск города: '{city_name}'")

    try:
        # Находим поле поиска
        search_input = page.locator("input[placeholder='Направление доставки']").first
        if search_input.count() == 0:
            logger.error("Поле поиска не найдено")
            return "failed"

        # Очищаем и вводим название города
        search_input.click(timeout=3_000)
        page.wait_for_timeout(200)
        search_input.fill("")
        page.wait_for_timeout(200)
        search_input.fill(city_name)
        page.wait_for_timeout(SEARCH_DEBOUNCE_MS)

        # Ждём результаты поиска (появляются в таблице или dropdown)
        # Ищем строку с чекбоксом, содержащую название города
        for attempt in range(10):
            result = page.evaluate("""
                (cityName) => {
                    const norm = cityName.toLowerCase().trim();
                    // Ищем в таблице строки с городами
                    const rows = document.querySelectorAll('tr, [role="row"]');
                    for (const row of rows) {
                        const text = (row.textContent || '').toLowerCase();
                        if (!text.includes(norm)) continue;

                        // Проверяем чекбокс
                        const checkbox = row.querySelector('input[type="checkbox"]');
                        if (!checkbox) continue;

                        if (checkbox.checked) {
                            return {status: 'already', text: row.textContent.trim().substring(0, 100)};
                        }

                        // Кликаем по чекбоксу или его label
                        const label = checkbox.closest('label') || checkbox.parentElement;
                        if (label) {
                            label.click();
                        } else {
                            checkbox.click();
                        }
                        return {status: 'checked', text: row.textContent.trim().substring(0, 100)};
                    }

                    // Ищем в popover/dropdown
                    const popovers = document.querySelectorAll('[ods-popover-reference], [role="listbox"], [role="option"]');
                    for (const p of popovers) {
                        const text = (p.textContent || '').toLowerCase();
                        if (text.includes(norm)) {
                            p.click();
                            return {status: 'clicked_popover', text: p.textContent.trim().substring(0, 100)};
                        }
                    }

                    // Ищем span/div с именем города (результат поиска)
                    const spans = document.querySelectorAll('.nd4-ak3 span, td span');
                    for (const span of spans) {
                        if (span.textContent.trim().toLowerCase() === norm) {
                            // Найти чекбокс в том же ряду
                            const row = span.closest('tr') || span.closest('[role="row"]') || span.parentElement?.parentElement;
                            if (row) {
                                const cb = row.querySelector('input[type="checkbox"]');
                                if (cb) {
                                    if (cb.checked) return {status: 'already', text: span.textContent.trim()};
                                    const lbl = cb.closest('label') || cb.parentElement;
                                    if (lbl) lbl.click(); else cb.click();
                                    return {status: 'checked', text: span.textContent.trim()};
                                }
                            }
                        }
                    }

                    return {status: 'not_found', rowsCount: rows.length};
                }
            """, city_name)

            status = result.get("status", "not_found") if result else "not_found"

            if status == "checked":
                logger.info(f"  Город '{city_name}' отмечен: {result.get('text', '')[:60]}")
                return "checked"
            elif status == "already":
                logger.info(f"  Город '{city_name}' уже отмечен")
                return "already"
            elif status == "clicked_popover":
                logger.info(f"  Город '{city_name}' выбран из popover")
                page.wait_for_timeout(500)
                continue  # После выбора может появиться таблица
            else:
                # Подождать загрузки результатов
                page.wait_for_timeout(500)

        logger.warning(f"  Город '{city_name}' не найден в результатах поиска")
        return "not_found"

    except Exception as e:
        logger.error(f"Ошибка при поиске города '{city_name}': {e}")
        return "failed"


# ══════════════════════════════════════════════
#  ОСНОВНОЙ FLOW — обработка одной зоны
# ══════════════════════════════════════════════
def process_zone(page, warehouse_name: str, zone_title: str, cities: List[str]) -> Dict:
    """
    Обрабатывает одну зону: открывает редактирование, ищет города, сохраняет.
    Возвращает отчёт.
    """
    report = {
        "zone": zone_title,
        "total": len(cities),
        "checked": [],
        "already": [],
        "not_found": [],
        "failed": [],
    }

    ui_sep(f"{zone_title} -- {len(cities)} городов", "blue")

    # 1. Кликаем карандаш для зоны
    if not find_zone_edit_button(page, warehouse_name, zone_title):
        ui_status("", f"Не удалось открыть '{zone_title}' для редактирования", "red")
        ui_wait_enter(f"Откройте '{zone_title}' вручную и нажмите Enter")

    # 2. Кликаем "Далее" чтобы перейти на страницу поиска городов
    page.wait_for_timeout(1_500)
    if not click_next_button(page):
        ui_status("", "Кнопка 'Далее' не найдена, пробуем подождать...", "yellow")
        page.wait_for_timeout(3_000)
        click_next_button(page)

    # 3. Ждём страницу с поиском
    if not wait_for_search_page(page):
        ui_status("", "Страница поиска не загрузилась", "red")
        ui_wait_enter("Перейдите на страницу поиска вручную и нажмите Enter")

    # 4. Ищем и отмечаем каждый город
    for i, city in enumerate(cities, 1):
        ui_status("", f"[{i}/{len(cities)}] {city}", "cyan")

        status = search_and_select_city(page, city)

        if status == "checked":
            ui_status("", f"[{i}/{len(cities)}] {city} -- отмечен", "green")
            report["checked"].append(city)
        elif status == "already":
            ui_status("", f"[{i}/{len(cities)}] {city} -- уже отмечен", "grey")
            report["already"].append(city)
        elif status == "not_found":
            ui_status("", f"[{i}/{len(cities)}] {city} -- НЕ НАЙДЕН", "yellow")
            report["not_found"].append(city)
        else:
            ui_status("", f"[{i}/{len(cities)}] {city} -- ОШИБКА", "red")
            report["failed"].append(city)

    # 5. Сохраняем: "Далее" (2 раза) + "Сохранить"
    ui_status("", "Сохраняем зону...", "cyan")
    click_next_button(page)
    page.wait_for_timeout(1_000)
    click_next_button(page)
    page.wait_for_timeout(1_000)

    if not click_save_button(page):
        ui_status("", "Кнопка 'Сохранить' не найдена, пробуем 'Далее'...", "yellow")
        # Может быть ещё один "Далее" перед "Сохранить"
        click_next_button(page)
        page.wait_for_timeout(1_000)
        click_save_button(page)

    # Ждём возврата на страницу складов
    page.wait_for_timeout(AFTER_SAVE_MS)

    # 6. Отчёт
    print_zone_report(report)

    return report


def print_zone_report(report: Dict):
    """Печатает отчёт по зоне."""
    ui_sep(f"ОТЧЁТ: {report['zone']}", "blue")
    ui_status("", f"Всего городов: {report['total']}", "white")
    ui_status("", f"Отмечено сейчас: {len(report['checked'])}", "green")
    ui_status("", f"Уже были отмечены: {len(report['already'])}", "grey")
    ui_status("", f"Не найдено: {len(report['not_found'])}", "yellow")
    if report["failed"]:
        ui_status("", f"Ошибки: {len(report['failed'])}", "red")

    if report["not_found"]:
        for c in report["not_found"][:15]:
            print(_c("yellow", f"    {c}"))
        if len(report["not_found"]) > 15:
            print(_c("yellow", f"    ... и ещё {len(report['not_found']) - 15}"))

    found_total = len(report["checked"]) + len(report["already"])
    ui_status("", f"{report['zone']}: найдено {found_total}, не найдено {len(report['not_found'])}", "cyan")


# ══════════════════════════════════════════════
#  МАППИНГ ЗОН — сопоставление Excel и страницы
# ══════════════════════════════════════════════
def match_zones_to_page(page_zones: List[Dict], excel_zones: Dict[str, List[str]],
                        warehouse_name: str) -> Dict[str, str]:
    """
    Сопоставляет зоны из Excel (напр. "Зона 1") с зонами на странице (напр. "Пенза Зона 1").
    Возвращает: {excel_zone_name: page_zone_title}
    """
    mapping = {}
    page_zone_titles = [z["title"] for z in page_zones]

    ui_status("", f"Сопоставляем зоны Excel с зонами на странице...", "cyan")
    logger.info(f"Page zones: {page_zone_titles}")
    logger.info(f"Excel zones: {list(excel_zones.keys())}")

    for excel_zone in excel_zones.keys():
        # Извлекаем номер зоны
        zone_match = re.search(r'(\d+)', excel_zone)
        zone_num = zone_match.group(1) if zone_match else ""

        matched = None
        for pz_title in page_zone_titles:
            if pz_title in mapping.values():
                continue  # Уже сопоставлена

            # Точное совпадение
            if excel_zone.lower() == pz_title.lower():
                matched = pz_title
                break

            # По номеру зоны: "Зона 1" ↔ "Пенза Зона 1"
            if zone_num:
                pattern = rf'зона\s*\.?\s*{zone_num}(?!\d)'
                if re.search(pattern, pz_title, re.IGNORECASE):
                    matched = pz_title
                    break

        if matched:
            mapping[excel_zone] = matched
            ui_status("", f"  {excel_zone} -> {matched}", "green")
            logger.info(f"Matched: {excel_zone} -> {matched}")
        else:
            ui_status("", f"  {excel_zone} -> НЕ НАЙДЕНА на странице!", "red")
            logger.warning(f"No match for: {excel_zone}")

    return mapping


def prompt_zone_mapping(excel_zone: str, page_zone_titles: List[str]) -> Optional[str]:
    """Ручной выбор зоны пользователем."""
    print()
    ui_status("", f"Зона '{excel_zone}' не найдена автоматически", "yellow")
    print(_c("white", "  Доступные зоны на странице:"))
    available = [t for t in page_zone_titles]
    for i, t in enumerate(available, 1):
        print(_c("white", f"    {i}. {t}"))
    print(_c("white", f"    0. Пропустить"))
    print()

    while True:
        try:
            raw = input(_c("yellow", f" Какая зона соответствует '{excel_zone}'? ")).strip()
            if raw == "0":
                return None
            if raw.isdigit() and 1 <= int(raw) <= len(available):
                return available[int(raw) - 1]
        except (KeyboardInterrupt, EOFError):
            return None


# ══════════════════════════════════════════════
#  ГЛАВНЫЙ ПРОЦЕСС
# ══════════════════════════════════════════════
def run(cities_data: Dict[str, List[str]], warehouse_filter: str = ""):
    """
    Основной процесс:
    1. Открыть страницу складов
    2. Найти нужный склад
    3. Раскрыть скрытые методы
    4. Для каждой зоны: открыть редактирование, найти города, сохранить
    """
    pw, ctx, page = launch_browser()

    try:
        # 1. Открываем страницу складов
        ui_status("", f"Открываем: {OZON_WAREHOUSE_URL}", "cyan")
        page.goto(OZON_WAREHOUSE_URL, wait_until="domcontentloaded", timeout=30_000)
        page.wait_for_timeout(3_000)

        # 2. Проверяем авторизацию
        if not check_auth(page):
            ui_status("", "Требуется авторизация!", "yellow")
            ui_wait_enter("Авторизуйтесь в браузере и нажмите Enter")
            page.goto(OZON_WAREHOUSE_URL, wait_until="domcontentloaded", timeout=30_000)
            page.wait_for_timeout(3_000)
            if not check_auth(page):
                ui_status("", "Авторизация не удалась!", "red")
                return

        ui_status("", "Авторизация подтверждена", "green")

        # 3. Сканируем склады
        warehouses = scan_warehouses(page)
        if not warehouses:
            ui_status("", "Склады не найдены на странице!", "red")
            return

        # 4. Выбираем склад
        ui_sep("ВЫБОР СКЛАДА", "yellow")
        if warehouse_filter:
            # Автоматический поиск
            matched_wh = None
            for wh in warehouses:
                if warehouse_filter.lower() in wh["name"].lower():
                    matched_wh = wh
                    break
            if not matched_wh:
                ui_status("", f"Склад '{warehouse_filter}' не найден!", "red")
                for wh in warehouses:
                    ui_status("", f"  Доступен: {wh['name']}", "grey")
                return
        else:
            for i, wh in enumerate(warehouses, 1):
                zones_str = f", зон: {wh['zonesCount']}" if wh["zonesCount"] else ""
                print(_c("white", f"  {i}. {wh['name']}{zones_str}"))
            print()

            while True:
                try:
                    raw = input(_c("yellow", " Выберите склад: ")).strip()
                    if raw.isdigit() and 1 <= int(raw) <= len(warehouses):
                        matched_wh = warehouses[int(raw) - 1]
                        break
                except (KeyboardInterrupt, EOFError):
                    return

        warehouse_name = matched_wh["name"]
        ui_status("", f"Выбран склад: {warehouse_name}", "green")

        # 5. Раскрываем скрытые методы
        expand_hidden_methods(page, warehouse_name)

        # 6. Пересканируем зоны после раскрытия
        page.wait_for_timeout(1_000)
        warehouses_updated = scan_warehouses(page)
        wh_updated = None
        for wh in warehouses_updated:
            if wh["name"] == warehouse_name:
                wh_updated = wh
                break

        if not wh_updated:
            ui_status("", "Склад не найден после обновления!", "red")
            return

        page_zones = wh_updated.get("zones", [])
        ui_status("", f"Зон на странице: {len(page_zones)}", "cyan")
        for z in page_zones:
            ui_status("", f"  {z['title']}", "grey")

        # 7. Маппинг зон Excel -> страница
        zone_mapping = match_zones_to_page(page_zones, cities_data, warehouse_name)

        # Для незамапленных — спросить пользователя
        page_zone_titles = [z["title"] for z in page_zones]
        for excel_zone in cities_data.keys():
            if excel_zone not in zone_mapping:
                chosen = prompt_zone_mapping(excel_zone, page_zone_titles)
                if chosen:
                    zone_mapping[excel_zone] = chosen

        if not zone_mapping:
            ui_status("", "Ни одна зона не сопоставлена!", "red")
            return

        # 8. Обрабатываем каждую зону
        all_reports = []
        for excel_zone, page_zone_title in zone_mapping.items():
            cities = cities_data.get(excel_zone, [])
            if not cities:
                ui_status("", f"Зона {excel_zone}: пустая, пропускаем", "grey")
                continue

            report = process_zone(page, warehouse_name, page_zone_title, cities)
            all_reports.append(report)

            # После сохранения — возвращаемся на страницу складов
            page.wait_for_timeout(2_000)
            current_url = page.url
            if "warehouse" not in current_url:
                ui_status("", "Возвращаемся на страницу складов...", "cyan")
                page.goto(OZON_WAREHOUSE_URL, wait_until="domcontentloaded", timeout=30_000)
                page.wait_for_timeout(3_000)

            # Пересканируем и раскрываем для следующей зоны
            expand_hidden_methods(page, warehouse_name)
            page.wait_for_timeout(1_000)

        # 9. Итоговый отчёт
        ui_sep("ИТОГОВЫЙ ОТЧЁТ", "magenta")
        total_checked = sum(len(r["checked"]) for r in all_reports)
        total_already = sum(len(r["already"]) for r in all_reports)
        total_not_found = sum(len(r["not_found"]) for r in all_reports)
        total_cities = sum(r["total"] for r in all_reports)

        ui_status("", f"Обработано зон: {len(all_reports)}", "white")
        ui_status("", f"Всего городов: {total_cities}", "white")
        ui_status("", f"Отмечено: {total_checked}", "green")
        ui_status("", f"Уже были: {total_already}", "grey")
        ui_status("", f"Не найдено: {total_not_found}", "yellow")
        ui_sep()

    except KeyboardInterrupt:
        ui_status("", "Прервано пользователем", "yellow")
    except Exception as e:
        ui_status("", f"Ошибка: {e}", "red")
        logger.exception("Fatal error")
    finally:
        try:
            ctx.close()
            pw.stop()
        except Exception:
            pass


# ══════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════
def main():
    print()
    ui_sep("OZON -- Настройка зон доставки", "magenta")
    print()

    # Выбор Excel
    excel_file = select_excel_file()
    cities_data = load_cities_data(excel_file)

    if not cities_data:
        ui_status("", "Нет данных о городах!", "red")
        sys.exit(1)

    ui_sep("ЗАГРУЖЕННЫЕ ЗОНЫ", "yellow")
    for zone_name, zone_cities in cities_data.items():
        print(_c("white", f"  {zone_name}: {len(zone_cities)} городов"))
    print()

    # Выбор зон
    cities_data = select_zones(cities_data)

    total = sum(len(c) for c in cities_data.values())
    ui_status("", f"Выбрано: {len(cities_data)} зон, {total} городов", "cyan")

    # Запуск
    run(cities_data)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        ui_status("", "Прервано", "yellow")
        sys.exit(0)
    except Exception as e:
        ui_status("", f"Ошибка: {e}", "red")
        logger.exception("Fatal error")
        sys.exit(1)
