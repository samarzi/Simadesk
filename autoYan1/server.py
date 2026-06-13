#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Локальный сервер автоматизации для SimaDesk.
Запускает Playwright-скрипты (kur.py, tariff.py, ozon_warehouse.py)
через API-эндпоинты и стримит логи через SSE.

Запуск: python server.py
Порт:   8899 (SimaDesk обращается к http://localhost:8899)
"""

import os
import sys
import json
import time
import uuid
import signal
import logging
import tempfile
import subprocess
import threading
from pathlib import Path
from datetime import datetime
from collections import deque
from typing import Dict, List, Optional

# ── HTTP сервер на stdlib (без зависимостей кроме playwright/openpyxl) ──
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
import io

# ══════════════════════════════════════════════
#  КОНФИГУРАЦИЯ
# ══════════════════════════════════════════════
PORT = 8899
BASE_DIR = Path(__file__).parent.resolve()
SCRIPTS_DIR_YA = BASE_DIR / "Яндекс скрипты"
SCRIPTS_DIR_OZ = BASE_DIR / "Озон скрипты"
CITIES_DIR = BASE_DIR / "Склады города"
WAREHOUSES_DIR = BASE_DIR / "Склады"
REPORTS_DIR = BASE_DIR / "Отчёты"

# Активные процессы: task_id → { process, logs, status, ... }
tasks: Dict[str, dict] = {}

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("AutoServer")


# ══════════════════════════════════════════════
#  УПРАВЛЕНИЕ ЗАДАЧАМИ
# ══════════════════════════════════════════════
def create_task(script_type: str, params: dict) -> str:
    """Создаёт задачу и запускает скрипт в фоне."""
    task_id = str(uuid.uuid4())[:8]

    # Определяем скрипт и рабочую директорию
    if script_type == "yandex-warehouses":
        script = str(SCRIPTS_DIR_YA / "kur.py")
        cwd = str(SCRIPTS_DIR_YA)
    elif script_type == "yandex-tariffs":
        script = str(SCRIPTS_DIR_YA / "tariff.py")
        cwd = str(SCRIPTS_DIR_YA)
    elif script_type == "ozon-warehouses":
        script = str(SCRIPTS_DIR_OZ / "ozon_warehouse.py")
        cwd = str(SCRIPTS_DIR_OZ)
    else:
        raise ValueError(f"Unknown script type: {script_type}")

    # Сохраняем города во временный Excel если переданы
    cities_file = None
    if params.get("cities"):
        cities_file = save_cities_to_excel(params["cities"], task_id)

    # Формируем конфиг для запуска
    config = {
        "YD_START_URL": params.get("warehouseUrl", ""),
        "YD_CABINET": params.get("cabinet", ""),
        "YD_WAREHOUSE": params.get("warehouse", ""),
        "YD_MODE": params.get("mode", ""),
        "YD_CITIES": params.get("cities", []),
        "YD_PROFILE_DIR": str(BASE_DIR / "profiles" / f"browser_profile_{task_id}"),
        "YD_INSTANCE_NAME": task_id,
    }

    config_path = str(BASE_DIR / f"config_{task_id}.json")
    with open(config_path, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

    # Запускаем процесс
    env = os.environ.copy()
    env["YD_PROFILE_DIR"] = config["YD_PROFILE_DIR"]
    env["YD_INSTANCE_NAME"] = task_id
    env["PYTHONUNBUFFERED"] = "1"

    cmd = [sys.executable, script, "--instance", config_path]

    log_buffer = deque(maxlen=5000)

    task = {
        "id": task_id,
        "type": script_type,
        "params": params,
        "status": "running",
        "started_at": datetime.now().isoformat(),
        "finished_at": None,
        "logs": log_buffer,
        "process": None,
        "config_path": config_path,
        "cities_file": cities_file,
    }

    tasks[task_id] = task

    # Запускаем в фоновом потоке
    thread = threading.Thread(target=_run_process, args=(task_id, cmd, cwd, env), daemon=True)
    thread.start()

    logger.info(f"Task {task_id} started: {script_type}")
    return task_id


def _run_process(task_id: str, cmd: list, cwd: str, env: dict):
    """Запускает процесс и читает stdout/stderr в буфер логов."""
    task = tasks.get(task_id)
    if not task:
        return

    try:
        proc = subprocess.Popen(
            cmd,
            cwd=cwd,
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            bufsize=1,
            universal_newlines=True,
        )
        task["process"] = proc

        # Читаем вывод построчно
        for line in proc.stdout:
            line = line.rstrip("\n")
            # Убираем ANSI-коды для веб-отображения
            import re
            clean = re.sub(r'\033\[[^m]*m', '', line)
            timestamp = datetime.now().strftime("%H:%M:%S")
            task["logs"].append(f"[{timestamp}] {clean}")

        proc.wait()
        task["status"] = "completed" if proc.returncode == 0 else "error"
        task["finished_at"] = datetime.now().isoformat()

        exit_msg = f"Процесс завершён (код: {proc.returncode})"
        task["logs"].append(f"[{datetime.now().strftime('%H:%M:%S')}] {'✅' if proc.returncode == 0 else '❌'} {exit_msg}")

    except Exception as e:
        task["status"] = "error"
        task["finished_at"] = datetime.now().isoformat()
        task["logs"].append(f"[{datetime.now().strftime('%H:%M:%S')}] ❌ Ошибка: {e}")
        logger.error(f"Task {task_id} error: {e}")

    # Очистка временных файлов
    config_path = task.get("config_path")
    if config_path and Path(config_path).exists():
        try:
            Path(config_path).unlink()
        except Exception:
            pass


def stop_task(task_id: str) -> bool:
    """Останавливает задачу."""
    task = tasks.get(task_id)
    if not task:
        return False
    proc = task.get("process")
    if proc and proc.poll() is None:
        proc.terminate()
        try:
            proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            proc.kill()
        task["status"] = "stopped"
        task["finished_at"] = datetime.now().isoformat()
        task["logs"].append(f"[{datetime.now().strftime('%H:%M:%S')}] ⏹ Остановлено пользователем")
        return True
    return False


def save_cities_to_excel(cities: list, task_id: str) -> str:
    """Сохраняет список городов в Excel для скрипта."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1, "Город")
    for i, city in enumerate(cities, 2):
        ws.cell(i, 1, city.strip())
    filepath = str(WAREHOUSES_DIR / f"task_{task_id}.xlsx")
    WAREHOUSES_DIR.mkdir(exist_ok=True)
    wb.save(filepath)
    return filepath


# ══════════════════════════════════════════════
#  HTTP ОБРАБОТЧИК
# ══════════════════════════════════════════════
class AutoHandler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        """CORS preflight."""
        self.send_response(200)
        self._cors()
        self.send_header("Content-Length", "0")
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        qs = parse_qs(parsed.query)

        if path == "/api/health":
            self._json_response({"status": "ok", "version": "1.0"})

        elif path == "/api/tasks":
            result = []
            for t in tasks.values():
                result.append({
                    "id": t["id"],
                    "type": t["type"],
                    "status": t["status"],
                    "started_at": t["started_at"],
                    "finished_at": t["finished_at"],
                })
            self._json_response(result)

        elif path.startswith("/api/tasks/") and path.endswith("/logs"):
            task_id = path.split("/")[3]
            task = tasks.get(task_id)
            if not task:
                self._json_response({"error": "Task not found"}, 404)
                return
            since = int(qs.get("since", ["0"])[0])
            logs = list(task["logs"])
            self._json_response({
                "id": task_id,
                "status": task["status"],
                "logs": logs[since:],
                "total": len(logs),
            })

        elif path == "/api/template/cities":
            # Генерируем шаблон Excel для городов
            self._send_template_cities()

        elif path == "/api/template/zones":
            # Генерируем шаблон Excel для зон
            self._send_template_zones()

        elif path == "/api/reports":
            # Список отчётов
            reports = []
            if REPORTS_DIR.exists():
                for f in sorted(REPORTS_DIR.glob("*.xlsx"), reverse=True):
                    reports.append({
                        "name": f.name,
                        "size": f.stat().st_size,
                        "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
                    })
            self._json_response(reports)

        else:
            self._json_response({"error": "Not found"}, 404)

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path

        body = self._read_body()

        if path == "/api/tasks/start":
            try:
                data = json.loads(body) if body else {}
                task_id = create_task(
                    script_type=data.get("type", ""),
                    params=data.get("params", {}),
                )
                self._json_response({"task_id": task_id, "status": "started"})
            except Exception as e:
                self._json_response({"error": str(e)}, 400)

        elif path.startswith("/api/tasks/") and path.endswith("/stop"):
            task_id = path.split("/")[3]
            success = stop_task(task_id)
            self._json_response({"stopped": success})

        else:
            self._json_response({"error": "Not found"}, 404)

    # ── Helpers ──

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def _json_response(self, data, code=200):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(code)
        self._cors()
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _read_body(self) -> str:
        length = int(self.headers.get("Content-Length", 0))
        if length > 0:
            return self.rfile.read(length).decode("utf-8")
        return ""

    def _send_template_cities(self):
        """Отправляет шаблон Excel для списка городов."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Города"

        # Заголовок
        ws.cell(1, 1, "Город").font = Font(bold=True, size=12)
        ws.column_dimensions["A"].width = 30

        # Примеры
        examples = ["Москва", "Санкт-Петербург", "Нижний Новгород", "Пенза", "Смоленск"]
        for i, city in enumerate(examples, 2):
            ws.cell(i, 1, city).font = Font(color="999999")

        # Инструкция
        ws.cell(8, 1, "Инструкция:").font = Font(bold=True, color="FF6600")
        ws.cell(9, 1, "Столбец A — по одному городу в каждой строке")
        ws.cell(10, 1, "Строка 1 — заголовок (пропускается)")
        ws.cell(11, 1, "Удалите примеры и вставьте свои города")

        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()

        self.send_response(200)
        self._cors()
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", "attachment; filename=shablon_goroda.xlsx")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _send_template_zones(self):
        """Отправляет шаблон Excel для зон (используется в тарифах и Ozon)."""
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Зоны"

        # Заголовки зон
        zones = ["Зона 1", "Зона 2", "Зона 3"]
        for i, z in enumerate(zones, 1):
            ws.cell(1, i, z).font = Font(bold=True, size=12)
            ws.column_dimensions[ws.cell(1, i).column_letter].width = 25

        # Примеры
        examples = {
            1: ["Москва", "Подольск", "Мытищи"],
            2: ["Тула", "Рязань", "Калуга"],
            3: ["Нижний Новгород", "Казань"],
        }
        for col, cities in examples.items():
            for row, city in enumerate(cities, 2):
                ws.cell(row, col, city).font = Font(color="999999")

        ws.cell(7, 1, "Колонки = зоны, строки = города").font = Font(bold=True, color="FF6600")

        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()

        self.send_response(200)
        self._cors()
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", "attachment; filename=shablon_zony.xlsx")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, format, *args):
        """Тише логи HTTP."""
        if "/api/tasks/" in str(args[0]) and "/logs" in str(args[0]):
            return  # Не логируем polling
        logger.info(f"HTTP {args[0]}")


# ══════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════
def main():
    # Создаём нужные папки
    for d in [CITIES_DIR, WAREHOUSES_DIR, REPORTS_DIR]:
        d.mkdir(exist_ok=True)

    server = HTTPServer(("0.0.0.0", PORT), AutoHandler)

    print()
    print(f"  ╔══════════════════════════════════════════════╗")
    print(f"  ║   SimaDesk Automation Server v1.0            ║")
    print(f"  ║   http://localhost:{PORT}                      ║")
    print(f"  ║                                              ║")
    print(f"  ║   Скрипты:                                   ║")
    print(f"  ║   • Яндекс Склады  (kur.py)                  ║")
    print(f"  ║   • Яндекс Тарифы  (tariff.py)               ║")
    print(f"  ║   • Озон Склады    (ozon_warehouse.py)        ║")
    print(f"  ║                                              ║")
    print(f"  ║   Нажмите Ctrl+C для остановки               ║")
    print(f"  ╚══════════════════════════════════════════════╝")
    print()

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  Сервер остановлен")
        # Останавливаем все активные задачи
        for task_id in list(tasks.keys()):
            stop_task(task_id)
        server.server_close()


if __name__ == "__main__":
    main()
